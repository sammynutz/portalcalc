import sys
import math
import os
import re
import json
import copy
import html
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, replace

# The envelope solver can run load combinations in parallel. Keep BLAS from
# creating another large thread pool inside each combination worker.
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("OMP_NUM_THREADS", "1")

import numpy as np
from openpyxl import load_workbook

from PySide6.QtCore import Qt, QTimer

from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QFormLayout,
    QLineEdit,
    QPushButton,
    QLabel,
    QTextEdit,
    QComboBox,
    QMessageBox,
    QTabWidget,
    QVBoxLayout,
    QGridLayout,
    QButtonGroup,
    QRadioButton,
    QFileDialog,
    QToolButton,
    QCheckBox,
    QGroupBox,
    QScrollArea,
)
from PySide6.QtGui import QFontMetrics, QPainter
from PySide6.QtCore import QRectF

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from loads import BasicLoadCase, BasicRoofLoads, RoofLoadPlacement, WallLoads, apply_roof_vertical_loads, apply_structure_self_weight, build_load_combinations
from loads import apply_roof_vertical_loads_in_zones, apply_wall_horizontal_loads
from member_checks import MemberActions, SectionCheckProperties, check_member
from wind import (
    WindInputs,
    calculate_wind,
    apply_roof_wind_case_to_top_nodes,
    apply_column_wind_cpe_loads,
    apply_roof_cpi_case_to_top_nodes,
    apply_roof_cpi_to_top_nodes,
    apply_column_wind_cpi_loads,
    CPE_CASE_OPTIONS,
    CPI_CASE_OPTIONS,
    WIND_CASE_OPTIONS,
    active_cpi_for_case,
    effective_cpi_case_for_wind,
    action_combination_factors,
    external_pressure_factor,
    internal_pressure_factor,
    wind_pressure_scale,
    adjacent_wall_cpe_for_side,
    wall_surface_for_cpe_case,
    wall_cpe_for_surface,
    opening_area_reduction_factor,
    free_roof_cpn,
    local_pressure_dimension_a,
    roof_local_pressure_factor,
    wall_cpe_line_load_kn_m,
    wall_cpi_line_load_kn_m,
)


SNOW_REGION_K1 = {
    "Region AN": 0.2,
    "Region AC": 0.7,
    "Region AS": 1.0,
    "Region AT": 1.6,
}

COSTING_DEFAULTS = {
    "shs_rate_per_kg": "2.07",
    "ub_rate_per_kg": "2.07",
    "wb_rate_per_kg": "2.80",
    "purlin_rate_per_kg": "1.90",
}

UB_STOCK_LENGTHS_M = [4.5, 5.0, 5.5, 6.0, 6.75, 7.5, 8.25, 9.0, 10.5, 12.0, 13.5, 15.0, 16.5, 18.0, 20.0]
STANDARD_PIER_DIAMETERS_MM = [600, 750, 900, 1200]


# ==========================================================
# RESOURCE PATH
# ==========================================================

def resource_path(relative_path):
    """Works both from source and from a PyInstaller bundle."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)


# ==========================================================
# PROFILE CATALOG
# ==========================================================

@dataclass
class SectionProfile:
    name: str
    group_name: str
    shape_type: str
    A: float       # mm^2
    J: float       # mm^4
    Iyp: float     # mm^4
    Izp: float     # mm^4
    Iw: float      # mm^6
    Syp: float = 0.0
    Szp: float = 0.0
    fabrication: str = ""
    fy: float = 300.0
    fyw: float = 300.0
    fu: float = 440.0
    depth: float = 0.0
    flange_width: float = 0.0
    flange_thickness: float = 0.0
    web_thickness: float = 0.0

    @property
    def I_analysis(self):
        """
        Temporary 2D analysis inertia assumption.
        For a final design tool, this should become a selectable strong/weak-axis setting.
        """
        return max(self.Iyp, self.Izp)

    @property
    def display_name(self):
        return self.name

    def check_properties(self):
        return SectionCheckProperties(
            name=self.name,
            shape_type=self.shape_type,
            fabrication=self.fabrication,
            area_mm2=self.A,
            j_mm4=self.J,
            iy_mm4=self.Iyp,
            iz_mm4=self.Izp,
            iw_mm6=self.Iw,
            sy_mm3=self.Syp or None,
            sz_mm3=self.Szp or None,
            fy_mpa=self.fy,
            fyw_mpa=self.fyw,
            fu_mpa=self.fu,
            depth_mm=self.depth or None,
            flange_width_mm=self.flange_width or None,
            flange_thickness_mm=self.flange_thickness or None,
            web_thickness_mm=self.web_thickness or None,
        )


@dataclass
class PurlinLayout:
    points: list
    max_spacing_mm: float
    end_spacing_limit_mm: float
    mid_spacing_limit_mm: float


class ShapeCatalog:

    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.profiles = []
        self.profile_by_name = {}
        self.load()

    def load(self):
        wb = load_workbook(self.xlsx_path, data_only=True)
        ws = wb.active

        header_row = None
        headers = {}

        for row in ws.iter_rows(min_row=1, max_row=20):
            values = [cell.value for cell in row]
            if "Name" in values and "A" in values and "Iyp" in values and "Izp" in values:
                header_row = row[0].row
                headers = {str(cell.value).strip(): cell.column for cell in row if cell.value is not None}
                break

        if header_row is None:
            raise ValueError("Could not find catalog header row containing Name, A, Iyp and Izp.")

        name_col = headers.get("Name")
        group_col = headers.get("Group Name")
        shape_type_col = headers.get("Shape Type")
        area_col = headers.get("A")
        j_col = headers.get("J")
        iyp_col = headers.get("Iyp")
        izp_col = headers.get("Izp")
        iw_col = headers.get("Iw")
        syp_col = headers.get("Syp")
        szp_col = headers.get("Szp")
        fabrication_col = headers.get("Fabrication")
        depth_col = headers.get("D")
        bt_col = headers.get("Bt")
        bb_col = headers.get("Bb")
        tt_col = headers.get("Tt")
        tb_col = headers.get("Tb")
        tw_col = headers.get("Tw")

        def header_col(*tokens):
            for label, column in headers.items():
                normalized = str(label).replace("_x000D_", " ").replace("\n", " ").lower()
                if all(token.lower() in normalized for token in tokens):
                    return column
            return None

        fy_col = header_col("fy", "normal")
        fyw_col = header_col("fyw", "normal")
        fu_col = header_col("fu", "normal")

        def cell_float(row_number, column, default=0.0):
            if column is None:
                return default
            value = ws.cell(row=row_number, column=column).value
            if value is None or value == "":
                return default
            return float(value)

        def cell_text(row_number, column, default=""):
            if column is None:
                return default
            value = ws.cell(row=row_number, column=column).value
            return str(value or default).strip()

        for row in range(header_row + 2, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            if not name:
                continue

            A = ws.cell(row=row, column=area_col).value
            Iyp = ws.cell(row=row, column=iyp_col).value
            Izp = ws.cell(row=row, column=izp_col).value

            if A is None or Iyp is None or Izp is None:
                continue

            try:
                profile = SectionProfile(
                    name=str(name).strip(),
                    group_name=str(ws.cell(row=row, column=group_col).value or "").strip(),
                    shape_type=str(ws.cell(row=row, column=shape_type_col).value or "").strip(),
                    A=float(A),
                    J=cell_float(row, j_col),
                    Iyp=float(Iyp),
                    Izp=float(Izp),
                    Iw=cell_float(row, iw_col),
                    Syp=cell_float(row, syp_col),
                    Szp=cell_float(row, szp_col),
                    fabrication=cell_text(row, fabrication_col),
                    fy=cell_float(row, fy_col, 300.0),
                    fyw=cell_float(row, fyw_col) or cell_float(row, fy_col, 300.0),
                    fu=cell_float(row, fu_col, 440.0),
                    depth=cell_float(row, depth_col),
                    flange_width=max(cell_float(row, bt_col), cell_float(row, bb_col)),
                    flange_thickness=max(cell_float(row, tt_col), cell_float(row, tb_col)),
                    web_thickness=cell_float(row, tw_col),
                )
            except (TypeError, ValueError):
                continue

            self.profiles.append(profile)
            self.profile_by_name[profile.name] = profile

        self.profiles.sort(key=self.profile_sort_key)

    def profile_sort_key(self, profile):
        text = f"{profile.name} {profile.group_name} {profile.shape_type}".upper()
        family_order = [
            ("PURLIN", 0),
            ("ZED", 0),
            ("CEE", 0),
            ("UB", 0),
            ("UC", 1),
            ("PFC", 2),
            ("RHS", 3),
            ("SHS", 4),
            ("CHS", 5),
        ]
        family_rank = 99
        family_name = ""
        for token, rank in family_order:
            if re.search(rf"(^|[^A-Z]){token}([^A-Z]|$)", text):
                family_rank = rank
                family_name = token
                break
        numbers = [float(value) for value in re.findall(r"\d+(?:\.\d+)?", profile.name)]
        if family_name in {"SHS", "RHS", "CHS"}:
            # Tube names are typically depth * thickness; sort by depth first.
            numeric_key = numbers
        else:
            # UB/UC/PFC names are typically depth then mass; this still gives
            # smallest-to-largest behaviour within each family.
            numeric_key = numbers
        return (family_rank, numeric_key, profile.name)

    def all_profiles(self):
        return self.profiles

    def shs_profiles(self):
        shs = []
        for p in self.profiles:
            text = f"{p.name} {p.group_name} {p.shape_type}".upper()
            if "SHS" in text or "SQUARE HOLLOW" in text:
                shs.append(p)
        return shs

    def ub_profiles(self):
        ub = []
        for p in self.profiles:
            text = f"{p.name} {p.group_name} {p.shape_type}".upper()
            if re.search(r"(^|[^A-Z])UB([^A-Z]|$)", text):
                ub.append(p)
        return ub

    def purlin_profiles(self):
        purlins = []
        for p in self.profiles:
            text = f"{p.name} {p.group_name} {p.shape_type}".upper()
            if any(token in text for token in ["PURLIN", "ZED", "CEE", "Z PURLIN", "C PURLIN"]):
                purlins.append(p)
        return purlins

    def get(self, name):
        return self.profile_by_name[name]


# ==========================================================
# NODE
# ==========================================================

class Node:

    def __init__(self, node_id, x, y, ux=None, uy=None, rz=None):
        self.id = node_id
        self.x = x
        self.y = y

        self.native_ux = 3 * (node_id - 1)
        self.native_uy = 3 * (node_id - 1) + 1
        self.native_rz = 3 * (node_id - 1) + 2
        self.ux = self.native_ux if ux is None else ux
        self.uy = self.native_uy if uy is None else uy
        self.rz = self.native_rz if rz is None else rz

    @property
    def unused_native_dofs(self):
        assigned = {self.ux, self.uy, self.rz}
        return [dof for dof in [self.native_ux, self.native_uy, self.native_rz] if dof not in assigned]


# ==========================================================
# 2D FRAME ELEMENT
# ==========================================================

class FrameElement2D:

    def __init__(self, element_id, start_node, end_node, E, profile, group="FRAME", analysis_i_factor=1.0):
        self.id = element_id
        self.start = start_node
        self.end = end_node
        self.E = E
        self.profile = profile
        self.A = profile.A
        self.analysis_i_factor = max(float(analysis_i_factor), 1e-9)
        self.I = profile.I_analysis * self.analysis_i_factor
        self.group = group

    def length(self):
        dx = self.end.x - self.start.x
        dy = self.end.y - self.start.y
        return math.sqrt(dx**2 + dy**2)

    def angle(self):
        dx = self.end.x - self.start.x
        dy = self.end.y - self.start.y
        return math.atan2(dy, dx)

    def transformation_matrix(self):
        theta = self.angle()
        c = math.cos(theta)
        s = math.sin(theta)

        return np.array([
            [ c, s, 0, 0, 0, 0],
            [-s, c, 0, 0, 0, 0],
            [ 0, 0, 1, 0, 0, 0],
            [ 0, 0, 0, c, s, 0],
            [ 0, 0, 0,-s, c, 0],
            [ 0, 0, 0, 0, 0, 1]
        ])

    def local_stiffness(self):
        L = self.length()
        E = self.E
        A = self.A
        I = self.I

        return np.array([
            [ A*E/L,           0,              0, -A*E/L,           0,              0],
            [ 0,      12*E*I/L**3,   6*E*I/L**2,       0, -12*E*I/L**3,   6*E*I/L**2],
            [ 0,       6*E*I/L**2,     4*E*I/L,       0,  -6*E*I/L**2,     2*E*I/L],
            [-A*E/L,           0,              0,  A*E/L,           0,              0],
            [ 0,     -12*E*I/L**3,  -6*E*I/L**2,       0,  12*E*I/L**3,  -6*E*I/L**2],
            [ 0,       6*E*I/L**2,     2*E*I/L,       0,  -6*E*I/L**2,     4*E*I/L]
        ])

    def global_stiffness(self):
        T = self.transformation_matrix()
        return T.T @ self.local_stiffness() @ T

    def local_geometric_stiffness(self, axial_tension_n):
        L = self.length()
        if L <= 1e-9 or abs(axial_tension_n) <= 1e-9:
            return np.zeros((6, 6))
        n = axial_tension_n / (30.0 * L)
        return n * np.array([
            [0, 0, 0, 0, 0, 0],
            [0, 36, 3 * L, 0, -36, 3 * L],
            [0, 3 * L, 4 * L**2, 0, -3 * L, -L**2],
            [0, 0, 0, 0, 0, 0],
            [0, -36, -3 * L, 0, 36, -3 * L],
            [0, 3 * L, -L**2, 0, -3 * L, 4 * L**2],
        ])

    def global_geometric_stiffness(self, axial_tension_n):
        T = self.transformation_matrix()
        return T.T @ self.local_geometric_stiffness(axial_tension_n) @ T

    def dof_indices(self):
        return [
            self.start.ux,
            self.start.uy,
            self.start.rz,
            self.end.ux,
            self.end.uy,
            self.end.rz
        ]

    def local_displacements(self, global_displacements):
        dofs = self.dof_indices()
        d_global = global_displacements[dofs]
        return self.transformation_matrix() @ d_global

    def local_end_forces(self, global_displacements):
        return self.local_stiffness() @ self.local_displacements(global_displacements)

    def axial_tension_force(self, global_displacements):
        f = self.local_end_forces(global_displacements)
        return (f[3] - f[0]) / 2.0

    def force_summary(self, global_displacements):
        f = self.local_end_forces(global_displacements)

        # Positive internal axial force is tension, negative is compression.
        axial_n = self.axial_tension_force(global_displacements)
        compression_kn = max(-axial_n, 0.0) / 1000
        tension_kn = max(axial_n, 0.0) / 1000

        shear_kn = max(abs(f[1]), abs(f[4])) / 1000
        moment_knm = max(abs(f[2]), abs(f[5])) / 1_000_000

        return {
            "compression_kn": compression_kn,
            "tension_kn": tension_kn,
            "shear_kn": shear_kn,
            "moment_knm": moment_knm,
            "end_forces": f
    }


# ==========================================================
# STRUCTURE
# ==========================================================

class Structure2D:

    def __init__(self):
        self.nodes = []
        self.elements = []
        self.loads = {}
        self.springs = {}
        self.restrained_dofs = []

    def add_node(self, node):
        self.nodes.append(node)
        for dof in node.unused_native_dofs:
            self.add_support(dof)

    def add_element(self, element):
        self.elements.append(element)

    def add_support(self, dof):
        if dof not in self.restrained_dofs:
            self.restrained_dofs.append(dof)

    def add_load(self, dof, value):
        self.loads[dof] = self.loads.get(dof, 0.0) + value

    def add_spring(self, dof, stiffness_n_per_mm):
        if stiffness_n_per_mm > 0.0:
            self.springs[dof] = self.springs.get(dof, 0.0) + stiffness_n_per_mm

    def elastic_stiffness_matrix(self):
        total_dofs = len(self.nodes) * 3
        K = np.zeros((total_dofs, total_dofs))
        for element in self.elements:
            k = element.global_stiffness()
            dofs = element.dof_indices()
            for i in range(6):
                for j in range(6):
                    K[dofs[i], dofs[j]] += k[i, j]
        for dof, stiffness in self.springs.items():
            K[dof, dof] += stiffness
        return K

    def load_vector(self):
        total_dofs = len(self.nodes) * 3
        F = np.zeros(total_dofs)
        for dof, value in self.loads.items():
            F[dof] += value
        return F

    def solve_with_stiffness(self, K, F):
        total_dofs = len(self.nodes) * 3
        free_dofs = [i for i in range(total_dofs) if i not in self.restrained_dofs]
        Kff = K[np.ix_(free_dofs, free_dofs)]
        Ff = F[free_dofs]

        df = np.linalg.solve(Kff, Ff)
        d = np.zeros(total_dofs)
        d[free_dofs] = df
        return d

    def solve(self):
        K = self.elastic_stiffness_matrix()
        F = self.load_vector()
        d = self.solve_with_stiffness(K, F)
        R = K @ d - F
        return d, R, F

    def geometric_stiffness_matrix(self, displacements):
        total_dofs = len(self.nodes) * 3
        K = np.zeros((total_dofs, total_dofs))
        for element in self.elements:
            axial_tension_n = element.axial_tension_force(displacements)
            kg = element.global_geometric_stiffness(axial_tension_n)
            dofs = element.dof_indices()
            for i in range(6):
                for j in range(6):
                    K[dofs[i], dofs[j]] += kg[i, j]
        return K

    def solve_second_order(self, max_iterations=20, tolerance=1e-4):
        elastic_k = self.elastic_stiffness_matrix()
        F = self.load_vector()
        first_order_d = self.solve_with_stiffness(elastic_k, F)
        d = first_order_d.copy()
        converged = False
        iterations = 0

        for iteration in range(1, max_iterations + 1):
            tangent_k = elastic_k + self.geometric_stiffness_matrix(d)
            next_d = self.solve_with_stiffness(tangent_k, F)
            change = np.max(np.abs(next_d - d))
            scale = max(np.max(np.abs(next_d)), 1.0)
            iterations = iteration
            d = next_d
            if change / scale <= tolerance:
                converged = True
                break

        R = elastic_k @ d - F
        first_max = max(float(np.max(np.abs(first_order_d))), 0.0)
        second_max = max(float(np.max(np.abs(d))), 0.0)
        amplification = second_max / first_max if first_max > 1e-12 else 1.0
        info = {
            "type": "second_order",
            "converged": converged,
            "iterations": iterations,
            "max_iterations": max_iterations,
            "tolerance": tolerance,
            "first_order_max_abs_mm": first_max,
            "second_order_max_abs_mm": second_max,
            "displacement_amplification": amplification,
        }
        return d, R, F, info


# ==========================================================
# COMPASS SELECTOR
# ==========================================================

class CompassSelector(QWidget):
    """Eight-point north selector, read as viewed from the shed elevation."""

    DIRECTIONS = [
        ("NW", "↖", 0, 0), ("N", "↑", 0, 2), ("NE", "↗", 0, 4),
        ("W",  "←", 2, 0),                         ("E",  "→", 2, 4),
        ("SW", "↙", 4, 0), ("S", "↓", 4, 2), ("SE", "↘", 4, 4),
    ]

    def __init__(self, default="E"):
        super().__init__()
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

        note = QLabel("Select where NORTH is when looking at the shed elevation")
        note.setWordWrap(True)
        layout.addWidget(note)

        grid = QGridLayout()
        grid.setSpacing(4)
        self.button_group = QButtonGroup(self)

        centre = QLabel("NORTH")
        centre.setAlignment(Qt.AlignCenter)
        centre.setStyleSheet("font-weight: bold; padding: 8px; border: 1px solid #999; border-radius: 6px;")
        grid.addWidget(centre, 2, 2)

        for direction, arrow, row, col in self.DIRECTIONS:
            button = QRadioButton(arrow)
            button.setProperty("direction", direction)
            self.button_group.addButton(button)
            grid.addWidget(button, row, col)
            if direction == default:
                button.setChecked(True)

        # Light arrow guides pointing away from the NORTH reference label.
        guide_labels = [
            ("↖", 1, 1), ("↑", 1, 2), ("↗", 1, 3),
            ("←", 2, 1),             ("→", 2, 3),
            ("↙", 3, 1), ("↓", 3, 2), ("↘", 3, 3),
        ]
        for text, row, col in guide_labels:
            label = QLabel(text)
            label.setAlignment(Qt.AlignCenter)
            grid.addWidget(label, row, col)

        layout.addLayout(grid)

    def currentText(self):
        checked = self.button_group.checkedButton()
        if checked is None:
            return "E"
        return checked.property("direction") or "E"

    def setCurrentText(self, value):
        for button in self.button_group.buttons():
            if button.property("direction") == value:
                button.setChecked(True)
                return


class EnvelopePlanSelector(QWidget):
    WALLS = [
        ("back", "Back"),
        ("left", "Left"),
        ("right", "Right"),
        ("front", "Front"),
    ]

    def __init__(self, labels=None, default_states=None, centre_text="PLAN\nVIEW"):
        super().__init__()
        layout = QGridLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        self.setLayout(layout)
        self.wall_buttons = {}
        self.side_keys = {"left", "right"}
        self.labels = labels or dict(self.WALLS)
        self.default_states = default_states or {"left": True, "right": True, "front": True, "back": True}

        def add_wall_button(key, label, row, col):
            button = RotatedButton() if key in {"left", "right"} else QPushButton()
            button.setCheckable(True)
            if key in {"left", "right"}:
                button.setFixedWidth(34)
                button.setMinimumHeight(92)
            else:
                button.setFixedWidth(88)
                button.setMinimumHeight(30)
            button.clicked.connect(self.update_button_styles)
            self.wall_buttons[key] = button
            layout.addWidget(button, row, col)

        add_wall_button("back", "Back", 0, 1)
        add_wall_button("left", "Left", 1, 0)

        footprint = QLabel(centre_text)
        footprint.setAlignment(Qt.AlignCenter)
        footprint.setFixedSize(74, 56)
        footprint.setStyleSheet("border: 1px solid #777; background: #f7f7f7; font-weight: bold;")
        layout.addWidget(footprint, 1, 1)

        add_wall_button("right", "Right", 1, 2)
        add_wall_button("front", "Front", 2, 1)
        self.set_wall_states(self.default_states)

    def wall_states(self):
        return {key: button.isChecked() for key, button in self.wall_buttons.items()}

    def set_wall_states(self, states):
        for key, button in self.wall_buttons.items():
            button.setChecked(bool(states.get(key, self.default_states.get(key, True))))
        self.update_button_styles()

    def update_button_styles(self):
        for key, button in self.wall_buttons.items():
            label = self.labels.get(key, key.title())
            state = "Clad" if button.isChecked() else "Open"
            if key in self.side_keys:
                label_text = f"{label} {state}"
                metrics = QFontMetrics(button.font())
                button.setText(metrics.elidedText(label_text, Qt.ElideRight, max(button.height() - 12, 40)))
            else:
                button.setText(f"{label}\n{state}")
            if button.isChecked():
                button.setStyleSheet("QPushButton { background: #dfeee4; border: 2px solid #2f7d44; font-weight: bold; }")
            else:
                button.setStyleSheet("QPushButton { background: #f7e1df; border: 2px dashed #a43d34; font-weight: bold; }")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_button_styles()


class RotatedButton(QPushButton):
    def __init__(self):
        super().__init__()
        self._display_text = ""

    def setText(self, text):
        self._display_text = text
        super().setText("")
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.translate(self.width() / 2, self.height() / 2)
        painter.rotate(-90)
        rect = QRectF(-self.height() / 2, -self.width() / 2, self.height(), self.width())
        painter.drawText(rect, Qt.AlignCenter, self._display_text)
        painter.end()


class LeanToEnvelopeProxy:
    def __init__(self, selector, side):
        self.selector = selector
        self.side = side

    def wall_states(self):
        return self.selector.lean_wall_states(self.side)

    def set_wall_states(self, states):
        self.selector.set_lean_wall_states(self.side, states)


class BuildingEnvelopeSelector(QWidget):
    def __init__(self):
        super().__init__()
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        self.setLayout(layout)

        self.left_selector = EnvelopePlanSelector(
            labels={"left": "Outer", "right": "Inner", "front": "Front", "back": "Back"},
            default_states={"left": False, "right": False, "front": False, "back": False},
            centre_text="LEFT\nLEAN-TO",
        )
        self.main_selector = EnvelopePlanSelector(centre_text="MAIN\nGABLE")
        self.right_selector = EnvelopePlanSelector(
            labels={"left": "Inner", "right": "Outer", "front": "Front", "back": "Back"},
            default_states={"left": False, "right": False, "front": False, "back": False},
            centre_text="RIGHT\nLEAN-TO",
        )

        for selector in [self.left_selector, self.main_selector, self.right_selector]:
            layout.addWidget(selector)

        self.left_selector.wall_buttons["right"].setVisible(False)
        self.right_selector.wall_buttons["left"].setVisible(False)
        self.left_lean_proxy = LeanToEnvelopeProxy(self, "left")
        self.right_lean_proxy = LeanToEnvelopeProxy(self, "right")
        self.wall_buttons = self.main_selector.wall_buttons
        self.set_lean_to_visibility(False, False)

    def all_wall_buttons(self):
        buttons = list(self.main_selector.wall_buttons.values())
        buttons.extend(button for key, button in self.left_selector.wall_buttons.items() if key != "right")
        buttons.extend(button for key, button in self.right_selector.wall_buttons.items() if key != "left")
        return buttons

    def set_lean_to_visibility(self, left_visible, right_visible):
        self.left_selector.setVisible(bool(left_visible))
        self.right_selector.setVisible(bool(right_visible))
        self.sync_inner_walls()

    def sync_inner_walls(self):
        main_states = self.wall_states()
        self.left_selector.wall_buttons["right"].setChecked(main_states.get("left", False))
        self.right_selector.wall_buttons["left"].setChecked(main_states.get("right", False))
        self.left_selector.update_button_styles()
        self.right_selector.update_button_styles()

    def wall_states(self):
        return self.main_selector.wall_states()

    def set_wall_states(self, states):
        self.main_selector.set_wall_states(states)
        self.sync_inner_walls()

    def lean_wall_states(self, side):
        self.sync_inner_walls()
        if side == "left":
            return self.left_selector.wall_states()
        return self.right_selector.wall_states()

    def set_lean_wall_states(self, side, states):
        states = states or {}
        if side == "left" and states.get("right"):
            main_states = self.wall_states()
            main_states["left"] = True
            self.main_selector.set_wall_states(main_states)
        if side == "right" and states.get("left"):
            main_states = self.wall_states()
            main_states["right"] = True
            self.main_selector.set_wall_states(main_states)
        selector = self.left_selector if side == "left" else self.right_selector
        selector.set_wall_states(states)
        self.sync_inner_walls()


# ==========================================================
# PLOT CANVAS
# ==========================================================

class PlotCanvas(FigureCanvasQTAgg):

    def __init__(self):
        self.fig = Figure(figsize=(12, 8))
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)

    def plot_structure(self, nodes, elements, base_nodes, support_types, load_arrows,
                       result_type="Deflection", displacements=None, scale=100, diagram_scale_factor=1.0,
                       purlin_layout=None, wall_cladding_segments=None):
        self.ax.clear()

        for element in elements:
            x1 = element.start.x
            y1 = element.start.y
            x2 = element.end.x
            y2 = element.end.y
            linewidth = 2.5 if element.group in ["TOP", "BOTTOM", "COLUMN", "LEFT_COLUMN", "RIGHT_COLUMN", "INTERNAL_COLUMN"] else 1.4
            self.ax.plot([x1, x2], [y1, y2], linewidth=linewidth)

        if result_type == "Geometry":
            pass
        elif result_type == "Load Diagram":
            self.plot_load_diagram_visual(load_arrows, diagram_scale_factor)
        elif result_type == "Deflection" and displacements is not None:
            self.plot_deflected_shape(elements, displacements, scale)
        elif result_type in ["Bending Moment", "Shear Force", "Axial Compression"] and displacements is not None:
            self.plot_force_diagram(elements, displacements, result_type, diagram_scale_factor)

        for node in nodes:
            self.ax.plot(node.x, node.y, marker="o", markersize=3)

        if purlin_layout is not None:
            self.plot_purlins(purlin_layout)
        if wall_cladding_segments:
            self.plot_wall_cladding(wall_cladding_segments)

        self.plot_supports(base_nodes, support_types)
        if result_type != "Geometry":
            self.plot_load_arrows(load_arrows, show_labels=(result_type == "Load Diagram"))

        self.ax.set_aspect("equal")
        self.ax.grid(True)
        self.ax.set_xlabel("X (mm)")
        self.ax.set_ylabel("Y (mm)")
        self.ax.set_title(f"2D Warren Portal Frame Analysis — {result_type}")
        if result_type == "Deflection":
            self.ax.text(0.01, 0.98, f"Deflection exaggerated x{scale:g}", transform=self.ax.transAxes,
                         va="top", ha="left", fontsize=8,
                         bbox={"boxstyle": "round,pad=0.25", "fc": "white", "alpha": 0.75})
        elif result_type in ["Load Diagram", "Bending Moment", "Shear Force", "Axial Compression"]:
            self.ax.text(0.01, 0.98, f"Diagram scale factor x{diagram_scale_factor:g}", transform=self.ax.transAxes,
                         va="top", ha="left", fontsize=8,
                         bbox={"boxstyle": "round,pad=0.25", "fc": "white", "alpha": 0.75})
        self.draw()

    def plot_purlins(self, purlin_layout):
        for x, y in purlin_layout.points:
            self.ax.plot(x, y, marker="s", markersize=4, color="#7a4c00")
            self.ax.plot([x, x], [y - 180, y + 180], color="#7a4c00", linewidth=1.1)

    def plot_wall_cladding(self, wall_cladding_segments):
        for segment in wall_cladding_segments:
            x = segment["x"]
            y1 = min(segment["y1"], segment["y2"])
            y2 = max(segment["y1"], segment["y2"])
            self.ax.plot([x, x], [y1, y2], color="#2f6f73", linewidth=3.0, alpha=0.65)
            height = max(y2 - y1, 0.0)
            if height <= 1e-9:
                continue
            girt_spacing = 1500.0
            count = max(1, int(math.floor(height / girt_spacing)))
            tick = 260.0
            for i in range(count + 1):
                y = y1 + min(i * girt_spacing, height)
                self.ax.plot([x - tick, x + tick], [y, y], color="#2f6f73", linewidth=1.1, alpha=0.85)

    def plot_deflected_shape(self, elements, displacements, scale):
        for element in elements:
            n1 = element.start
            n2 = element.end
            x1 = n1.x + displacements[n1.ux] * scale
            y1 = n1.y + displacements[n1.uy] * scale
            x2 = n2.x + displacements[n2.ux] * scale
            y2 = n2.y + displacements[n2.uy] * scale
            self.ax.plot([x1, x2], [y1, y2], linestyle="--", linewidth=1.4)

    def plot_force_diagram(self, elements, displacements, result_type, diagram_scale_factor=1.0):
        summaries = [el.force_summary(displacements) for el in elements]

        if result_type == "Bending Moment":
            max_value = max([s["moment_knm"] for s in summaries] + [1e-9])
            unit = "kNm"
        elif result_type == "Shear Force":
            max_value = max([s["shear_kn"] for s in summaries] + [1e-9])
            unit = "kN"
        else:
            max_value = max([s["compression_kn"] for s in summaries] + [1e-9])
            unit = "kN"

        all_x = [n for el in elements for n in [el.start.x, el.end.x]]
        structure_span = max(all_x) - min(all_x)
        diagram_scale = (max(structure_span * 0.04, 600) / max_value) * diagram_scale_factor

        for element, summary in zip(elements, summaries):
            x1 = element.start.x
            y1 = element.start.y
            x2 = element.end.x
            y2 = element.end.y
            dx = x2 - x1
            dy = y2 - y1
            L = math.sqrt(dx**2 + dy**2)
            if L == 0:
                continue

            nx = -dy / L
            ny = dx / L
            f = summary["end_forces"]

            if result_type == "Bending Moment":
                v1 = f[2] / 1_000_000
                v2 = -f[5] / 1_000_000
                label_value = summary["moment_knm"]
            elif result_type == "Shear Force":
                v1 = f[1] / 1000
                v2 = -f[4] / 1000
                label_value = summary["shear_kn"]
            else:
                v1 = summary["compression_kn"]
                v2 = summary["compression_kn"]
                label_value = summary["compression_kn"]

            ox1 = nx * v1 * diagram_scale
            oy1 = ny * v1 * diagram_scale
            ox2 = nx * v2 * diagram_scale
            oy2 = ny * v2 * diagram_scale

            self.ax.plot([x1, x1 + ox1, x2 + ox2, x2],
                         [y1, y1 + oy1, y2 + oy2, y2], linestyle="--", linewidth=1.2)

            mx = (x1 + x2) / 2 + nx * ((v1 + v2) / 2) * diagram_scale
            my = (y1 + y2) / 2 + ny * ((v1 + v2) / 2) * diagram_scale
            self.ax.text(mx, my, f"M{element.id}: {label_value:.1f} {unit}", fontsize=7)

    def plot_supports(self, base_nodes, support_types):
        if not base_nodes:
            return
        span = abs(base_nodes[1].x - base_nodes[0].x)
        size = max(span * 0.015, 250)

        for node, support_type in zip(base_nodes, support_types):
            x = node.x
            y = node.y
            if support_type == "Pinned":
                self.ax.plot([x - size, x, x + size, x - size], [y - size, y, y - size, y - size])
            elif support_type == "Fixed":
                self.ax.plot([x - size, x + size], [y, y], linewidth=3)
                for i in range(5):
                    hx = x - size + i * (2 * size / 4)
                    self.ax.plot([hx - size * 0.2, hx + size * 0.2], [y - size * 0.5, y], linewidth=1)
            elif support_type == "X restraint":
                self.ax.plot([x, x], [y - size * 0.65, y + size * 0.65], linewidth=2)
                self.ax.plot([x - size * 0.4, x], [y - size * 0.4, y], linewidth=1)
                self.ax.plot([x - size * 0.4, x], [y, y + size * 0.4], linewidth=1)
            elif support_type == "X spring":
                coils = 5
                xs = [x]
                ys = [y]
                for i in range(coils * 2):
                    xs.append(x + (-1 if i % 2 == 0 else 1) * size * 0.18)
                    ys.append(y - size * 0.5 + size * i / (coils * 2 - 1))
                xs.append(x)
                ys.append(y + size * 0.5)
                self.ax.plot(xs, ys, linewidth=1.5)
                self.ax.plot([x - size * 0.35, x + size * 0.35], [y + size * 0.5, y + size * 0.5], linewidth=2)

    def combined_load_arrows(self, load_arrows):
        """
        Combine coincident load arrows before plotting.

        This avoids the old sawtooth effect when, for example, dead load and wind
        load were drawn as separate arrows at the same roof panel midpoint.
        """
        combined = {}
        for x, y, fx, fy in load_arrows:
            key = (round(x, 6), round(y, 6))
            old_fx, old_fy = combined.get(key, (0.0, 0.0))
            combined[key] = (old_fx + fx, old_fy + fy)
        return [(x, y, fx, fy) for (x, y), (fx, fy) in combined.items()]

    def plot_load_diagram_visual(self, load_arrows, diagram_scale_factor=1.0):
        """Draw a scaled resultant load diagram using combined loads."""
        combined = self.combined_load_arrows(load_arrows)
        combined = [(x, y, fx, fy) for x, y, fx, fy in combined if math.sqrt(fx**2 + fy**2) > 1e-9]
        if not combined:
            return

        max_mag = max([math.sqrt(fx**2 + fy**2) for _, _, fx, fy in combined] + [1e-9])
        all_x = [x for x, _, _, _ in combined]
        structure_span = max(all_x) - min(all_x) if len(all_x) > 1 else 10000
        diagram_scale = (max(structure_span * 0.04, 600) / max_mag) * diagram_scale_factor

        vertical_points = []
        horizontal_points = []

        for x, y, fx, fy in combined:
            mag = math.sqrt(fx**2 + fy**2)
            if mag == 0:
                continue

            ox = fx * diagram_scale
            oy = fy * diagram_scale

            self.ax.plot([x, x + ox], [y, y + oy], linestyle="--", linewidth=1.1)
            self.ax.plot(x + ox, y + oy, marker=".", markersize=5)

            if abs(fy) >= abs(fx):
                vertical_points.append((x, y, y + oy))
            else:
                horizontal_points.append((y, x, x + ox))

        # Plot horizontal load envelopes per column/member station.
        # Previously all left/right acting horizontal loads were connected together.
        # On an enclosed frame, both columns can have horizontal Cpe arrows acting in
        # the same global direction, so sorting only by height connected the left and
        # right column points into a jagged zig-zag across the bay.  Group by the
        # original arrow x-position first so each column gets its own clean line.
        horizontal_by_station = {}
        for y, xbase, xoff in horizontal_points:
            station_key = round(xbase, 6)
            horizontal_by_station.setdefault(station_key, []).append((xoff, y, xbase))

        for station_points in horizontal_by_station.values():
            left = sorted([(xoff, y) for xoff, y, xbase in station_points if xoff < xbase], key=lambda p: p[1])
            right = sorted([(xoff, y) for xoff, y, xbase in station_points if xoff >= xbase], key=lambda p: p[1])
            if len(left) > 1:
                self.ax.plot([p[0] for p in left], [p[1] for p in left], linestyle="--", linewidth=1.3)
            if len(right) > 1:
                self.ax.plot([p[0] for p in right], [p[1] for p in right], linestyle="--", linewidth=1.3)

    def vertical_load_bands(self, vertical_points):
        if not vertical_points:
            return []
        points = sorted(vertical_points, key=lambda p: p[0])
        xs = [p[0] for p in points]
        ys = [p[1] for p in points]
        x_span = max(xs) - min(xs)
        y_span = max(ys) - min(ys)
        max_gap = max(x_span * 0.12, 1200.0)
        y_tol = max(y_span * 0.08, 500.0)
        bands = []
        current = []
        last = None

        for x, ybase, yoff in points:
            sign = 1 if yoff >= ybase else -1
            if last is None:
                current = [(x, yoff)]
            else:
                last_x, last_ybase, last_sign = last
                gap = abs(x - last_x)
                base_jump = abs(ybase - last_ybase)
                if sign != last_sign or gap > max_gap or base_jump > y_tol:
                    if len(current) > 1:
                        bands.append(current)
                    current = [(x, yoff)]
                else:
                    current.append((x, yoff))
            last = (x, ybase, sign)

        if len(current) > 1:
            bands.append(current)
        return bands

    def plot_load_arrows(self, load_arrows, show_labels=False):
        arrows = self.combined_load_arrows(load_arrows) if show_labels else load_arrows
        for x, y, fx, fy in arrows:
            mag = math.sqrt(fx**2 + fy**2)
            if mag == 0:
                continue
            arrow_scale = 0.00008
            dx = fx * arrow_scale
            dy = fy * arrow_scale
            self.ax.arrow(x - dx, y - dy, dx, dy, head_width=250, head_length=300,
                          length_includes_head=True)
            if show_labels:
                label = f"Fx={fx / 1000:+.1f} kN\nFy={fy / 1000:+.1f} kN"
                self.ax.text(x, y, label, fontsize=7)


# ==========================================================
# MAIN WINDOW
# ==========================================================

class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("SpaceGASSBuilder — 2D Solver with Profile Catalog")
        self.resize(1820, 980)

        self.catalog = None
        self.metroll_tables = None
        self.structure = None
        self.displacements = None
        self.reactions = None
        self.force_vector = None
        self.base_nodes = None
        self.support_types = None
        self.load_arrows = None
        self.single_load_arrows = {}
        self.design_envelope_actions = None
        self.member_check_restraints = None
        self.purlin_layout = None
        self.purlin_check = None
        self.wall_girt_check = None
        self.costing_summary = None
        self.foundation_summary = None
        self.wall_cladding_segments = []
        self.concept_roof_top_nodes = []
        self.concept_main_top_nodes = []
        self.concept_building_length_m = 0.0
        self.concept_bay_size_m = 0.0
        self.concept_wall_states = {}
        self.serviceability_height_mm = None
        self.serviceability_span_mm = None
        self.serviceability_importance_level = None
        self.serviceability_crane_active = False
        self.analysis_info = None
        self.auto_size_report_text = ""
        self.load_combinations = build_load_combinations(CPE_CASE_OPTIONS, CPI_CASE_OPTIONS)
        self.factory_default_data = None
        self.current_preset_path = None
        self.confirm_on_close = True

        self.load_catalog()
        self.load_metroll_tables()
        self.build_ui()
        self.factory_default_data = self.collect_preset_data()
        self.load_default_inputs(silent=True)

    def load_catalog(self):
        catalog_path = resource_path("Shape_Catalog.xlsx")
        if not os.path.exists(catalog_path):
            raise FileNotFoundError(f"Could not find Shape_Catalog.xlsx at {catalog_path}")
        self.catalog = ShapeCatalog(catalog_path)

    def load_metroll_tables(self):
        tables_path = resource_path("metroll_megaspan_capacity_tables.json")
        if os.path.exists(tables_path):
            with open(tables_path, "r", encoding="utf-8") as file:
                self.metroll_tables = json.load(file)

    def build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QHBoxLayout()
        central.setLayout(layout)

        left_panel = QWidget()
        left_layout = QVBoxLayout()
        left_panel.setLayout(left_layout)

        tabs = QTabWidget()
        left_layout.addWidget(tabs)

        # ---------------- Geometry / Wind tab ----------------
        geometry_wind_tab = QScrollArea()
        geometry_wind_tab.setWidgetResizable(True)
        geometry_wind_content = QWidget()
        geometry_wind_layout = QFormLayout()
        geometry_wind_content.setLayout(geometry_wind_layout)
        geometry_wind_tab.setWidget(geometry_wind_content)

        advanced_tab = QScrollArea()
        advanced_tab.setWidgetResizable(True)
        advanced_content = QWidget()
        advanced_layout = QFormLayout()
        advanced_content.setLayout(advanced_layout)
        advanced_tab.setWidget(advanced_content)

        self.span_input = QLineEdit("30000")
        self.building_length_input = QLineEdit("40.0")
        self.pitch_input = QLineEdit("10")
        self.eave_input = QLineEdit("7000")
        self.depth_input = QLineEdit("900")
        self.frame_system_combo = QComboBox()
        self.frame_system_combo.addItems(["Truss", "Rafter"])
        self.rafter_haunch_length_input = QLineEdit("3000")
        self.load_bay_size_input = QLineEdit("8000")
        self.snow_region_combo = QComboBox()
        self.snow_region_combo.addItems(["None", "Region AN", "Region AC", "Region AS", "Region AT"])
        self.snow_ahd_height_input = QLineEdit("0")
        self.internal_column_count_combo = QComboBox()
        self.internal_column_count_combo.addItems(["None", "1 central", "2 columns"])
        self.internal_column_offset_input = QLineEdit("3000")
        self.span_input.editingFinished.connect(self.update_truss_depth_from_span)

        geometry_wind_layout.addRow("Span (mm)", self.span_input)
        geometry_wind_layout.addRow("Building Length (m)", self.building_length_input)
        geometry_wind_layout.addRow("Roof Pitch (deg)", self.pitch_input)
        geometry_wind_layout.addRow("Eave Height (mm)", self.eave_input)
        geometry_wind_layout.addRow("Frame system", self.frame_system_combo)
        geometry_wind_layout.addRow("Rafter haunch length (mm)", self.rafter_haunch_length_input)
        self.rafter_haunch_rows = [
            geometry_wind_layout.labelForField(self.rafter_haunch_length_input),
            self.rafter_haunch_length_input,
        ]
        geometry_wind_layout.addRow("Tributary bay size (mm)", self.load_bay_size_input)
        geometry_wind_layout.addRow("Internal columns", self.internal_column_count_combo)
        geometry_wind_layout.addRow("Internal column offset from centre (mm)", self.internal_column_offset_input)
        self.internal_column_offset_rows = [
            geometry_wind_layout.labelForField(self.internal_column_offset_input),
            self.internal_column_offset_input,
        ]

        # ---------------- Canopy / lean-to geometry ----------------
        canopies_layout = geometry_wind_layout

        self.left_canopy_length_input = QLineEdit("0")
        self.right_canopy_length_input = QLineEdit("0")
        self.left_canopy_type_combo = QComboBox()
        self.left_canopy_type_combo.addItems(["None", "Canopy", "Lean-to"])
        self.right_canopy_type_combo = QComboBox()
        self.right_canopy_type_combo.addItems(["None", "Canopy", "Lean-to"])
        self.left_canopy_wind_model_combo = QComboBox()
        self.left_canopy_wind_model_combo.addItems(["Main roof Cpe + Cpi", "AS/NZS B.5 free roof", "Manual net Cpn"])
        self.right_canopy_wind_model_combo = QComboBox()
        self.right_canopy_wind_model_combo.addItems(["Main roof Cpe + Cpi", "AS/NZS B.5 free roof", "Manual net Cpn"])
        self.left_canopy_underside_combo = QComboBox()
        self.left_canopy_underside_combo.addItems(["Empty under", "Blocked under"])
        self.right_canopy_underside_combo = QComboBox()
        self.right_canopy_underside_combo.addItems(["Empty under", "Blocked under"])
        self.left_lean_outer_wall_clad_combo = QComboBox()
        self.left_lean_outer_wall_clad_combo.addItems(["No", "Yes"])
        self.right_lean_outer_wall_clad_combo = QComboBox()
        self.right_lean_outer_wall_clad_combo.addItems(["No", "Yes"])
        self.left_canopy_eave_height_input = QLineEdit("")
        self.right_canopy_eave_height_input = QLineEdit("")
        self.left_canopy_pitch_input = QLineEdit("5")
        self.right_canopy_pitch_input = QLineEdit("5")
        self.left_canopy_cpn_uplift_input = QLineEdit("0.0")
        self.left_canopy_cpn_downward_input = QLineEdit("0.0")
        self.right_canopy_cpn_uplift_input = QLineEdit("0.0")
        self.right_canopy_cpn_downward_input = QLineEdit("0.0")
        self.left_canopy_shallow_depth_input = QLineEdit("300")
        self.right_canopy_shallow_depth_input = QLineEdit("300")
        self.left_canopy_g_load_input = QLineEdit("0.0")
        self.left_canopy_q_load_input = QLineEdit("0.0")
        self.left_canopy_solar_load_input = QLineEdit("0.0")
        self.right_canopy_g_load_input = QLineEdit("0.0")
        self.right_canopy_q_load_input = QLineEdit("0.0")
        self.right_canopy_solar_load_input = QLineEdit("0.0")

        canopy_heading = QLabel("Canopy / lean-to geometry")
        canopy_heading.setStyleSheet("font-weight: bold; margin-top: 8px;")
        canopies_layout.addRow(canopy_heading)
        canopies_layout.addRow("Left addition", self.left_canopy_type_combo)
        canopies_layout.addRow("Left width (mm)", self.left_canopy_length_input)
        canopies_layout.addRow("Left eave height override (mm)", self.left_canopy_eave_height_input)
        canopies_layout.addRow("Left pitch (deg)", self.left_canopy_pitch_input)
        canopies_layout.addRow("Left wind model", self.left_canopy_wind_model_combo)
        self.left_canopy_wind_model_rows = [
            canopies_layout.labelForField(self.left_canopy_wind_model_combo),
            self.left_canopy_wind_model_combo,
        ]
        canopies_layout.addRow("Left underside", self.left_canopy_underside_combo)
        self.left_canopy_underside_rows = [
            canopies_layout.labelForField(self.left_canopy_underside_combo),
            self.left_canopy_underside_combo,
        ]
        canopies_layout.addRow("Left manual Cpn uplift (+ up)", self.left_canopy_cpn_uplift_input)
        canopies_layout.addRow("Left manual Cpn downward (+ down)", self.left_canopy_cpn_downward_input)
        self.left_canopy_manual_cpn_rows = [
            canopies_layout.labelForField(self.left_canopy_cpn_uplift_input),
            self.left_canopy_cpn_uplift_input,
            canopies_layout.labelForField(self.left_canopy_cpn_downward_input),
            self.left_canopy_cpn_downward_input,
        ]
        self.left_canopy_geometry_rows = [
            canopies_layout.labelForField(self.left_canopy_length_input),
            self.left_canopy_length_input,
            canopies_layout.labelForField(self.left_canopy_eave_height_input),
            self.left_canopy_eave_height_input,
            canopies_layout.labelForField(self.left_canopy_pitch_input),
            self.left_canopy_pitch_input,
        ]
        self.left_lean_enclosure_rows = []

        canopies_layout.addRow("Right addition", self.right_canopy_type_combo)
        canopies_layout.addRow("Right width (mm)", self.right_canopy_length_input)
        canopies_layout.addRow("Right eave height override (mm)", self.right_canopy_eave_height_input)
        canopies_layout.addRow("Right pitch (deg)", self.right_canopy_pitch_input)
        canopies_layout.addRow("Right wind model", self.right_canopy_wind_model_combo)
        self.right_canopy_wind_model_rows = [
            canopies_layout.labelForField(self.right_canopy_wind_model_combo),
            self.right_canopy_wind_model_combo,
        ]
        canopies_layout.addRow("Right underside", self.right_canopy_underside_combo)
        self.right_canopy_underside_rows = [
            canopies_layout.labelForField(self.right_canopy_underside_combo),
            self.right_canopy_underside_combo,
        ]
        canopies_layout.addRow("Right manual Cpn uplift (+ up)", self.right_canopy_cpn_uplift_input)
        canopies_layout.addRow("Right manual Cpn downward (+ down)", self.right_canopy_cpn_downward_input)
        self.right_canopy_manual_cpn_rows = [
            canopies_layout.labelForField(self.right_canopy_cpn_uplift_input),
            self.right_canopy_cpn_uplift_input,
            canopies_layout.labelForField(self.right_canopy_cpn_downward_input),
            self.right_canopy_cpn_downward_input,
        ]
        self.right_canopy_geometry_rows = [
            canopies_layout.labelForField(self.right_canopy_length_input),
            self.right_canopy_length_input,
            canopies_layout.labelForField(self.right_canopy_eave_height_input),
            self.right_canopy_eave_height_input,
            canopies_layout.labelForField(self.right_canopy_pitch_input),
            self.right_canopy_pitch_input,
        ]
        self.right_lean_enclosure_rows = []

        self.left_support_combo = QComboBox()
        self.left_support_combo.addItems(["Pinned", "Fixed"])
        self.right_support_combo = QComboBox()
        self.right_support_combo.addItems(["Pinned", "Fixed"])
        self.eave_x_restraint_combo = QComboBox()
        self.eave_x_restraint_combo.addItems(["None", "Left eave", "Right eave", "Both eaves", "End wall bracing approx"])
        self.eave_x_spring_input = QLineEdit("0.0")
        self.end_wall_brace_dia_input = QLineEdit("12.0")
        self.license_status_label = QLabel()
        self.license_status_label.setWordWrap(True)
        self.refresh_license_status_button = QPushButton("Refresh")
        self.refresh_license_status_button.clicked.connect(lambda: self.update_license_status_label(refresh_server=True))
        license_status_row = QWidget()
        license_status_layout = QHBoxLayout()
        license_status_layout.setContentsMargins(0, 0, 0, 0)
        license_status_row.setLayout(license_status_layout)
        license_status_layout.addWidget(self.license_status_label, 1)
        license_status_layout.addWidget(self.refresh_license_status_button)
        advanced_layout.addRow("License", license_status_row)
        advanced_layout.addRow("Left Base Support", self.left_support_combo)
        advanced_layout.addRow("Right Base Support", self.right_support_combo)
        advanced_layout.addRow("Eave X restraint", self.eave_x_restraint_combo)
        advanced_layout.addRow("Manual eave X spring (kN/mm)", self.eave_x_spring_input)
        self.eave_x_spring_rows = [
            advanced_layout.labelForField(self.eave_x_spring_input),
            self.eave_x_spring_input,
        ]
        advanced_layout.addRow("End wall brace dia (mm)", self.end_wall_brace_dia_input)
        self.end_wall_brace_rows = [
            advanced_layout.labelForField(self.end_wall_brace_dia_input),
            self.end_wall_brace_dia_input,
        ]

        self.wind_combination_note = QLabel("Wu combinations use every Cpe/Cpi pairing from these wind inputs. Md uses the worst multiplier in the selected direction's 90 degree window.")
        self.wind_combination_note.setWordWrap(True)
        self.wind_region_combo = QComboBox()
        self.wind_region_combo.addItems(["A0", "A1", "A2", "A3", "A4", "A5", "B1", "B2", "C", "D"])
        self.wind_region_combo.setCurrentText("A5")

        self.importance_combo = QComboBox()
        self.importance_combo.addItems(["1", "2", "3", "4"])
        self.terrain_combo = QComboBox()
        self.terrain_combo.addItems(["1", "2", "2.5", "3", "4"])
        self.terrain_combo.setCurrentText("2")

        self.wind_orientation_selector = CompassSelector(default="E")

        self.frame_type_combo = QComboBox()
        self.frame_type_combo.addItems(["Enclosed", "Roof Only", "3 Sided", "2 Sided", "1 Sided", "No Wind"])
        self.frame_type_combo.setCurrentText("3 Sided")
        self.envelope_plan_selector = BuildingEnvelopeSelector()
        self.left_lean_enclosure_selector = self.envelope_plan_selector.left_lean_proxy
        self.right_lean_enclosure_selector = self.envelope_plan_selector.right_lean_proxy
        self.derived_frame_type_label = QLabel("-")

        self.wind_case_combo = QComboBox()
        self.wind_case_combo.addItems(WIND_CASE_OPTIONS)

        self.wind_reduction_input = QLineEdit("1.0")

        geometry_wind_layout.addRow(self.wind_combination_note)
        geometry_wind_layout.addRow("Wind Region", self.wind_region_combo)
        geometry_wind_layout.addRow("Importance Level", self.importance_combo)
        geometry_wind_layout.addRow("Terrain Category", self.terrain_combo)
        geometry_wind_layout.addRow("Snow Region", self.snow_region_combo)
        geometry_wind_layout.addRow("AHD Height h0 (m)", self.snow_ahd_height_input)
        self.snow_ahd_height_rows = [
            geometry_wind_layout.labelForField(self.snow_ahd_height_input),
            self.snow_ahd_height_input,
        ]
        geometry_wind_layout.addRow("North direction in section view", self.wind_orientation_selector)
        geometry_wind_layout.addRow("Wall enclosure", self.envelope_plan_selector)
        geometry_wind_layout.addRow("Derived frame type", self.derived_frame_type_label)
        geometry_wind_layout.addRow("Legacy wind mode", self.frame_type_combo)

        tabs.addTab(geometry_wind_tab, "Geometry/Wind")
        tabs.addTab(advanced_tab, "Advanced")

        # ---------------- Sections tab ----------------
        sections_tab = QWidget()
        sections_layout = QFormLayout()
        sections_tab.setLayout(sections_layout)

        self.left_column_profile_combo = QComboBox()
        self.right_column_profile_combo = QComboBox()
        self.top_profile_combo = QComboBox()
        self.bottom_profile_combo = self.top_profile_combo
        self.purlin_span_type_combo = QComboBox()
        self.purlin_span_type_combo.addItems(["Single span", "Double span", "Continuous lapped"])
        self.wall_girt_span_type_combo = QComboBox()
        self.wall_girt_span_type_combo.addItems(["Single span", "Double span", "Continuous lapped"])
        self.web_profile_combo = QComboBox()
        self.internal_column_profile_combo = QComboBox()
        self.left_canopy_top_profile_combo = QComboBox()
        self.left_canopy_bottom_profile_combo = self.left_canopy_top_profile_combo
        self.left_canopy_web_profile_combo = QComboBox()
        self.left_lean_column_profile_combo = QComboBox()
        self.right_canopy_top_profile_combo = QComboBox()
        self.right_canopy_bottom_profile_combo = self.right_canopy_top_profile_combo
        self.right_canopy_web_profile_combo = QComboBox()
        self.right_lean_column_profile_combo = QComboBox()

        for profile in self.catalog.all_profiles():
            self.left_column_profile_combo.addItem(profile.display_name)
            self.right_column_profile_combo.addItem(profile.display_name)
            self.internal_column_profile_combo.addItem(profile.display_name)
            self.left_lean_column_profile_combo.addItem(profile.display_name)
            self.right_lean_column_profile_combo.addItem(profile.display_name)

        shs_profiles = self.catalog.shs_profiles()
        main_chord_profiles = shs_profiles + self.catalog.ub_profiles()
        for profile in main_chord_profiles:
            self.top_profile_combo.addItem(profile.display_name)
        for combo in [
            self.web_profile_combo,
            self.left_canopy_top_profile_combo, self.left_canopy_web_profile_combo,
            self.right_canopy_top_profile_combo, self.right_canopy_web_profile_combo,
        ]:
            for profile in shs_profiles:
                combo.addItem(profile.display_name)

        self.set_default_combo(self.left_column_profile_combo, ["250 UC", "200 UC", "310 UB", "250 PFC"])
        self.set_default_combo(self.right_column_profile_combo, ["250 UC", "200 UC", "310 UB", "250 PFC"])
        self.set_default_combo(self.internal_column_profile_combo, ["200 UC", "150 UC", "250 UC", "100 SHS"])
        self.set_default_combo(self.top_profile_combo, ["150 SHS", "125 SHS", "100 SHS"])
        self.set_default_combo(self.web_profile_combo, ["50*2.5 SHS", "50 x 2.5 SHS", "50 SHS"])
        for combo in [self.left_canopy_top_profile_combo, self.right_canopy_top_profile_combo]:
            self.set_default_combo(combo, ["100 SHS", "89 SHS", "75 SHS"])
        for combo in [self.left_canopy_web_profile_combo, self.right_canopy_web_profile_combo]:
            self.set_default_combo(combo, ["50*2.5 SHS", "50 x 2.5 SHS", "50 SHS"])
        for combo in [self.left_lean_column_profile_combo, self.right_lean_column_profile_combo]:
            self.set_default_combo(combo, ["150 UC", "100 UC", "150 PFC", "100 SHS"])

        self.left_column_check_label = QLabel("Not checked")
        self.right_column_check_label = QLabel("Not checked")
        self.internal_column_check_label = QLabel("Not checked")
        self.top_chord_check_label = QLabel("Not checked")
        self.bottom_chord_check_label = QLabel("Not checked")
        self.purlin_check_label = QLabel("Not checked")
        self.wall_girt_check_label = QLabel("Not checked")
        self.web_check_label = QLabel("Not checked")
        self.left_canopy_top_check_label = QLabel("Not checked")
        self.left_canopy_bottom_check_label = QLabel("Not checked")
        self.left_canopy_web_check_label = QLabel("Not checked")
        self.left_lean_column_check_label = QLabel("Not checked")
        self.right_canopy_top_check_label = QLabel("Not checked")
        self.right_canopy_bottom_check_label = QLabel("Not checked")
        self.right_canopy_web_check_label = QLabel("Not checked")
        self.right_lean_column_check_label = QLabel("Not checked")
        self.column_deflection_check_label = QLabel("Not checked")
        self.truss_deflection_check_label = QLabel("Not checked")
        self.total_cost_label = QLabel("Not calculated")
        for label in [
            self.left_column_check_label,
            self.right_column_check_label,
            self.internal_column_check_label,
            self.top_chord_check_label,
            self.bottom_chord_check_label,
            self.purlin_check_label,
            self.wall_girt_check_label,
            self.web_check_label,
            self.left_canopy_top_check_label,
            self.left_canopy_bottom_check_label,
            self.left_canopy_web_check_label,
            self.left_lean_column_check_label,
            self.right_canopy_top_check_label,
            self.right_canopy_bottom_check_label,
            self.right_canopy_web_check_label,
            self.right_lean_column_check_label,
            self.column_deflection_check_label,
            self.truss_deflection_check_label,
            self.total_cost_label,
        ]:
            label.setStyleSheet("font-weight: bold;")

        def section_group(title):
            group = QGroupBox(title)
            layout = QFormLayout()
            group.setLayout(layout)
            sections_layout.addRow(group)
            return layout

        columns_layout = section_group("Columns")
        columns_layout.addRow("Left column profile", self.section_selector_row(self.left_column_profile_combo))
        columns_layout.addRow("Left column check", self.left_column_check_label)
        columns_layout.addRow("Right column profile", self.section_selector_row(self.right_column_profile_combo))
        columns_layout.addRow("Right column check", self.right_column_check_label)
        internal_column_profile_widget = self.section_selector_row(self.internal_column_profile_combo)
        columns_layout.addRow("Internal column profile", internal_column_profile_widget)
        self.internal_column_section_rows = [columns_layout.labelForField(internal_column_profile_widget), internal_column_profile_widget]
        columns_layout.addRow("Internal column check", self.internal_column_check_label)
        self.internal_column_check_rows = [columns_layout.labelForField(self.internal_column_check_label), self.internal_column_check_label]
        columns_layout.addRow("Column deflection", self.column_deflection_check_label)

        truss_layout = section_group("Truss")
        truss_layout.addRow("Top/bottom chord / rafter", self.section_selector_row(self.top_profile_combo))
        truss_layout.addRow("Top chord check", self.top_chord_check_label)
        truss_layout.addRow("Bottom chord check", self.bottom_chord_check_label)
        self.bottom_chord_check_rows = [truss_layout.labelForField(self.bottom_chord_check_label), self.bottom_chord_check_label]
        web_profile_widget = self.section_selector_row(self.web_profile_combo)
        truss_layout.addRow("Web/post SHS", web_profile_widget)
        self.web_section_rows = [truss_layout.labelForField(web_profile_widget), web_profile_widget]
        truss_layout.addRow("Web/post check", self.web_check_label)
        self.web_check_rows = [truss_layout.labelForField(self.web_check_label), self.web_check_label]
        truss_depth_widget = self.depth_selector_row()
        truss_layout.addRow("Truss depth (mm)", truss_depth_widget)
        self.truss_depth_rows = [truss_layout.labelForField(truss_depth_widget), truss_depth_widget]
        truss_layout.addRow("Truss midspan deflection", self.truss_deflection_check_label)

        purlin_layout = section_group("Purlins / Girts")
        purlin_layout.addRow("Roof purlin span type", self.purlin_span_type_combo)
        purlin_layout.addRow("Roof purlin output", self.purlin_check_label)
        purlin_layout.addRow("Wall girt span type", self.wall_girt_span_type_combo)
        purlin_layout.addRow("Wall girt output", self.wall_girt_check_label)

        costing_summary_layout = section_group("Costing")
        costing_summary_layout.addRow("Total estimated steel cost", self.total_cost_label)

        canopy_layout = section_group("Canopies / Lean-Tos")
        self.left_canopy_section_rows = []
        self.right_canopy_section_rows = []
        self.left_lean_section_rows = []
        self.right_lean_section_rows = []
        self.left_canopy_check_rows = []
        self.right_canopy_check_rows = []
        self.left_lean_check_rows = []
        self.right_lean_check_rows = []
        left_depth_widget = self.canopy_depth_selector_row(self.left_canopy_shallow_depth_input)
        canopy_layout.addRow("Left canopy/lean-to depth (mm)", left_depth_widget)
        self.left_canopy_section_rows.extend([canopy_layout.labelForField(left_depth_widget), left_depth_widget])
        for label, widget, check_label, check_widget in [
            ("Left canopy/lean-to top/bottom chord", self.section_selector_row(self.left_canopy_top_profile_combo), "Left canopy top check", self.left_canopy_top_check_label),
            ("Left canopy/lean-to webs", self.section_selector_row(self.left_canopy_web_profile_combo), "Left canopy web/post check", self.left_canopy_web_check_label),
        ]:
            canopy_layout.addRow(label, widget)
            self.left_canopy_section_rows.extend([canopy_layout.labelForField(widget), widget])
            if "web" in label.lower():
                self.left_canopy_web_section_rows = [canopy_layout.labelForField(widget), widget]
            canopy_layout.addRow(check_label, check_widget)
            self.left_canopy_check_rows.extend([canopy_layout.labelForField(check_widget), check_widget])
            if "web" in check_label.lower():
                self.left_canopy_web_check_rows = [canopy_layout.labelForField(check_widget), check_widget]
        canopy_layout.addRow("Left canopy bottom check", self.left_canopy_bottom_check_label)
        self.left_canopy_check_rows.extend([canopy_layout.labelForField(self.left_canopy_bottom_check_label), self.left_canopy_bottom_check_label])
        left_lean_widget = self.section_selector_row(self.left_lean_column_profile_combo)
        canopy_layout.addRow("Left lean-to outer column", left_lean_widget)
        self.left_lean_section_rows.extend([canopy_layout.labelForField(left_lean_widget), left_lean_widget])
        canopy_layout.addRow("Left lean-to column check", self.left_lean_column_check_label)
        self.left_lean_check_rows = [canopy_layout.labelForField(self.left_lean_column_check_label), self.left_lean_column_check_label]

        right_depth_widget = self.canopy_depth_selector_row(self.right_canopy_shallow_depth_input)
        canopy_layout.addRow("Right canopy/lean-to depth (mm)", right_depth_widget)
        self.right_canopy_section_rows.extend([canopy_layout.labelForField(right_depth_widget), right_depth_widget])
        for label, widget, check_label, check_widget in [
            ("Right canopy/lean-to top/bottom chord", self.section_selector_row(self.right_canopy_top_profile_combo), "Right canopy top check", self.right_canopy_top_check_label),
            ("Right canopy/lean-to webs", self.section_selector_row(self.right_canopy_web_profile_combo), "Right canopy web/post check", self.right_canopy_web_check_label),
        ]:
            canopy_layout.addRow(label, widget)
            self.right_canopy_section_rows.extend([canopy_layout.labelForField(widget), widget])
            if "web" in label.lower():
                self.right_canopy_web_section_rows = [canopy_layout.labelForField(widget), widget]
            canopy_layout.addRow(check_label, check_widget)
            self.right_canopy_check_rows.extend([canopy_layout.labelForField(check_widget), check_widget])
            if "web" in check_label.lower():
                self.right_canopy_web_check_rows = [canopy_layout.labelForField(check_widget), check_widget]
        canopy_layout.addRow("Right canopy bottom check", self.right_canopy_bottom_check_label)
        self.right_canopy_check_rows.extend([canopy_layout.labelForField(self.right_canopy_bottom_check_label), self.right_canopy_bottom_check_label])
        right_lean_widget = self.section_selector_row(self.right_lean_column_profile_combo)
        canopy_layout.addRow("Right lean-to outer column", right_lean_widget)
        self.right_lean_section_rows.extend([canopy_layout.labelForField(right_lean_widget), right_lean_widget])
        canopy_layout.addRow("Right lean-to column check", self.right_lean_column_check_label)
        self.right_lean_check_rows = [canopy_layout.labelForField(self.right_lean_column_check_label), self.right_lean_column_check_label]
        self.update_canopy_section_visibility(False, False, False, False)

        tabs.addTab(sections_tab, "Sections")

        # ---------------- Loads tab ----------------
        loads_tab = QWidget()
        loads_layout = QFormLayout()
        loads_tab.setLayout(loads_layout)

        self.g_load_input = QLineEdit("0.10")
        self.q_load_input = QLineEdit("0.0")
        self.solar_load_input = QLineEdit("0.0")
        self.fire_service_load_input = QLineEdit("0.0")
        self.hvac_load_input = QLineEdit("0.0")
        self.other_load_input = QLineEdit("0.0")
        self.crane_rating_input = QLineEdit("0.0")
        self.left_wall_load_input = QLineEdit("0.0")
        self.right_wall_load_input = QLineEdit("0.0")

        loads_layout.addRow("G roof area load downward (kPa)", self.g_load_input)
        loads_layout.addRow("G roof extent", self.roof_load_placement_row("g"))
        loads_layout.addRow("Q roof area load downward (kPa)", self.q_load_input)
        loads_layout.addRow("Q roof extent", self.roof_load_placement_row("q"))
        loads_layout.addRow("Solar roof area load downward (kPa)", self.solar_load_input)
        loads_layout.addRow("Solar roof extent", self.roof_load_placement_row("solar"))
        loads_layout.addRow("Fire service area load downward (kPa)", self.fire_service_load_input)
        loads_layout.addRow("HVAC roof area load downward (kPa)", self.hvac_load_input)
        loads_layout.addRow("Other roof area load downward (kPa)", self.other_load_input)
        loads_layout.addRow("Other roof extent", self.roof_load_placement_row("other"))
        loads_layout.addRow("Crane Rating (t)", self.crane_rating_input)
        loads_layout.addRow("Left Wall Load, +X (kN/m)", self.left_wall_load_input)
        self.left_wall_load_rows = [loads_layout.labelForField(self.left_wall_load_input), self.left_wall_load_input]
        loads_layout.addRow("Right Wall Load, -X (kN/m)", self.right_wall_load_input)
        self.right_wall_load_rows = [loads_layout.labelForField(self.right_wall_load_input), self.right_wall_load_input]
        loads_layout.addRow("Left canopy G load downward (kPa)", self.left_canopy_g_load_input)
        self.left_canopy_load_rows = [loads_layout.labelForField(self.left_canopy_g_load_input), self.left_canopy_g_load_input]
        loads_layout.addRow("Left canopy Q load downward (kPa)", self.left_canopy_q_load_input)
        self.left_canopy_load_rows.extend([loads_layout.labelForField(self.left_canopy_q_load_input), self.left_canopy_q_load_input])
        loads_layout.addRow("Left canopy solar load downward (kPa)", self.left_canopy_solar_load_input)
        self.left_canopy_load_rows.extend([loads_layout.labelForField(self.left_canopy_solar_load_input), self.left_canopy_solar_load_input])
        loads_layout.addRow("Right canopy G load downward (kPa)", self.right_canopy_g_load_input)
        self.right_canopy_load_rows = [loads_layout.labelForField(self.right_canopy_g_load_input), self.right_canopy_g_load_input]
        loads_layout.addRow("Right canopy Q load downward (kPa)", self.right_canopy_q_load_input)
        self.right_canopy_load_rows.extend([loads_layout.labelForField(self.right_canopy_q_load_input), self.right_canopy_q_load_input])
        loads_layout.addRow("Right canopy solar load downward (kPa)", self.right_canopy_solar_load_input)
        self.right_canopy_load_rows.extend([loads_layout.labelForField(self.right_canopy_solar_load_input), self.right_canopy_solar_load_input])

        for widget in self.left_wall_load_rows + self.right_wall_load_rows:
            if widget is not None:
                widget.setVisible(False)

        self.load_combination_combo = QComboBox()
        self.load_combination_combo.addItems([combo.name for combo in self.load_combinations])

        tabs.addTab(loads_tab, "Loads")

        # ---------------- Costing tab ----------------
        costing_tab = QWidget()
        costing_layout = QVBoxLayout()
        costing_tab.setLayout(costing_layout)

        costing_form = QFormLayout()
        self.shs_cost_input = QLineEdit(COSTING_DEFAULTS["shs_rate_per_kg"])
        self.ub_cost_input = QLineEdit(COSTING_DEFAULTS["ub_rate_per_kg"])
        self.wb_cost_input = QLineEdit(COSTING_DEFAULTS["wb_rate_per_kg"])
        self.purlin_cost_input = QLineEdit(COSTING_DEFAULTS["purlin_rate_per_kg"])
        costing_form.addRow("SHS/RHS/CHS rate ($/kg)", self.shs_cost_input)
        costing_form.addRow("UB/UC/PFC rate ($/kg)", self.ub_cost_input)
        costing_form.addRow("WB rate ($/kg)", self.wb_cost_input)
        costing_form.addRow("Purlin/girt rate ($/kg)", self.purlin_cost_input)
        costing_layout.addLayout(costing_form)

        self.costing_box = QTextEdit()
        self.costing_box.setMinimumWidth(520)
        self.costing_box.setReadOnly(True)
        costing_layout.addWidget(self.costing_box)

        tabs.addTab(costing_tab, "Costing")

        # ---------------- Foundations tab ----------------
        foundations_tab = QWidget()
        foundations_layout = QVBoxLayout()
        foundations_tab.setLayout(foundations_layout)

        foundations_form = QFormLayout()
        self.foundation_bearing_input = QLineEdit("150")
        self.foundation_skin_friction_input = QLineEdit("15")
        self.foundation_skin_bearing_checkbox = QCheckBox("Use skin friction for bearing")
        self.foundation_skin_bearing_checkbox.setChecked(False)
        foundations_form.addRow("Allowable bearing (kPa)", self.foundation_bearing_input)
        foundations_form.addRow("Skin friction from 1 m (kPa)", self.foundation_skin_friction_input)
        foundations_form.addRow("Bearing", self.foundation_skin_bearing_checkbox)
        foundations_layout.addLayout(foundations_form)

        self.foundation_box = QTextEdit()
        self.foundation_box.setMinimumWidth(520)
        self.foundation_box.setReadOnly(True)
        foundations_layout.addWidget(self.foundation_box)

        tabs.addTab(foundations_tab, "Foundations")

        # ---------------- Results tab ----------------
        results_tab = QWidget()
        results_layout = QVBoxLayout()
        results_tab.setLayout(results_layout)

        results_form = QFormLayout()
        self.result_combo = QComboBox()
        self.result_combo.addItems(["Geometry", "Load Diagram", "Deflection", "Bending Moment", "Shear Force", "Axial Compression"])
        self.result_combo.currentTextChanged.connect(self.refresh_plot)
        results_form.addRow("Diagram View", self.result_combo)
        self.load_view_mode_combo = QComboBox()
        self.load_view_mode_combo.addItems(["Combination loads", "Single load"])
        self.load_view_mode_combo.currentTextChanged.connect(self.refresh_plot)
        self.single_load_combo = QComboBox()
        self.single_load_combo.addItem("G")
        self.single_load_combo.currentTextChanged.connect(self.refresh_plot)
        results_form.addRow("Load View Mode", self.load_view_mode_combo)
        results_form.addRow("Single Load", self.single_load_combo)
        results_form.addRow("Load Combination", self.load_combination_combo)

        self.second_order_checkbox = QCheckBox("Second-order analysis")
        self.second_order_checkbox.setChecked(False)
        self.second_order_checkbox.setToolTip("Iterate with member geometric stiffness from axial force.")
        results_form.addRow("Analysis", self.second_order_checkbox)

        self.deflection_scale_input = QLineEdit("100")
        self.deflection_scale_input.editingFinished.connect(self.refresh_plot)
        self.diagram_scale_input = QLineEdit("1.0")
        self.diagram_scale_input.editingFinished.connect(self.refresh_plot)
        results_form.addRow("Deflection exaggeration", self.deflection_scale_input)
        results_form.addRow("Force/load diagram scale", self.diagram_scale_input)

        self.panel_label = QLabel("-")
        self.web_angle_label = QLabel("-")
        self.panel_count_label = QLabel("-")
        results_form.addRow("Actual Panel Length", self.panel_label)
        results_form.addRow("Approx Web Angle", self.web_angle_label)
        results_form.addRow("Total Panels", self.panel_count_label)
        results_layout.addLayout(results_form)

        self.results_box = QTextEdit()
        self.results_box.setMinimumWidth(520)
        results_layout.addWidget(self.results_box)

        tabs.addTab(results_tab, "Results")

        # ---------------- Section checks tab ----------------
        section_checks_tab = QWidget()
        section_checks_layout = QVBoxLayout()
        section_checks_tab.setLayout(section_checks_layout)

        self.section_checks_box = QTextEdit()
        self.section_checks_box.setMinimumWidth(520)
        section_checks_layout.addWidget(self.section_checks_box)

        tabs.addTab(section_checks_tab, "Section Checks")

        preset_buttons = QHBoxLayout()
        save_preset_button = QPushButton("Save")
        save_as_preset_button = QPushButton("Save As")
        load_preset_button = QPushButton("Load")
        save_preset_button.clicked.connect(self.save_preset)
        save_as_preset_button.clicked.connect(self.save_preset_as)
        load_preset_button.clicked.connect(self.load_preset)
        preset_buttons.addWidget(save_preset_button)
        preset_buttons.addWidget(save_as_preset_button)
        preset_buttons.addWidget(load_preset_button)
        left_layout.addLayout(preset_buttons)

        solve_button = QPushButton("Generate + Solve")
        solve_button.clicked.connect(self.generate_and_solve)
        left_layout.addWidget(solve_button)
        export_3d_button = QPushButton("Export 3D Concept")
        export_3d_button.clicked.connect(self.export_3d_concept)
        left_layout.addWidget(export_3d_button)

        self.canvas = PlotCanvas()
        left_panel.setMinimumWidth(680)
        layout.addWidget(left_panel, 2)
        layout.addWidget(self.canvas, 3)

        self.update_web_section_visibility()
        self.update_internal_column_visibility()
        self.update_current_canopy_visibility()
        self.update_snow_visibility()
        self.update_frame_system_visibility()
        self.update_eave_restraint_visibility()
        self.update_license_status_label()
        self.update_snow_visibility()
        self.internal_column_count_combo.currentTextChanged.connect(lambda _: self.update_internal_column_visibility())
        self.frame_system_combo.currentTextChanged.connect(lambda _: self.update_frame_system_visibility())
        self.eave_x_restraint_combo.currentTextChanged.connect(lambda _: self.update_eave_restraint_visibility())
        for widget in [self.left_canopy_length_input, self.right_canopy_length_input]:
            widget.editingFinished.connect(self.update_current_canopy_visibility)
        self.left_canopy_type_combo.currentTextChanged.connect(lambda _: self.update_current_canopy_visibility())
        self.right_canopy_type_combo.currentTextChanged.connect(lambda _: self.update_current_canopy_visibility())
        self.snow_region_combo.currentTextChanged.connect(lambda _: self.update_snow_visibility())
        for button in self.envelope_plan_selector.all_wall_buttons():
            button.clicked.connect(self.update_derived_frame_type_label)
        self.frame_type_combo.currentTextChanged.connect(lambda _: self.apply_legacy_frame_type_to_plan())
        self.update_derived_frame_type_label()

        if not shs_profiles:
            QMessageBox.warning(self, "Catalog warning", "No SHS profiles were found in the catalog. Check the naming/group fields.")

    def set_default_combo(self, combo, preferred_fragments):
        for fragment in preferred_fragments:
            fragment = fragment.upper()
            for i in range(combo.count()):
                if fragment in combo.itemText(i).upper():
                    combo.setCurrentIndex(i)
                    return

    def selected_profile(self, combo):
        return self.catalog.get(combo.currentText())

    def bay_size_m_from_ui(self):
        return float(self.load_bay_size_input.text()) / 1000.0

    def downward_roof_load_from_input(self, widget):
        return -abs(float(widget.text()))

    def downward_roof_load_text(self, value):
        try:
            return str(abs(float(value)))
        except (TypeError, ValueError):
            return str(value)

    def optional_height_from_input(self, widget):
        text = widget.text().strip()
        if not text:
            return None
        return max(float(text), 0.0)

    def side_addition_selected(self, combo):
        return combo.currentText() in {"Canopy", "Lean-to"}

    def frame_type_from_wall_states(self, states=None):
        if getattr(self, "frame_type_combo", None) is not None and self.frame_type_combo.currentText() == "No Wind":
            return "No Wind"
        states = states or self.envelope_plan_selector.wall_states()
        clad_count = sum(1 for clad in states.values() if clad)
        if clad_count == 4:
            return "Enclosed"
        if clad_count == 0:
            return "Roof Only"
        if clad_count == 3:
            return "3 Sided"
        if clad_count == 2:
            return "2 Sided"
        return "1 Sided"

    def wall_states_from_legacy_frame_type(self, frame_type):
        if frame_type == "Enclosed":
            return {"left": True, "right": True, "front": True, "back": True}
        if frame_type in {"Roof Only", "No Wind"}:
            return {"left": False, "right": False, "front": False, "back": False}
        if frame_type == "3 Sided":
            return {"left": True, "right": True, "front": False, "back": True}
        if frame_type == "2 Sided":
            return {"left": True, "right": False, "front": False, "back": True}
        if frame_type == "1 Sided":
            return {"left": True, "right": False, "front": False, "back": False}
        return {"left": True, "right": True, "front": True, "back": True}

    def update_derived_frame_type_label(self):
        if getattr(self, "envelope_plan_selector", None) is not None and hasattr(self.envelope_plan_selector, "sync_inner_walls"):
            self.envelope_plan_selector.sync_inner_walls()
        if getattr(self, "derived_frame_type_label", None) is not None:
            self.derived_frame_type_label.setText(self.frame_type_from_wall_states())

    def apply_legacy_frame_type_to_plan(self):
        if getattr(self, "envelope_plan_selector", None) is not None:
            self.envelope_plan_selector.set_wall_states(
                self.wall_states_from_legacy_frame_type(self.frame_type_combo.currentText())
            )
        self.update_derived_frame_type_label()

    def round_up_to_increment(self, value, increment):
        if increment <= 0.0:
            return value
        return math.ceil(value / increment) * increment

    def truss_depth_from_span(self, ratio=0.03):
        span_mm = float(self.span_input.text())
        return self.round_up_to_increment(ratio * span_mm, 50.0)

    def update_truss_depth_from_span(self):
        try:
            depth_mm = self.truss_depth_from_span(0.03)
        except ValueError:
            return
        if depth_mm <= 0.0:
            return
        self.depth_input.setText(f"{depth_mm:.0f}")
        self.generate_and_solve()

    def snow_kp_for_importance(self, importance_level):
        return {
            1: 1.4,
            2: 1.5,
            3: 1.65,
            4: 1.85,
        }.get(int(importance_level), 1.5)

    def snow_loads_from_ui(self):
        region = self.snow_region_combo.currentText()
        if region == "None":
            return {
                "region": region,
                "enabled": False,
                "k1": 0.0,
                "kp": 0.0,
                "mu": 0.0,
                "sg_kpa": 0.0,
                "ultimate_kpa": 0.0,
                "service_kpa": 0.0,
            }
        importance = int(self.importance_combo.currentText())
        h0_m = max(float(self.snow_ahd_height_input.text()), 0.0)
        pitch_deg = float(self.pitch_input.text())
        k1 = SNOW_REGION_K1.get(region, 0.0)
        kp = self.snow_kp_for_importance(importance)
        mu = min(0.7 * (60.0 - pitch_deg) / 50.0, 0.7)
        mu = max(mu, 0.0)
        sg_kpa = max(kp * k1 * ((2.8 * h0_m) / 1000.0 - 1.2), 0.0)
        ultimate_kpa = mu * sg_kpa
        service_kpa = ultimate_kpa / 1.5
        return {
            "region": region,
            "enabled": True,
            "importance": importance,
            "h0_m": h0_m,
            "pitch_deg": pitch_deg,
            "k1": k1,
            "kp": kp,
            "mu": mu,
            "sg_kpa": sg_kpa,
            "ultimate_kpa": ultimate_kpa,
            "service_kpa": service_kpa,
        }

    def roof_load_placement_from_ui(self, prefix):
        return RoofLoadPlacement(
            side=getattr(self, f"{prefix}_load_side_combo").currentText(),
            from_percent=float(getattr(self, f"{prefix}_load_from_input").text()),
            to_percent=float(getattr(self, f"{prefix}_load_to_input").text()),
        )

    def roof_load_placement_data(self, prefix):
        return {
            "side": getattr(self, f"{prefix}_load_side_combo").currentText(),
            "from_percent": getattr(self, f"{prefix}_load_from_input").text(),
            "to_percent": getattr(self, f"{prefix}_load_to_input").text(),
        }

    def apply_roof_load_placement_data(self, prefix, data):
        data = data or {}
        self.set_combo_text_if_present(getattr(self, f"{prefix}_load_side_combo"), data.get("side"))
        getattr(self, f"{prefix}_load_from_input").setText(str(data.get("from_percent", getattr(self, f"{prefix}_load_from_input").text())))
        getattr(self, f"{prefix}_load_to_input").setText(str(data.get("to_percent", getattr(self, f"{prefix}_load_to_input").text())))

    def section_selector_row(self, combo):
        row = QWidget()
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        row.setLayout(layout)
        down_button = QToolButton()
        up_button = QToolButton()
        down_button.setText("↓")
        up_button.setText("↑")
        down_button.setToolTip("Move one section smaller and solve")
        up_button.setToolTip("Move one section larger and solve")
        down_button.clicked.connect(lambda: self.step_section_combo(combo, -1))
        up_button.clicked.connect(lambda: self.step_section_combo(combo, +1))
        layout.addWidget(combo, 1)
        layout.addWidget(down_button)
        layout.addWidget(up_button)
        return row

    def depth_selector_row(self):
        return self.stepped_depth_row(
            self.depth_input,
            lambda delta: self.step_depth_input(self.depth_input, delta),
            "truss depth",
        )

    def canopy_depth_selector_row(self, depth_input):
        return self.stepped_depth_row(
            depth_input,
            lambda delta: self.step_depth_input(depth_input, delta),
            "canopy/lean-to depth",
        )

    def stepped_depth_row(self, depth_input, step_callback, label):
        row = QWidget()
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        row.setLayout(layout)
        down_button = QToolButton()
        up_button = QToolButton()
        down_button.setText("-50")
        up_button.setText("+50")
        down_button.setToolTip(f"Reduce {label} by 50 mm and solve")
        up_button.setToolTip(f"Increase {label} by 50 mm and solve")
        down_button.clicked.connect(lambda: step_callback(-50.0))
        up_button.clicked.connect(lambda: step_callback(50.0))
        layout.addWidget(depth_input, 1)
        layout.addWidget(down_button)
        layout.addWidget(up_button)
        return row

    def roof_load_placement_row(self, prefix):
        row = QWidget()
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        row.setLayout(layout)
        side_combo = QComboBox()
        side_combo.addItems(["Both", "Left", "Right"])
        from_input = QLineEdit("0")
        to_input = QLineEdit("100")
        from_input.setMaximumWidth(55)
        to_input.setMaximumWidth(55)
        setattr(self, f"{prefix}_load_side_combo", side_combo)
        setattr(self, f"{prefix}_load_from_input", from_input)
        setattr(self, f"{prefix}_load_to_input", to_input)
        layout.addWidget(side_combo)
        layout.addWidget(QLabel("from"))
        layout.addWidget(from_input)
        layout.addWidget(QLabel("to"))
        layout.addWidget(to_input)
        layout.addWidget(QLabel("%"))
        return row

    def step_section_combo(self, combo, direction):
        next_index = combo.currentIndex() + direction
        if next_index < 0 or next_index >= combo.count():
            return
        combo.setCurrentIndex(next_index)
        self.generate_and_solve()

    def step_truss_depth(self, delta_mm):
        self.step_depth_input(self.depth_input, delta_mm)

    def step_depth_input(self, depth_input, delta_mm):
        try:
            current = float(depth_input.text())
        except ValueError:
            current = 0.0
        next_depth = max(50.0, current + delta_mm)
        depth_input.setText(f"{next_depth:.0f}")
        self.generate_and_solve()

    def default_preset_dir(self):
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "presets")
        os.makedirs(path, exist_ok=True)
        return path

    def default_settings_path(self):
        return os.path.join(self.default_preset_dir(), "portal_frame_defaults.json")

    def collect_preset_data(self):
        return {
            "version": 1,
            "geometry": {
                "span_mm": self.span_input.text(),
                "roof_pitch_deg": self.pitch_input.text(),
                "eave_height_mm": self.eave_input.text(),
                "truss_depth_mm": self.depth_input.text(),
                "frame_system": self.frame_system_combo.currentText(),
                "rafter_haunch_length_mm": self.rafter_haunch_length_input.text(),
                "internal_column_count": self.internal_column_count_combo.currentText(),
                "internal_column_offset_mm": self.internal_column_offset_input.text(),
                "left_support": self.left_support_combo.currentText(),
                "right_support": self.right_support_combo.currentText(),
                "eave_x_restraint": self.eave_x_restraint_combo.currentText(),
                "eave_x_spring_kn_mm": self.eave_x_spring_input.text(),
                "end_wall_brace_dia_mm": self.end_wall_brace_dia_input.text(),
            },
            "canopies": {
                "left_length_mm": self.left_canopy_length_input.text(),
                "left_type": self.left_canopy_type_combo.currentText(),
                "left_wind_model": self.left_canopy_wind_model_combo.currentText(),
                "left_underside": self.left_canopy_underside_combo.currentText(),
                "left_cpn_uplift": self.left_canopy_cpn_uplift_input.text(),
                "left_cpn_downward": self.left_canopy_cpn_downward_input.text(),
                "left_lean_outer_wall_clad": self.left_lean_outer_wall_clad_combo.currentText(),
                "left_lean_enclosure": self.left_lean_enclosure_selector.wall_states(),
                "left_eave_height_mm": self.left_canopy_eave_height_input.text(),
                "left_pitch_deg": self.left_canopy_pitch_input.text(),
                "left_shallow_depth_mm": self.left_canopy_shallow_depth_input.text(),
                "left_g_kpa": self.left_canopy_g_load_input.text(),
                "left_q_kpa": self.left_canopy_q_load_input.text(),
                "left_solar_kpa": self.left_canopy_solar_load_input.text(),
                "right_length_mm": self.right_canopy_length_input.text(),
                "right_type": self.right_canopy_type_combo.currentText(),
                "right_wind_model": self.right_canopy_wind_model_combo.currentText(),
                "right_underside": self.right_canopy_underside_combo.currentText(),
                "right_cpn_uplift": self.right_canopy_cpn_uplift_input.text(),
                "right_cpn_downward": self.right_canopy_cpn_downward_input.text(),
                "right_lean_outer_wall_clad": self.right_lean_outer_wall_clad_combo.currentText(),
                "right_lean_enclosure": self.right_lean_enclosure_selector.wall_states(),
                "right_eave_height_mm": self.right_canopy_eave_height_input.text(),
                "right_pitch_deg": self.right_canopy_pitch_input.text(),
                "right_shallow_depth_mm": self.right_canopy_shallow_depth_input.text(),
                "right_g_kpa": self.right_canopy_g_load_input.text(),
                "right_q_kpa": self.right_canopy_q_load_input.text(),
                "right_solar_kpa": self.right_canopy_solar_load_input.text(),
            },
            "sections": {
                "left_column": self.left_column_profile_combo.currentText(),
                "right_column": self.right_column_profile_combo.currentText(),
                "internal_column": self.internal_column_profile_combo.currentText(),
                "top_chord": self.top_profile_combo.currentText(),
                "bottom_chord": self.top_profile_combo.currentText(),
                "purlin_span_type": self.purlin_span_type_combo.currentText(),
                "wall_girt_span_type": self.wall_girt_span_type_combo.currentText(),
                "webs_posts": self.web_profile_combo.currentText(),
                "left_canopy_top_chord": self.left_canopy_top_profile_combo.currentText(),
                "left_canopy_bottom_chord": self.left_canopy_bottom_profile_combo.currentText(),
                "left_canopy_webs_posts": self.left_canopy_web_profile_combo.currentText(),
                "left_lean_to_column": self.left_lean_column_profile_combo.currentText(),
                "right_canopy_top_chord": self.right_canopy_top_profile_combo.currentText(),
                "right_canopy_bottom_chord": self.right_canopy_bottom_profile_combo.currentText(),
                "right_canopy_webs_posts": self.right_canopy_web_profile_combo.currentText(),
                "right_lean_to_column": self.right_lean_column_profile_combo.currentText(),
            },
            "costing": {
                "shs_rate_per_kg": self.shs_cost_input.text(),
                "ub_rate_per_kg": self.ub_cost_input.text(),
                "wb_rate_per_kg": self.wb_cost_input.text(),
                "purlin_rate_per_kg": self.purlin_cost_input.text(),
            },
            "foundations": {
                "allowable_bearing_kpa": self.foundation_bearing_input.text(),
                "skin_friction_kpa": self.foundation_skin_friction_input.text(),
                "use_skin_friction_for_bearing": self.foundation_skin_bearing_checkbox.isChecked(),
            },
            "loads": {
                "bay_size_mm": self.load_bay_size_input.text(),
                "bay_size_m": self.bay_size_m_from_ui(),
                "g_kpa": self.g_load_input.text(),
                "q_kpa": self.q_load_input.text(),
                "solar_kpa": self.solar_load_input.text(),
                "fire_service_kpa": self.fire_service_load_input.text(),
                "hvac_kpa": self.hvac_load_input.text(),
                "other_kpa": self.other_load_input.text(),
                "crane_rating_t": self.crane_rating_input.text(),
                "snow_region": self.snow_region_combo.currentText(),
                "snow_ahd_height_m": self.snow_ahd_height_input.text(),
                "roof_load_placements": {
                    "g": self.roof_load_placement_data("g"),
                    "q": self.roof_load_placement_data("q"),
                    "solar": self.roof_load_placement_data("solar"),
                    "other": self.roof_load_placement_data("other"),
                },
                "left_wall_kn_m": self.left_wall_load_input.text(),
                "right_wall_kn_m": self.right_wall_load_input.text(),
                "load_combination": self.load_combination_combo.currentText(),
            },
            "wind": {
                "region": self.wind_region_combo.currentText(),
                "importance_level": self.importance_combo.currentText(),
                "terrain_category": self.terrain_combo.currentText(),
                "orientation": self.wind_orientation_selector.currentText(),
                "frame_type": self.frame_type_combo.currentText(),
                "wall_enclosure": self.envelope_plan_selector.wall_states(),
                "single_wind_case": self.wind_case_combo.currentText(),
                "building_length_m": self.building_length_input.text(),
                "additional_pressure_factor": self.wind_reduction_input.text(),
            },
            "display": {
                "diagram_view": self.result_combo.currentText(),
                "load_view_mode": self.load_view_mode_combo.currentText(),
                "single_load": self.single_load_combo.currentText(),
                "second_order_analysis": self.second_order_checkbox.isChecked(),
                "deflection_scale": self.deflection_scale_input.text(),
                "diagram_scale": self.diagram_scale_input.text(),
            },
        }

    def set_combo_text_if_present(self, combo, value):
        if value is None:
            return
        index = combo.findText(str(value))
        if index >= 0:
            combo.setCurrentIndex(index)

    def normalized_canopy_wind_model(self, value):
        if value == "Free roof net Cpn":
            return "Manual net Cpn"
        return value

    def normalized_single_load_label(self, value):
        if value is None:
            return value
        text = str(value)
        if text.startswith("Wu Cpe ") or text.startswith("Ws Cpe ") or text.startswith("Wu Cpi ") or text.startswith("Ws Cpi "):
            return text
        wind_prefixes = ["Wu Left", "Wu Right", "Wu End", "Ws Left", "Ws Right", "Ws End"]
        for prefix in wind_prefixes:
            if text.startswith(prefix):
                return text.replace("Wu ", "Wu Cpe ", 1).replace("Ws ", "Ws Cpe ", 1)
        if text.startswith("Cpi "):
            return f"Wu {text}"
        if text.startswith("Ws Cpi "):
            return text
        return text

    def apply_preset_data(self, data):
        geometry = data.get("geometry", {})
        self.span_input.setText(str(geometry.get("span_mm", self.span_input.text())))
        self.pitch_input.setText(str(geometry.get("roof_pitch_deg", self.pitch_input.text())))
        self.eave_input.setText(str(geometry.get("eave_height_mm", self.eave_input.text())))
        self.depth_input.setText(str(geometry.get("truss_depth_mm", self.depth_input.text())))
        self.set_combo_text_if_present(self.frame_system_combo, geometry.get("frame_system"))
        self.rafter_haunch_length_input.setText(str(geometry.get("rafter_haunch_length_mm", self.rafter_haunch_length_input.text())))
        self.set_combo_text_if_present(self.internal_column_count_combo, geometry.get("internal_column_count"))
        self.internal_column_offset_input.setText(str(geometry.get("internal_column_offset_mm", self.internal_column_offset_input.text())))
        self.set_combo_text_if_present(self.left_support_combo, geometry.get("left_support"))
        self.set_combo_text_if_present(self.right_support_combo, geometry.get("right_support"))
        self.set_combo_text_if_present(self.eave_x_restraint_combo, geometry.get("eave_x_restraint"))
        self.eave_x_spring_input.setText(str(geometry.get("eave_x_spring_kn_mm", self.eave_x_spring_input.text())))
        self.end_wall_brace_dia_input.setText(str(geometry.get("end_wall_brace_dia_mm", self.end_wall_brace_dia_input.text())))

        canopies = data.get("canopies", {})
        self.left_canopy_length_input.setText(str(canopies.get("left_length_mm", self.left_canopy_length_input.text())))
        left_type = canopies.get("left_type")
        if left_type is None:
            try:
                left_type = "Canopy" if float(self.left_canopy_length_input.text()) > 0.0 else None
            except ValueError:
                left_type = None
        self.set_combo_text_if_present(self.left_canopy_type_combo, left_type)
        self.set_combo_text_if_present(self.left_canopy_wind_model_combo, self.normalized_canopy_wind_model(canopies.get("left_wind_model")))
        self.set_combo_text_if_present(self.left_canopy_underside_combo, canopies.get("left_underside"))
        self.left_canopy_cpn_uplift_input.setText(str(canopies.get("left_cpn_uplift", self.left_canopy_cpn_uplift_input.text())))
        self.left_canopy_cpn_downward_input.setText(str(canopies.get("left_cpn_downward", self.left_canopy_cpn_downward_input.text())))
        self.set_combo_text_if_present(self.left_lean_outer_wall_clad_combo, canopies.get("left_lean_outer_wall_clad"))
        left_lean_enclosure = canopies.get("left_lean_enclosure")
        if isinstance(left_lean_enclosure, dict):
            self.left_lean_enclosure_selector.set_wall_states(left_lean_enclosure)
        elif canopies.get("left_lean_outer_wall_clad") == "Yes":
            self.left_lean_enclosure_selector.set_wall_states({"left": True})
        self.left_canopy_eave_height_input.setText(str(canopies.get("left_eave_height_mm", canopies.get("eave_height_mm", self.left_canopy_eave_height_input.text()))))
        self.left_canopy_pitch_input.setText(str(canopies.get("left_pitch_deg", canopies.get("pitch_deg", self.left_canopy_pitch_input.text()))))
        self.left_canopy_shallow_depth_input.setText(str(canopies.get("left_shallow_depth_mm", canopies.get("shallow_depth_mm", self.left_canopy_shallow_depth_input.text()))))
        self.left_canopy_g_load_input.setText(self.downward_roof_load_text(canopies.get("left_g_kpa", canopies.get("g_kpa", self.left_canopy_g_load_input.text()))))
        self.left_canopy_q_load_input.setText(self.downward_roof_load_text(canopies.get("left_q_kpa", canopies.get("q_kpa", self.left_canopy_q_load_input.text()))))
        self.left_canopy_solar_load_input.setText(self.downward_roof_load_text(canopies.get("left_solar_kpa", canopies.get("solar_kpa", self.left_canopy_solar_load_input.text()))))
        self.right_canopy_length_input.setText(str(canopies.get("right_length_mm", self.right_canopy_length_input.text())))
        right_type = canopies.get("right_type")
        if right_type is None:
            try:
                right_type = "Canopy" if float(self.right_canopy_length_input.text()) > 0.0 else None
            except ValueError:
                right_type = None
        self.set_combo_text_if_present(self.right_canopy_type_combo, right_type)
        self.set_combo_text_if_present(self.right_canopy_wind_model_combo, self.normalized_canopy_wind_model(canopies.get("right_wind_model")))
        self.set_combo_text_if_present(self.right_canopy_underside_combo, canopies.get("right_underside"))
        self.right_canopy_cpn_uplift_input.setText(str(canopies.get("right_cpn_uplift", self.right_canopy_cpn_uplift_input.text())))
        self.right_canopy_cpn_downward_input.setText(str(canopies.get("right_cpn_downward", self.right_canopy_cpn_downward_input.text())))
        self.set_combo_text_if_present(self.right_lean_outer_wall_clad_combo, canopies.get("right_lean_outer_wall_clad"))
        right_lean_enclosure = canopies.get("right_lean_enclosure")
        if isinstance(right_lean_enclosure, dict):
            self.right_lean_enclosure_selector.set_wall_states(right_lean_enclosure)
        elif canopies.get("right_lean_outer_wall_clad") == "Yes":
            self.right_lean_enclosure_selector.set_wall_states({"right": True})
        self.right_canopy_eave_height_input.setText(str(canopies.get("right_eave_height_mm", canopies.get("eave_height_mm", self.right_canopy_eave_height_input.text()))))
        self.right_canopy_pitch_input.setText(str(canopies.get("right_pitch_deg", canopies.get("pitch_deg", self.right_canopy_pitch_input.text()))))
        self.right_canopy_shallow_depth_input.setText(str(canopies.get("right_shallow_depth_mm", canopies.get("shallow_depth_mm", self.right_canopy_shallow_depth_input.text()))))
        self.right_canopy_g_load_input.setText(self.downward_roof_load_text(canopies.get("right_g_kpa", canopies.get("g_kpa", self.right_canopy_g_load_input.text()))))
        self.right_canopy_q_load_input.setText(self.downward_roof_load_text(canopies.get("right_q_kpa", canopies.get("q_kpa", self.right_canopy_q_load_input.text()))))
        self.right_canopy_solar_load_input.setText(self.downward_roof_load_text(canopies.get("right_solar_kpa", canopies.get("solar_kpa", self.right_canopy_solar_load_input.text()))))

        sections = data.get("sections", {})
        self.set_combo_text_if_present(self.left_column_profile_combo, sections.get("left_column"))
        self.set_combo_text_if_present(self.right_column_profile_combo, sections.get("right_column"))
        self.set_combo_text_if_present(self.internal_column_profile_combo, sections.get("internal_column"))
        self.set_combo_text_if_present(self.top_profile_combo, sections.get("top_chord", sections.get("bottom_chord")))
        self.set_combo_text_if_present(self.purlin_span_type_combo, sections.get("purlin_span_type"))
        self.set_combo_text_if_present(self.wall_girt_span_type_combo, sections.get("wall_girt_span_type"))
        self.set_combo_text_if_present(self.web_profile_combo, sections.get("webs_posts"))
        self.set_combo_text_if_present(
            self.left_canopy_top_profile_combo,
            sections.get("left_canopy_top_chord", sections.get("left_canopy_bottom_chord", sections.get("canopy_top_chord", sections.get("canopy_bottom_chord"))))
        )
        self.set_combo_text_if_present(self.left_canopy_web_profile_combo, sections.get("left_canopy_webs_posts", sections.get("canopy_webs_posts")))
        self.set_combo_text_if_present(self.left_lean_column_profile_combo, sections.get("left_lean_to_column", sections.get("lean_to_column")))
        self.set_combo_text_if_present(
            self.right_canopy_top_profile_combo,
            sections.get("right_canopy_top_chord", sections.get("right_canopy_bottom_chord", sections.get("canopy_top_chord", sections.get("canopy_bottom_chord"))))
        )
        self.set_combo_text_if_present(self.right_canopy_web_profile_combo, sections.get("right_canopy_webs_posts", sections.get("canopy_webs_posts")))
        self.set_combo_text_if_present(self.right_lean_column_profile_combo, sections.get("right_lean_to_column", sections.get("lean_to_column")))

        costing = data.get("costing", {})
        self.shs_cost_input.setText(str(costing.get("shs_rate_per_kg", COSTING_DEFAULTS["shs_rate_per_kg"])))
        self.ub_cost_input.setText(str(costing.get("ub_rate_per_kg", COSTING_DEFAULTS["ub_rate_per_kg"])))
        self.wb_cost_input.setText(str(costing.get("wb_rate_per_kg", COSTING_DEFAULTS["wb_rate_per_kg"])))
        self.purlin_cost_input.setText(str(costing.get("purlin_rate_per_kg", COSTING_DEFAULTS["purlin_rate_per_kg"])))

        foundations = data.get("foundations", {})
        self.foundation_bearing_input.setText(str(foundations.get("allowable_bearing_kpa", self.foundation_bearing_input.text())))
        self.foundation_skin_friction_input.setText(str(foundations.get("skin_friction_kpa", self.foundation_skin_friction_input.text())))
        if foundations.get("use_skin_friction_for_bearing") is not None:
            self.foundation_skin_bearing_checkbox.setChecked(bool(foundations.get("use_skin_friction_for_bearing")))

        loads = data.get("loads", {})
        if loads.get("bay_size_mm") is not None:
            self.load_bay_size_input.setText(str(loads.get("bay_size_mm")))
        elif loads.get("bay_size_m") is not None:
            self.load_bay_size_input.setText(f"{float(loads.get('bay_size_m')) * 1000.0:.0f}")
        self.g_load_input.setText(self.downward_roof_load_text(loads.get("g_kpa", self.g_load_input.text())))
        self.q_load_input.setText(self.downward_roof_load_text(loads.get("q_kpa", self.q_load_input.text())))
        self.solar_load_input.setText(self.downward_roof_load_text(loads.get("solar_kpa", self.solar_load_input.text())))
        self.fire_service_load_input.setText(self.downward_roof_load_text(loads.get("fire_service_kpa", self.fire_service_load_input.text())))
        self.hvac_load_input.setText(self.downward_roof_load_text(loads.get("hvac_kpa", self.hvac_load_input.text())))
        self.other_load_input.setText(self.downward_roof_load_text(loads.get("other_kpa", self.other_load_input.text())))
        self.crane_rating_input.setText(str(loads.get("crane_rating_t", self.crane_rating_input.text())))
        self.set_combo_text_if_present(self.snow_region_combo, loads.get("snow_region"))
        self.snow_ahd_height_input.setText(str(loads.get("snow_ahd_height_m", self.snow_ahd_height_input.text())))
        placements = loads.get("roof_load_placements", {})
        legacy_placement = {
            "side": loads.get("roof_load_side"),
            "from_percent": loads.get("roof_load_from_percent"),
            "to_percent": loads.get("roof_load_to_percent"),
        }
        for prefix in ["g", "q", "solar", "other"]:
            self.apply_roof_load_placement_data(prefix, placements.get(prefix) or legacy_placement)
        self.left_wall_load_input.setText(str(loads.get("left_wall_kn_m", self.left_wall_load_input.text())))
        self.right_wall_load_input.setText(str(loads.get("right_wall_kn_m", self.right_wall_load_input.text())))
        self.set_combo_text_if_present(self.load_combination_combo, loads.get("load_combination"))

        wind = data.get("wind", {})
        self.set_combo_text_if_present(self.wind_region_combo, wind.get("region"))
        self.set_combo_text_if_present(self.importance_combo, wind.get("importance_level"))
        self.set_combo_text_if_present(self.terrain_combo, wind.get("terrain_category"))
        self.wind_orientation_selector.setCurrentText(wind.get("orientation", self.wind_orientation_selector.currentText()))
        self.set_combo_text_if_present(self.frame_type_combo, wind.get("frame_type"))
        wall_enclosure = wind.get("wall_enclosure")
        if wall_enclosure is None:
            wall_enclosure = self.wall_states_from_legacy_frame_type(wind.get("frame_type", self.frame_type_combo.currentText()))
        self.envelope_plan_selector.set_wall_states(wall_enclosure)
        self.set_combo_text_if_present(self.wind_case_combo, wind.get("single_wind_case"))
        self.building_length_input.setText(str(wind.get("building_length_m", self.building_length_input.text())))
        self.wind_reduction_input.setText(str(wind.get("additional_pressure_factor", self.wind_reduction_input.text())))
        self.update_derived_frame_type_label()

        display = data.get("display", {})
        self.set_combo_text_if_present(self.result_combo, display.get("diagram_view"))
        self.set_combo_text_if_present(self.load_view_mode_combo, display.get("load_view_mode"))
        normalized_single_load = self.normalized_single_load_label(display.get("single_load"))
        self.set_combo_text_if_present(self.single_load_combo, normalized_single_load)
        self.pending_single_load_text = normalized_single_load
        if display.get("second_order_analysis") is not None:
            self.second_order_checkbox.setChecked(bool(display.get("second_order_analysis")))
        self.deflection_scale_input.setText(str(display.get("deflection_scale", self.deflection_scale_input.text())))
        self.diagram_scale_input.setText(str(display.get("diagram_scale", self.diagram_scale_input.text())))
        self.update_web_section_visibility()
        self.update_internal_column_visibility()
        self.update_current_canopy_visibility()

    def solve_structure(self, structure):
        if getattr(self, "second_order_checkbox", None) is not None and self.second_order_checkbox.isChecked():
            return structure.solve_second_order()
        d, R, F = structure.solve()
        info = {
            "type": "first_order",
            "converged": True,
            "iterations": 1,
            "first_order_max_abs_mm": max(float(np.max(np.abs(d))), 0.0),
            "second_order_max_abs_mm": max(float(np.max(np.abs(d))), 0.0),
            "displacement_amplification": 1.0,
        }
        return d, R, F, info

    def write_preset_file(self, path):
        path = str(path)
        if not path.lower().endswith(".json"):
            path += ".json"
        with open(path, "w", encoding="utf-8") as file:
            json.dump(self.collect_preset_data(), file, indent=2)
        self.current_preset_path = path
        return path

    def save_preset(self, show_message=True):
        path = self.current_preset_path
        if path:
            try:
                path = self.write_preset_file(path)
            except OSError as exc:
                QMessageBox.critical(self, "Save preset failed", str(exc))
                return False
            if show_message:
                QMessageBox.information(self, "Preset saved", f"Saved preset:\n{path}")
            return True
        return self.save_preset_as(show_message=show_message)

    def save_preset_as(self, show_message=True):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save portal frame preset",
            os.path.join(self.default_preset_dir(), "portal_frame_preset.json"),
            "Portal Frame Preset (*.json);;All Files (*)",
        )
        if not path:
            return False
        try:
            path = self.write_preset_file(path)
        except OSError as exc:
            QMessageBox.critical(self, "Save preset failed", str(exc))
            return False
        if show_message:
            QMessageBox.information(self, "Preset saved", f"Saved preset:\n{path}")
        return True

    def load_preset(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Load portal frame preset",
            self.default_preset_dir(),
            "Portal Frame Preset (*.json);;All Files (*)",
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as file:
                data = json.load(file)
            self.apply_preset_data(data)
            self.current_preset_path = path
            self.generate_and_solve()
        except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
            QMessageBox.critical(self, "Load preset failed", str(exc))
            return

    def closeEvent(self, event):
        if not getattr(self, "confirm_on_close", True):
            event.accept()
            return
        response = QMessageBox.question(
            self,
            "Save before closing?",
            "Do you want to save this preset before closing?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Save,
        )
        if response == QMessageBox.Save:
            if self.save_preset(show_message=False):
                event.accept()
            else:
                event.ignore()
            return
        if response == QMessageBox.Cancel:
            event.ignore()
            return
        event.accept()

    def save_default_inputs(self):
        path = self.default_settings_path()
        try:
            with open(path, "w", encoding="utf-8") as file:
                json.dump(self.collect_preset_data(), file, indent=2)
        except OSError as exc:
            QMessageBox.critical(self, "Save defaults failed", str(exc))
            return
        QMessageBox.information(self, "Defaults saved", f"Saved default inputs:\n{path}")

    def load_default_inputs(self, silent=False):
        path = self.default_settings_path()
        if not os.path.exists(path):
            return False
        try:
            with open(path, "r", encoding="utf-8") as file:
                data = json.load(file)
            self.apply_preset_data(data)
        except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
            if not silent:
                QMessageBox.critical(self, "Load defaults failed", str(exc))
            return False
        return True

    def reset_default_inputs(self):
        if self.factory_default_data:
            self.apply_preset_data(self.factory_default_data)
            self.generate_and_solve()
        path = self.default_settings_path()
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError as exc:
            QMessageBox.critical(self, "Reset defaults failed", str(exc))
            return
        QMessageBox.information(self, "Defaults reset", "Default inputs have been reset to the built-in values.")

    def concept_3d_data(self):
        if self.structure is None or not self.structure.nodes:
            raise ValueError("Generate + Solve before exporting a 3D concept.")

        length_m = max(float(getattr(self, "concept_building_length_m", 0.0) or self.building_length_input.text()), 0.1)
        bay_size_m = max(float(getattr(self, "concept_bay_size_m", 0.0) or self.bay_size_m_from_ui()), 0.1)
        bay_count = max(1, math.ceil(length_m / bay_size_m))
        z_positions = [length_m * i / bay_count for i in range(bay_count + 1)]
        nodes = self.structure.nodes
        node_index = {id(node): index for index, node in enumerate(nodes)}

        frame_lines = []
        for element in self.structure.elements:
            frame_lines.append({
                "a": node_index[id(element.start)],
                "b": node_index[id(element.end)],
                "group": element.group,
            })

        longitudinal_nodes = sorted(
            {line["a"] for line in frame_lines}.union({line["b"] for line in frame_lines}),
            key=lambda index: (nodes[index].x, nodes[index].y),
        )
        roof_indices = [node_index[id(node)] for node in getattr(self, "concept_roof_top_nodes", []) if id(node) in node_index]
        main_top_indices = [node_index[id(node)] for node in getattr(self, "concept_main_top_nodes", []) if id(node) in node_index]
        wall_states = self.envelope_plan_selector.wall_states()

        return {
            "title": "Portal frame 3D concept",
            "length_m": length_m,
            "bay_count": bay_count,
            "z_positions": z_positions,
            "nodes": [{"x": node.x / 1000.0, "y": node.y / 1000.0} for node in nodes],
            "frame_lines": frame_lines,
            "longitudinal_nodes": longitudinal_nodes,
            "roof_indices": roof_indices,
            "main_top_indices": main_top_indices,
            "wall_states": wall_states,
        }

    def concept_3d_html(self, data):
        payload = json.dumps(data)
        title = html.escape(data.get("title", "3D concept"))
        return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <style>
    html, body {{ margin: 0; height: 100%; overflow: hidden; font-family: Arial, sans-serif; background: #f4f1ec; color: #202020; }}
    #toolbar {{ position: fixed; left: 16px; top: 14px; background: rgba(255,255,255,.86); border: 1px solid #c8c2b7; padding: 10px 12px; border-radius: 6px; box-shadow: 0 2px 12px rgba(0,0,0,.12); }}
    #toolbar strong {{ display: block; margin-bottom: 4px; }}
    #toolbar span {{ font-size: 12px; color: #555; }}
    canvas {{ display: block; width: 100vw; height: 100vh; cursor: grab; }}
    canvas:active {{ cursor: grabbing; }}
  </style>
</head>
<body>
<canvas id="view"></canvas>
<div id="toolbar"><strong>{title}</strong><span>Drag to orbit, wheel to zoom. Concept geometry only.</span></div>
<script>
const data = {payload};
const canvas = document.getElementById('view');
const ctx = canvas.getContext('2d');
let yaw = -0.72, pitch = 0.42, zoom = 62;
let dragging = false, lastX = 0, lastY = 0;

function resize() {{
  canvas.width = Math.max(1, window.innerWidth * devicePixelRatio);
  canvas.height = Math.max(1, window.innerHeight * devicePixelRatio);
  draw();
}}
window.addEventListener('resize', resize);
canvas.addEventListener('pointerdown', e => {{ dragging = true; lastX = e.clientX; lastY = e.clientY; canvas.setPointerCapture(e.pointerId); }});
canvas.addEventListener('pointermove', e => {{
  if (!dragging) return;
  yaw += (e.clientX - lastX) * 0.008;
  pitch = Math.max(-1.25, Math.min(1.25, pitch + (e.clientY - lastY) * 0.006));
  lastX = e.clientX; lastY = e.clientY; draw();
}});
canvas.addEventListener('pointerup', () => dragging = false);
canvas.addEventListener('wheel', e => {{ e.preventDefault(); zoom *= Math.exp(-e.deltaY * 0.001); zoom = Math.max(15, Math.min(220, zoom)); draw(); }}, {{passive:false}});

function bounds() {{
  const xs = data.nodes.map(n => n.x), ys = data.nodes.map(n => n.y);
  return {{ cx: (Math.min(...xs) + Math.max(...xs)) / 2, cy: (Math.min(...ys) + Math.max(...ys)) / 2, cz: data.length_m / 2 }};
}}
const b = bounds();

function project(p) {{
  let x = p.x - b.cx, y = p.y - b.cy, z = p.z - b.cz;
  const cy = Math.cos(yaw), sy = Math.sin(yaw);
  const cp = Math.cos(pitch), sp = Math.sin(pitch);
  const x1 = x * cy - z * sy;
  const z1 = x * sy + z * cy;
  const y1 = y * cp - z1 * sp;
  const z2 = y * sp + z1 * cp;
  const perspective = 900 / (900 + z2 * zoom * 0.08);
  return {{
    x: canvas.width / 2 + x1 * zoom * perspective * devicePixelRatio,
    y: canvas.height / 2 - y1 * zoom * perspective * devicePixelRatio,
    depth: z2
  }};
}}

function drawLine(a, b, color, width=1.5) {{
  const pa = project(a), pb = project(b);
  ctx.strokeStyle = color;
  ctx.lineWidth = width * devicePixelRatio;
  ctx.beginPath();
  ctx.moveTo(pa.x, pa.y);
  ctx.lineTo(pb.x, pb.y);
  ctx.stroke();
}}

function drawPoly(points, color) {{
  if (points.length < 3) return;
  const ps = points.map(project);
  ctx.fillStyle = color;
  ctx.beginPath();
  ctx.moveTo(ps[0].x, ps[0].y);
  for (const p of ps.slice(1)) ctx.lineTo(p.x, p.y);
  ctx.closePath();
  ctx.fill();
}}

function point(nodeIndex, z) {{
  const n = data.nodes[nodeIndex];
  return {{x:n.x, y:n.y, z:z}};
}}

function drawCladding() {{
  const z0 = data.z_positions[0], z1 = data.z_positions[data.z_positions.length - 1];
  if (data.roof_indices.length >= 2) {{
    const ridge = data.roof_indices.reduce((best, i) => data.nodes[i].y > data.nodes[best].y ? i : best, data.roof_indices[0]);
    const ridgePos = data.roof_indices.indexOf(ridge);
    const left = data.roof_indices.slice(0, ridgePos + 1);
    const right = data.roof_indices.slice(ridgePos);
    for (const run of [left, right]) {{
      for (let i = 0; i < run.length - 1; i++) {{
        drawPoly([point(run[i], z0), point(run[i+1], z0), point(run[i+1], z1), point(run[i], z1)], 'rgba(108, 148, 166, .18)');
      }}
    }}
  }}
  if (data.wall_states.left && data.main_top_indices.length) {{
    const i = data.main_top_indices[0], n = data.nodes[i];
    drawPoly([{{x:n.x,y:0,z:z0}}, {{x:n.x,y:n.y,z:z0}}, {{x:n.x,y:n.y,z:z1}}, {{x:n.x,y:0,z:z1}}], 'rgba(194, 113, 78, .18)');
  }}
  if (data.wall_states.right && data.main_top_indices.length) {{
    const i = data.main_top_indices[data.main_top_indices.length-1], n = data.nodes[i];
    drawPoly([{{x:n.x,y:0,z:z0}}, {{x:n.x,y:n.y,z:z0}}, {{x:n.x,y:n.y,z:z1}}, {{x:n.x,y:0,z:z1}}], 'rgba(194, 113, 78, .18)');
  }}
  if (data.wall_states.front || data.wall_states.back) {{
    const zs = [];
    if (data.wall_states.front) zs.push(z0);
    if (data.wall_states.back) zs.push(z1);
    for (const z of zs) {{
      for (let i = 0; i < data.main_top_indices.length - 1; i++) {{
        const a = point(data.main_top_indices[i], z), c = point(data.main_top_indices[i+1], z);
        drawPoly([{{x:a.x,y:0,z:z}}, a, c, {{x:c.x,y:0,z:z}}], 'rgba(194, 113, 78, .12)');
      }}
    }}
  }}
}}

function draw() {{
  ctx.clearRect(0, 0, canvas.width, canvas.height);
  ctx.fillStyle = '#f4f1ec';
  ctx.fillRect(0, 0, canvas.width, canvas.height);
  drawCladding();
  for (const z of data.z_positions) {{
    for (const line of data.frame_lines) {{
      const major = ['TOP','BOTTOM','LEFT_COLUMN','RIGHT_COLUMN','INTERNAL_COLUMN','COLUMN'].includes(line.group);
      drawLine(point(line.a,z), point(line.b,z), major ? '#263238' : '#59656b', major ? 2.2 : 1.1);
    }}
  }}
  for (let zi = 0; zi < data.z_positions.length - 1; zi++) {{
    for (const i of data.longitudinal_nodes) {{
      drawLine(point(i, data.z_positions[zi]), point(i, data.z_positions[zi+1]), '#829099', .8);
    }}
  }}
}}
resize();
</script>
</body>
</html>
"""

    def concept_3d_dae(self, data):
        vertices = []
        triangles = []

        def add_vertex(point):
            vertices.append(point)
            return len(vertices) - 1

        def add_triangle(a, b, c):
            triangles.append((add_vertex(a), add_vertex(b), add_vertex(c)))

        def add_quad(a, b, c, d):
            add_triangle(a, b, c)
            add_triangle(a, c, d)

        def sub(a, b):
            return (a[0] - b[0], a[1] - b[1], a[2] - b[2])

        def cross(a, b):
            return (
                a[1] * b[2] - a[2] * b[1],
                a[2] * b[0] - a[0] * b[2],
                a[0] * b[1] - a[1] * b[0],
            )

        def length(v):
            return math.sqrt(v[0] * v[0] + v[1] * v[1] + v[2] * v[2])

        def normalize(v):
            size = length(v)
            if size <= 1e-12:
                return (0.0, 0.0, 0.0)
            return (v[0] / size, v[1] / size, v[2] / size)

        def scale(v, factor):
            return (v[0] * factor, v[1] * factor, v[2] * factor)

        def add(a, b):
            return (a[0] + b[0], a[1] + b[1], a[2] + b[2])

        def point(node_index, z_m):
            node = data["nodes"][node_index]
            # Collada is exported Z-up for SketchUp: X = frame width,
            # Y = building length, Z = height.
            return (node["x"], z_m, node["y"])

        def add_prism(a, b, thickness):
            axis = sub(b, a)
            if length(axis) <= 1e-9:
                return
            direction = normalize(axis)
            up = (0.0, 0.0, 1.0)
            side = normalize(cross(direction, up))
            if length(side) <= 1e-9:
                side = (1.0, 0.0, 0.0)
            other = normalize(cross(direction, side))
            half = thickness / 2.0
            offsets = [
                add(scale(side, -half), scale(other, -half)),
                add(scale(side, half), scale(other, -half)),
                add(scale(side, half), scale(other, half)),
                add(scale(side, -half), scale(other, half)),
            ]
            a_pts = [add(a, offset) for offset in offsets]
            b_pts = [add(b, offset) for offset in offsets]
            for i in range(4):
                add_quad(a_pts[i], a_pts[(i + 1) % 4], b_pts[(i + 1) % 4], b_pts[i])
            add_quad(a_pts[3], a_pts[2], a_pts[1], a_pts[0])
            add_quad(b_pts[0], b_pts[1], b_pts[2], b_pts[3])

        def add_cladding():
            z0 = data["z_positions"][0]
            z1 = data["z_positions"][-1]
            roof_indices = data.get("roof_indices", [])
            if len(roof_indices) >= 2:
                for a, b in zip(roof_indices[:-1], roof_indices[1:]):
                    add_quad(point(a, z0), point(b, z0), point(b, z1), point(a, z1))

            main_top = data.get("main_top_indices", [])
            wall_states = data.get("wall_states", {})
            if wall_states.get("left") and main_top:
                top = data["nodes"][main_top[0]]
                x = top["x"]
                add_quad((x, z0, 0.0), (x, z0, top["y"]), (x, z1, top["y"]), (x, z1, 0.0))
            if wall_states.get("right") and main_top:
                top = data["nodes"][main_top[-1]]
                x = top["x"]
                add_quad((x, z1, 0.0), (x, z1, top["y"]), (x, z0, top["y"]), (x, z0, 0.0))
            for wall_key, z in [("front", z0), ("back", z1)]:
                if not wall_states.get(wall_key) or len(main_top) < 2:
                    continue
                for a, b in zip(main_top[:-1], main_top[1:]):
                    pa = point(a, z)
                    pb = point(b, z)
                    add_quad((pa[0], z, 0.0), pa, pb, (pb[0], z, 0.0))

        add_cladding()
        member_thickness = 0.08
        longitudinal_thickness = 0.035
        for z_m in data["z_positions"]:
            for line in data["frame_lines"]:
                group = str(line.get("group", ""))
                thickness = 0.12 if group in {"TOP", "BOTTOM", "LEFT_COLUMN", "RIGHT_COLUMN", "INTERNAL_COLUMN"} else member_thickness
                add_prism(point(line["a"], z_m), point(line["b"], z_m), thickness)
        for z_a, z_b in zip(data["z_positions"][:-1], data["z_positions"][1:]):
            for node_index in data["longitudinal_nodes"]:
                add_prism(point(node_index, z_a), point(node_index, z_b), longitudinal_thickness)

        positions = " ".join(f"{coord:.6f}" for vertex in vertices for coord in vertex)
        indices = " ".join(str(index) for tri in triangles for index in tri)
        vertex_count = len(vertices)
        triangle_count = len(triangles)
        title = html.escape(data.get("title", "PortalCalc 3D concept"))
        return f"""<?xml version="1.0" encoding="utf-8"?>
<COLLADA xmlns="http://www.collada.org/2005/11/COLLADASchema" version="1.4.1">
  <asset>
    <contributor><authoring_tool>PortalCalc</authoring_tool></contributor>
    <unit name="meter" meter="1"/>
    <up_axis>Z_UP</up_axis>
  </asset>
  <library_effects>
    <effect id="portalcalc-effect">
      <profile_COMMON>
        <technique sid="common">
          <lambert>
            <diffuse><color>0.42 0.56 0.62 1</color></diffuse>
          </lambert>
        </technique>
      </profile_COMMON>
    </effect>
  </library_effects>
  <library_materials>
    <material id="portalcalc-material" name="PortalCalc Material">
      <instance_effect url="#portalcalc-effect"/>
    </material>
  </library_materials>
  <library_geometries>
    <geometry id="portalcalc-geometry" name="{title}">
      <mesh>
        <source id="portalcalc-positions">
          <float_array id="portalcalc-positions-array" count="{vertex_count * 3}">{positions}</float_array>
          <technique_common>
            <accessor source="#portalcalc-positions-array" count="{vertex_count}" stride="3">
              <param name="X" type="float"/>
              <param name="Y" type="float"/>
              <param name="Z" type="float"/>
            </accessor>
          </technique_common>
        </source>
        <vertices id="portalcalc-vertices">
          <input semantic="POSITION" source="#portalcalc-positions"/>
        </vertices>
        <triangles material="portalcalc-material" count="{triangle_count}">
          <input semantic="VERTEX" source="#portalcalc-vertices" offset="0"/>
          <p>{indices}</p>
        </triangles>
      </mesh>
    </geometry>
  </library_geometries>
  <library_visual_scenes>
    <visual_scene id="Scene" name="Scene">
      <node id="PortalCalcConcept" name="PortalCalc Concept">
        <instance_geometry url="#portalcalc-geometry">
          <bind_material>
            <technique_common>
              <instance_material symbol="portalcalc-material" target="#portalcalc-material"/>
            </technique_common>
          </bind_material>
        </instance_geometry>
      </node>
    </visual_scene>
  </library_visual_scenes>
  <scene>
    <instance_visual_scene url="#Scene"/>
  </scene>
</COLLADA>
"""

    def export_3d_concept(self):
        try:
            data = self.concept_3d_data()
        except (TypeError, ValueError) as exc:
            QMessageBox.warning(self, "3D concept export", str(exc))
            return
        default_path = os.path.join(os.getcwd(), "portal_3d_concept.dae")
        path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export 3D concept",
            default_path,
            "SketchUp Collada (*.dae);;HTML files (*.html);;All files (*.*)",
        )
        if not path:
            return
        if not os.path.splitext(path)[1]:
            path += ".dae" if selected_filter.startswith("SketchUp") else ".html"
        try:
            extension = os.path.splitext(path)[1].lower()
            content = self.concept_3d_dae(data) if extension == ".dae" else self.concept_3d_html(data)
            with open(path, "w", encoding="utf-8") as file:
                file.write(content)
        except OSError as exc:
            QMessageBox.critical(self, "3D concept export failed", str(exc))
            return
        QMessageBox.information(self, "3D concept exported", f"Exported:\n{path}")

    def apply_base_support(self, structure, node, support_type):
        structure.add_support(node.ux)
        structure.add_support(node.uy)
        if support_type == "Fixed":
            structure.add_support(node.rz)

    def column_cladding_restraints(self, frame_type):
        if hasattr(frame_type, "left_wall_clad") and hasattr(frame_type, "right_wall_clad"):
            return {
                "LEFT_COLUMN": bool(frame_type.left_wall_clad),
                "RIGHT_COLUMN": bool(frame_type.right_wall_clad),
            }
        if frame_type == "Enclosed":
            return {"LEFT_COLUMN": True, "RIGHT_COLUMN": True}
        if frame_type in {"3 Sided", "2 Sided", "1 Sided"}:
            return {"LEFT_COLUMN": True, "RIGHT_COLUMN": False}
        return {"LEFT_COLUMN": False, "RIGHT_COLUMN": False}

    def purlin_mid_span_limit_mm(self, wind_region, importance_level):
        if str(wind_region).upper() == "C":
            return 1100.0
        if int(importance_level) >= 2:
            return 1600.0
        return 1700.0

    def wall_girt_spacing_mm(self, wind_region):
        if str(wind_region).upper() == "C":
            return 1100.0
        return 1700.0

    def actual_wall_girt_spacing_mm(self, eave_height_mm, wind_region):
        max_spacing = self.wall_girt_spacing_mm(wind_region)
        if eave_height_mm <= 0.0:
            return max_spacing
        spaces = max(1, math.ceil(eave_height_mm / max_spacing))
        return eave_height_mm / spaces

    def profile_depth_mm(self, profile):
        if getattr(profile, "depth", 0.0) and profile.depth > 0.0:
            return profile.depth
        match = re.search(r"(\d+(?:\.\d+)?)", profile.name)
        return float(match.group(1)) if match else 0.0

    def build_purlin_layout(self, top_nodes, wind_region, importance_level):
        end_limit = 1200.0
        mid_limit = self.purlin_mid_span_limit_mm(wind_region, importance_level)
        if len(top_nodes) < 2:
            return PurlinLayout([], mid_limit, end_limit, mid_limit)

        apex_index = max(range(len(top_nodes)), key=lambda i: top_nodes[i].y)
        points = []
        spacings = []

        def point_at_distance(nodes, distance):
            remaining = distance
            for n1, n2 in zip(nodes[:-1], nodes[1:]):
                dx = n2.x - n1.x
                dy = n2.y - n1.y
                length = math.hypot(dx, dy)
                if length <= 1e-9:
                    continue
                if remaining <= length:
                    t = remaining / length
                    return (n1.x + dx * t, n1.y + dy * t)
                remaining -= length
            return (nodes[-1].x, nodes[-1].y)

        def slope_points(nodes):
            lengths = [math.hypot(n2.x - n1.x, n2.y - n1.y) for n1, n2 in zip(nodes[:-1], nodes[1:])]
            slope_length = sum(lengths)
            if slope_length <= 1e-9:
                return []
            if slope_length <= 2.0 * end_limit:
                stations = [0.0, slope_length]
            else:
                middle_length = slope_length - 2.0 * end_limit
                middle_spaces = max(1, math.ceil(middle_length / mid_limit))
                actual_mid_spacing = middle_length / middle_spaces
                stations = [0.0, end_limit]
                stations.extend(end_limit + actual_mid_spacing * i for i in range(1, middle_spaces))
                stations.extend([slope_length - end_limit, slope_length])
            return [(station, point_at_distance(nodes, station)) for station in sorted(set(round(s, 6) for s in stations))]

        left_slope = top_nodes[:apex_index + 1]
        right_slope = list(reversed(top_nodes[apex_index:]))
        for slope in [left_slope, right_slope]:
            slope_layout = slope_points(slope)
            points.extend(point for _, point in slope_layout)
            spacings.extend(b - a for (a, _), (b, _) in zip(slope_layout[:-1], slope_layout[1:]))

        unique_points = []
        seen = set()
        for x, y in points:
            key = (round(x, 3), round(y, 3))
            if key not in seen:
                seen.add(key)
                unique_points.append((x, y))

        return PurlinLayout(unique_points, max(spacings or [mid_limit]), end_limit, mid_limit)

    def member_check_restraint_lengths(self, frame_type, left_base, right_base, bottom_nodes, purlin_layout=None, column_top_deduction_mm=0.0):
        cladding = self.column_cladding_restraints(frame_type)
        column_top_deduction_mm = max(column_top_deduction_mm, 0.0)
        left_unclad_length = max(abs(bottom_nodes[0].y - left_base.y) - column_top_deduction_mm, 0.0)
        right_unclad_length = max(abs(bottom_nodes[-1].y - right_base.y) - column_top_deduction_mm, 0.0)
        top_chord_minor_length = purlin_layout.max_spacing_mm if purlin_layout is not None else 1700.0
        bottom_chord_minor_length = 2.0 * top_chord_minor_length

        def column_lengths(group, unclad_length):
            if cladding[group]:
                return {"major_mm": 3000.0, "minor_mm": 1500.0, "ke": 0.85, "am": 1.75, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": True}
            return {"major_mm": unclad_length, "minor_mm": unclad_length, "ke": 0.85, "am": 1.75, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": False}

        return {
            "LEFT_COLUMN": column_lengths("LEFT_COLUMN", left_unclad_length),
            "RIGHT_COLUMN": column_lengths("RIGHT_COLUMN", right_unclad_length),
            "WEB": {"major_mm": None, "minor_mm": None, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "LEFT_CANOPY_WEB": {"major_mm": None, "minor_mm": None, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "RIGHT_CANOPY_WEB": {"major_mm": None, "minor_mm": None, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "TOP": {"major_mm": None, "minor_mm": top_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "BOTTOM": {"major_mm": None, "minor_mm": bottom_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "LEFT_CANOPY_TOP": {"major_mm": None, "minor_mm": top_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "LEFT_CANOPY_BOTTOM": {"major_mm": None, "minor_mm": top_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "RIGHT_CANOPY_TOP": {"major_mm": None, "minor_mm": top_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "RIGHT_CANOPY_BOTTOM": {"major_mm": None, "minor_mm": top_chord_minor_length, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": None},
            "LEFT_LEAN_COLUMN": {"major_mm": None, "minor_mm": None, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": False},
            "RIGHT_LEAN_COLUMN": {"major_mm": None, "minor_mm": None, "ke": 1.0, "am": 1.0, "kt": 1.0, "kl": 1.0, "kr": 1.0, "clad": False},
        }

    def end_wall_bracing_eave_spring(self, building_length_m, bay_size_m, brace_diameter_mm):
        bay_size_m = max(bay_size_m, 1e-9)
        half_length_m = max(building_length_m, 0.0) / 2.0
        squares_to_end = max(1, math.ceil(half_length_m / bay_size_m))
        bay_size_mm = bay_size_m * 1000.0
        brace_area_mm2 = math.pi * max(brace_diameter_mm, 0.0) ** 2 / 4.0
        diagonal_length_mm = math.sqrt(2.0) * bay_size_mm
        base_spring_kn_mm = (200000.0 * brace_area_mm2 / diagonal_length_mm) * 0.5 / 1000.0
        effective_spring_kn_mm = 2.0 * base_spring_kn_mm / squares_to_end
        return {
            "squares_to_end": squares_to_end,
            "brace_diameter_mm": brace_diameter_mm,
            "brace_area_mm2": brace_area_mm2,
            "diagonal_length_mm": diagonal_length_mm,
            "base_spring_kn_mm": base_spring_kn_mm,
            "effective_spring_kn_mm": effective_spring_kn_mm,
        }

    def purlin_capacity_row(self, table, section, span_mm):
        for profile in table.get("profiles", []):
            if profile.get("section") != section:
                continue
            rows = profile.get("rows", [])
            if not rows:
                return None
            if span_mm <= rows[0].get("span_mm", 0.0):
                return dict(rows[0])
            if span_mm > rows[-1].get("span_mm", 0.0) + 1e-9:
                return None
            for lower, upper in zip(rows[:-1], rows[1:]):
                lower_span = lower.get("span_mm", 0.0)
                upper_span = upper.get("span_mm", 0.0)
                if lower_span <= span_mm <= upper_span + 1e-9:
                    if abs(upper_span - lower_span) <= 1e-9:
                        return dict(upper)
                    ratio = (span_mm - lower_span) / (upper_span - lower_span)
                    result = {"span_mm": span_mm, "interpolated": abs(span_mm - lower_span) > 1e-9 and abs(span_mm - upper_span) > 1e-9}
                    for key, lower_value in lower.items():
                        if key == "span_mm":
                            continue
                        upper_value = upper.get(key)
                        if isinstance(lower_value, (int, float)) and isinstance(upper_value, (int, float)):
                            result[key] = lower_value + (upper_value - lower_value) * ratio
                    return result
        return None

    def metroll_section_area_mm2(self, section):
        areas = {
            "C/Z 150 12": 354.0,
            "C/Z 150 15": 442.0,
            "C/Z 150 19": 561.0,
            "C/Z 150 24": 708.0,
            "C/Z 200 12": 444.0,
            "C/Z 200 15": 555.0,
            "C/Z 200 19": 713.0,
            "C/Z 200 24": 900.0,
            "C/Z 250 15": 638.0,
            "C/Z 250 19": 808.0,
            "C/Z 250 24": 1020.0,
            "C/Z 300 19": 998.0,
            "C/Z 300 24": 1260.0,
            "C/Z 300 30": 1575.0,
            "C/Z 350 19": 1207.0,
            "C/Z 350 24": 1524.0,
            "C/Z 350 30": 1905.0,
        }
        if section in areas:
            return areas[section]
        numbers = [float(value) for value in re.findall(r"\d+", section or "")]
        if len(numbers) >= 2:
            return numbers[0] * numbers[1] / 10.0
        return float("inf")

    def steel_weight_kg(self, area_mm2, length_mm):
        return area_mm2 * length_mm * 1e-9 * 7850.0

    def stock_length_m_for_cut(self, cut_length_m):
        if cut_length_m <= 0.0:
            return 0.0
        for stock_length in UB_STOCK_LENGTHS_M:
            if cut_length_m <= stock_length + 1e-9:
                return stock_length
        pieces = math.ceil(cut_length_m / UB_STOCK_LENGTHS_M[-1])
        return pieces * UB_STOCK_LENGTHS_M[-1]

    def section_cost_bucket(self, profile):
        text = f"{profile.name} {profile.group_name} {profile.shape_type}".upper()
        if any(token in text for token in ["SHS", "RHS", "CHS", "HOLLOW", "TUBE"]):
            return "shs"
        if re.search(r"(^|[^A-Z])WB([^A-Z]|$)", text):
            return "wb"
        return "ub"

    def metroll_section_weight_kg_m(self, section):
        area = self.metroll_section_area_mm2(section)
        if not math.isfinite(area):
            return 0.0
        return area * 1e-6 * 7850.0

    def costing_rate_from_input(self, widget, default_key):
        try:
            return max(float(widget.text()), 0.0)
        except ValueError:
            return float(COSTING_DEFAULTS[default_key])

    def foundation_float_from_input(self, widget, default):
        try:
            return max(float(widget.text()), 0.0)
        except ValueError:
            return default

    def design_pier_footing(self, compression_kn, uplift_kn, bearing_kpa, skin_friction_kpa, use_skin_friction_for_bearing=False):
        compression_kn = max(compression_kn, 0.0)
        uplift_kn = max(uplift_kn, 0.0)
        depth_step_m = 0.25
        for diameter_mm in STANDARD_PIER_DIAMETERS_MM:
            diameter_m = diameter_mm / 1000.0
            base_area_m2 = math.pi * diameter_m**2 / 4.0
            perimeter_m = math.pi * diameter_m
            depth_m = 1.0
            while depth_m <= 4.0 + 1e-9:
                skin_capacity_kn = skin_friction_kpa * perimeter_m * max(depth_m - 1.0, 0.0)
                bearing_skin_capacity_kn = skin_capacity_kn if use_skin_friction_for_bearing else 0.0
                bearing_capacity_kn = bearing_kpa * base_area_m2 + bearing_skin_capacity_kn
                if bearing_capacity_kn + 1e-9 < compression_kn:
                    depth_m += depth_step_m
                    continue
                skin_capacity_kn = skin_friction_kpa * perimeter_m * max(depth_m - 1.0, 0.0)
                if skin_capacity_kn + 1e-9 >= uplift_kn:
                    return {
                        "diameter_mm": diameter_mm,
                        "depth_m": depth_m,
                        "bearing_capacity_kn": bearing_capacity_kn,
                        "base_bearing_capacity_kn": bearing_kpa * base_area_m2,
                        "bearing_skin_capacity_kn": bearing_skin_capacity_kn,
                        "skin_capacity_kn": skin_capacity_kn,
                        "use_skin_friction_for_bearing": use_skin_friction_for_bearing,
                        "status": "OK",
                    }
                depth_m += depth_step_m

        diameter_mm = STANDARD_PIER_DIAMETERS_MM[-1]
        diameter_m = diameter_mm / 1000.0
        base_area_m2 = math.pi * diameter_m**2 / 4.0
        perimeter_m = math.pi * diameter_m
        required_skin_depth_m = 1.0
        if skin_friction_kpa * perimeter_m > 1e-9:
            required_skin_depth_m = 1.0 + uplift_kn / (skin_friction_kpa * perimeter_m)
        required_bearing_diameter_m = 0.0
        if bearing_kpa > 1e-9:
            required_bearing_diameter_m = math.sqrt(4.0 * compression_kn / (math.pi * bearing_kpa))
        depth_m = max(4.0, math.ceil(required_skin_depth_m / depth_step_m) * depth_step_m)
        skin_capacity_kn = skin_friction_kpa * perimeter_m * max(depth_m - 1.0, 0.0)
        bearing_skin_capacity_kn = skin_capacity_kn if use_skin_friction_for_bearing else 0.0
        return {
            "diameter_mm": diameter_mm,
            "depth_m": depth_m,
            "bearing_capacity_kn": bearing_kpa * base_area_m2 + bearing_skin_capacity_kn,
            "base_bearing_capacity_kn": bearing_kpa * base_area_m2,
            "bearing_skin_capacity_kn": bearing_skin_capacity_kn,
            "skin_capacity_kn": skin_capacity_kn,
            "required_bearing_diameter_mm": required_bearing_diameter_m * 1000.0,
            "use_skin_friction_for_bearing": use_skin_friction_for_bearing,
            "status": "CHECK" if required_bearing_diameter_m * 1000.0 > diameter_mm + 1e-9 else "OK >4m",
        }

    def foundation_design_from_reactions(self, structure, combinations, apply_combination_loads, foundation_nodes):
        bearing_kpa = self.foundation_float_from_input(self.foundation_bearing_input, 150.0)
        skin_friction_kpa = self.foundation_float_from_input(self.foundation_skin_friction_input, 15.0)
        use_skin_friction_for_bearing = self.foundation_skin_bearing_checkbox.isChecked()
        items = []
        for label, node in foundation_nodes:
            envelope = {
                "compression_kn": 0.0,
                "compression_combo": "-",
                "uplift_kn": 0.0,
                "uplift_combo": "-",
                "shear_kn": 0.0,
                "shear_combo": "-",
                "moment_knm": 0.0,
                "moment_combo": "-",
            }
            for combination in combinations:
                apply_combination_loads(combination)
                _, reactions, _, _ = self.solve_structure(structure)
                ry_kn = reactions[node.uy] / 1000.0
                rx_kn = reactions[node.ux] / 1000.0
                mz_knm = reactions[node.rz] / 1_000_000.0
                compression_kn = max(ry_kn, 0.0)
                uplift_kn = max(-ry_kn, 0.0)
                shear_kn = abs(rx_kn)
                moment_knm = abs(mz_knm)
                if combination.is_serviceability and compression_kn > envelope["compression_kn"]:
                    envelope["compression_kn"] = compression_kn
                    envelope["compression_combo"] = combination.name
                if combination.is_ultimate and uplift_kn > envelope["uplift_kn"]:
                    envelope["uplift_kn"] = uplift_kn
                    envelope["uplift_combo"] = combination.name
                if combination.is_ultimate and shear_kn > envelope["shear_kn"]:
                    envelope["shear_kn"] = shear_kn
                    envelope["shear_combo"] = combination.name
                if combination.is_ultimate and moment_knm > envelope["moment_knm"]:
                    envelope["moment_knm"] = moment_knm
                    envelope["moment_combo"] = combination.name
            design = self.design_pier_footing(
                envelope["compression_kn"],
                envelope["uplift_kn"],
                bearing_kpa,
                skin_friction_kpa,
                use_skin_friction_for_bearing,
            )
            items.append({"label": label, "node": node.id, **envelope, **design})
        return {
            "bearing_kpa": bearing_kpa,
            "skin_friction_kpa": skin_friction_kpa,
            "use_skin_friction_for_bearing": use_skin_friction_for_bearing,
            "diameters_mm": STANDARD_PIER_DIAMETERS_MM,
            "items": items,
        }

    def foundation_design_from_envelopes(self, reaction_envelopes):
        bearing_kpa = self.foundation_float_from_input(self.foundation_bearing_input, 150.0)
        skin_friction_kpa = self.foundation_float_from_input(self.foundation_skin_friction_input, 15.0)
        use_skin_friction_for_bearing = self.foundation_skin_bearing_checkbox.isChecked()
        items = []
        for envelope in reaction_envelopes:
            design = self.design_pier_footing(
                envelope.get("compression_kn", 0.0),
                envelope.get("uplift_kn", 0.0),
                bearing_kpa,
                skin_friction_kpa,
                use_skin_friction_for_bearing,
            )
            items.append({**envelope, **design})
        return {
            "bearing_kpa": bearing_kpa,
            "skin_friction_kpa": skin_friction_kpa,
            "use_skin_friction_for_bearing": use_skin_friction_for_bearing,
            "diameters_mm": STANDARD_PIER_DIAMETERS_MM,
            "items": items,
        }

    def costing_from_structure(self, structure, building_length_m, bay_size_m):
        shs_rate = self.costing_rate_from_input(self.shs_cost_input, "shs_rate_per_kg")
        ub_rate = self.costing_rate_from_input(self.ub_cost_input, "ub_rate_per_kg")
        wb_rate = self.costing_rate_from_input(self.wb_cost_input, "wb_rate_per_kg")
        purlin_rate = self.costing_rate_from_input(self.purlin_cost_input, "purlin_rate_per_kg")

        bay_count = max(1, math.ceil(max(building_length_m, 0.0) / max(bay_size_m, 1e-9)))
        portal_count = bay_count + 1

        frame_weights = {"shs": 0.0, "ub": 0.0, "wb": 0.0}
        frame_base_weights = {"shs": 0.0, "ub": 0.0, "wb": 0.0}
        stock_length_m = {"ub": 0.0, "wb": 0.0}
        cut_length_m = {"ub": 0.0, "wb": 0.0}
        for element in structure.elements:
            bucket = self.section_cost_bucket(element.profile)
            exact_weight = self.steel_weight_kg(element.profile.A, element.length())
            frame_base_weights[bucket] += exact_weight
            if bucket == "shs":
                frame_weights["shs"] += exact_weight * 1.07
            else:
                element_cut_length_m = element.length() / 1000.0
                element_stock_length_m = self.stock_length_m_for_cut(element_cut_length_m)
                cut_length_m[bucket] += element_cut_length_m
                stock_length_m[bucket] += element_stock_length_m
                frame_weights[bucket] += element.profile.A * element_stock_length_m * 1e-6 * 7850.0

        roof_general = (self.purlin_check or {}).get("match", {})
        roof_edge = (self.purlin_check or {}).get("local_match", {})
        edge_rows = int((self.purlin_check or {}).get("edge_rows", 0) or 0)
        roof_rows_total = len((self.purlin_layout.points if self.purlin_layout is not None else []) or [])
        roof_edge_rows_total = min(roof_rows_total, 2 * edge_rows)
        roof_general_rows = max(roof_rows_total - roof_edge_rows_total, 0)
        roof_length_m = building_length_m

        wall_match = (self.wall_girt_check or {}).get("match", {})
        wall_spacing_mm = (self.wall_girt_check or {}).get("spacing_mm", 0.0) or 0.0
        eave_mm = max(float(self.eave_input.text()), 0.0)
        wall_rows_each_side = max(1, math.ceil(eave_mm / wall_spacing_mm)) if wall_spacing_mm > 0.0 else 0
        wall_rows_total = 2 * wall_rows_each_side

        purlin_items = []
        for label, match, count, length_m, span_type in [
            ("Roof general", roof_general, roof_general_rows, roof_length_m, (self.purlin_check or {}).get("span_type", "")),
            ("Roof edge/local", roof_edge, roof_edge_rows_total, roof_length_m, (self.purlin_check or {}).get("span_type", "")),
            ("Wall girts", wall_match, wall_rows_total, roof_length_m, (self.wall_girt_check or {}).get("span_type", "")),
        ]:
            section = match.get("section")
            if not section or count <= 0 or length_m <= 0.0:
                continue
            kg_m = self.metroll_section_weight_kg_m(section)
            total_m = count * length_m
            lap_factor = 1.15 if span_type == "Continuous lapped" else 1.0
            purlin_items.append({
                "label": label,
                "section": section,
                "span_type": span_type,
                "count": count,
                "length_m": length_m,
                "total_m": total_m,
                "kg_m": kg_m,
                "lap_factor": lap_factor,
                "kg_base": kg_m * total_m,
                "kg": kg_m * total_m * lap_factor,
            })

        frame_total_kg = {key: value * portal_count for key, value in frame_weights.items()}
        purlin_total_kg = sum(item["kg"] for item in purlin_items)
        costs = {
            "shs": frame_total_kg["shs"] * shs_rate,
            "ub": frame_total_kg["ub"] * ub_rate,
            "wb": frame_total_kg["wb"] * wb_rate,
            "purlin": purlin_total_kg * purlin_rate,
        }
        return {
            "bay_count": bay_count,
            "portal_count": portal_count,
            "building_length_m": building_length_m,
            "bay_size_m": bay_size_m,
            "per_portal_weights": frame_weights,
            "per_portal_base_weights": frame_base_weights,
            "frame_total_kg": frame_total_kg,
            "ub_cut_length_m_per_portal": cut_length_m["ub"],
            "ub_stock_length_m_per_portal": stock_length_m["ub"],
            "wb_cut_length_m_per_portal": cut_length_m["wb"],
            "wb_stock_length_m_per_portal": stock_length_m["wb"],
            "shs_wastage_factor": 1.07,
            "continuous_lap_factor": 1.15,
            "ub_stock_lengths_m": UB_STOCK_LENGTHS_M,
            "purlin_items": purlin_items,
            "purlin_total_kg": purlin_total_kg,
            "rates": {"shs": shs_rate, "ub": ub_rate, "wb": wb_rate, "purlin": purlin_rate},
            "costs": costs,
            "total_cost": sum(costs.values()),
        }

    def purlin_design_pressures(self, wind_inputs, tributary_area_m2=None, edge_zone=None, return_info=False):
        max_upward_kpa = 0.0
        max_downward_kpa = 0.0
        max_local_pressure_factor = 1.0
        local_dimension_a_m = 0.0
        purlin_wind_inputs = replace(wind_inputs, left_canopy_length_m=0.0, right_canopy_length_m=0.0)
        for cpe_case in CPE_CASE_OPTIONS:
            source_direction = self.wind_source_direction_for_case(purlin_wind_inputs.orientation, cpe_case)
            result = calculate_wind(replace(purlin_wind_inputs, orientation=source_direction))
            roof_length = result.roof_slope_length_m
            result_a_m = local_pressure_dimension_a(result)
            local_dimension_a_m = max(local_dimension_a_m, result_a_m)
            edge_width = result_a_m
            for cpi_case in CPI_CASE_OPTIONS:
                cpi = active_cpi_for_case(result, cpi_case, cpe_case)
                cpi_pressure = cpi * result.wu_kn_m2 * internal_pressure_factor(result, cpi)
                for zone in result.roof_zones:
                    zone_overlaps_edge = (
                        zone.distance_from_eave_from_m < edge_width
                        or zone.distance_from_eave_to_m > roof_length - edge_width
                    )
                    zone_overlaps_non_edge = (
                        zone.distance_from_eave_to_m > edge_width
                        and zone.distance_from_eave_from_m < roof_length - edge_width
                    )
                    if edge_zone is True and not zone_overlaps_edge:
                        continue
                    if edge_zone is False and not zone_overlaps_non_edge:
                        continue
                    proximity_m = min(
                        max(zone.distance_from_eave_from_m, 0.0),
                        max(roof_length - zone.distance_from_eave_to_m, 0.0),
                    )
                    local_pressure_factor = 1.0
                    if edge_zone is True and tributary_area_m2 is not None:
                        local_pressure_factor = roof_local_pressure_factor(result, tributary_area_m2, proximity_m)
                    max_local_pressure_factor = max(max_local_pressure_factor, local_pressure_factor)
                    if tributary_area_m2 is None:
                        external_pressures = [
                            zone.pressure_max_kn_m2,
                            zone.pressure_min_kn_m2,
                        ]
                    else:
                        external_factor = external_pressure_factor(result, "roof", tributary_area_m2)
                        external_pressures = [
                            (-zone.cpe_max) * result.wu_kn_m2 * external_factor,
                            (-zone.cpe_min) * result.wu_kn_m2 * external_factor,
                        ]
                    for external_pressure in external_pressures:
                        if external_pressure > 0.0:
                            external_pressure *= local_pressure_factor
                        combined_pressure = external_pressure + cpi_pressure
                        max_upward_kpa = max(max_upward_kpa, combined_pressure)
                        max_downward_kpa = max(max_downward_kpa, -combined_pressure)
        if return_info:
            return max_upward_kpa, max_downward_kpa, {
                "local_pressure_factor": max_local_pressure_factor,
                "local_dimension_a_m": local_dimension_a_m,
            }
        return max_upward_kpa, max_downward_kpa

    def wall_girt_design_pressures(self, wind_inputs):
        max_outward_kpa = 0.0
        max_inward_kpa = 0.0
        max_inward_service_kpa = 0.0

        for cpe_case in CPE_CASE_OPTIONS:
            source_direction = self.wind_source_direction_for_case(wind_inputs.orientation, cpe_case)
            result = calculate_wind(replace(wind_inputs, orientation=source_direction))
            for cpi_case in CPI_CASE_OPTIONS:
                cpi = active_cpi_for_case(result, cpi_case, cpe_case)
                for cpe, surface in [(0.7, "windward_wall"), (result.leeward_wall_cpe, "leeward_wall")]:
                    for wind_pressure in ["ultimate", "serviceability"]:
                        external_factor = external_pressure_factor(result, surface, result.inputs.bay_size_m * result.inputs.eave_height_m)
                        external = cpe * result.wu_kn_m2 * wind_pressure_scale(result, wind_pressure) * external_factor
                        internal = cpi * result.wu_kn_m2 * wind_pressure_scale(result, wind_pressure) * internal_pressure_factor(result, cpi)
                        combined = external - internal
                        if wind_pressure == "ultimate":
                            max_inward_kpa = max(max_inward_kpa, combined)
                            max_outward_kpa = max(max_outward_kpa, -combined)
                        else:
                            max_inward_service_kpa = max(max_inward_service_kpa, abs(combined))

        return max_outward_kpa, max_inward_kpa, max_inward_service_kpa

    def match_metroll_purlin(self, span_type, span_mm, outward_kn_m, inward_strength_kn_m, inward_service_kn_m):
        if not self.metroll_tables:
            return {"status": "NOT CHECKED", "message": "Metroll capacity table cache not found."}
        table_ids = {
            "Single span": ("1A", "1B"),
            "Double span": ("2A", "2B"),
            "Continuous lapped": ("5A", "5B"),
        }.get(span_type)
        if table_ids is None:
            return {"status": "NOT CHECKED", "message": f"Unsupported purlin span type: {span_type}"}
        outward_table = self.metroll_tables["tables"].get(table_ids[0])
        inward_table = self.metroll_tables["tables"].get(table_ids[1])
        if not outward_table or not inward_table:
            return {"status": "NOT CHECKED", "message": f"Missing Metroll tables {table_ids[0]} / {table_ids[1]}."}
        first_profile = outward_table["profiles"][0]
        last_span = first_profile["rows"][-1]["span_mm"] if first_profile.get("rows") else 0.0
        if span_mm > last_span:
            return {"status": "NOT OK", "message": f"Bay span {span_mm:.0f} mm exceeds table maximum {last_span:.0f} mm."}

        passing = []
        candidates = []
        for outward_profile in outward_table.get("profiles", []):
            section = outward_profile.get("section")
            outward_row = self.purlin_capacity_row(outward_table, section, span_mm)
            inward_row = self.purlin_capacity_row(inward_table, section, span_mm)
            if not outward_row or not inward_row:
                continue
            outward_capacity = outward_row["strength_capacity_kn_m"]
            inward_strength_capacity = inward_row["strength_capacity_kn_m"]
            inward_service_capacity = inward_row["service_l150_capacity_kn_m"]
            capped_by_double = False
            if span_type == "Single span":
                double_outward_row = self.purlin_capacity_row(self.metroll_tables["tables"].get("2A", {}), section, span_mm)
                double_inward_row = self.purlin_capacity_row(self.metroll_tables["tables"].get("2B", {}), section, span_mm)
                if double_outward_row:
                    capped_by_double = capped_by_double or double_outward_row["strength_capacity_kn_m"] < outward_capacity
                    outward_capacity = min(outward_capacity, double_outward_row["strength_capacity_kn_m"])
                if double_inward_row:
                    capped_by_double = capped_by_double or double_inward_row["strength_capacity_kn_m"] < inward_strength_capacity
                    capped_by_double = capped_by_double or double_inward_row["service_l150_capacity_kn_m"] < inward_service_capacity
                    inward_strength_capacity = min(inward_strength_capacity, double_inward_row["strength_capacity_kn_m"])
                    inward_service_capacity = min(inward_service_capacity, double_inward_row["service_l150_capacity_kn_m"])
            section_numbers = [float(value) for value in re.findall(r"\d+", section or "")]
            depth_mm = section_numbers[0] if section_numbers else 0.0
            thickness_tenths_mm = section_numbers[1] if len(section_numbers) > 1 else float("inf")
            candidate = {
                "section": section,
                "area_mm2": self.metroll_section_area_mm2(section),
                "depth_mm": depth_mm,
                "thickness_tenths_mm": thickness_tenths_mm,
                "span_row_mm": outward_row["span_mm"],
                "outward_capacity_kn_m": outward_capacity,
                "inward_strength_capacity_kn_m": inward_strength_capacity,
                "inward_service_capacity_kn_m": inward_service_capacity,
                "capped_by_double_span": capped_by_double,
                "outward_ratio": outward_kn_m / outward_capacity if outward_capacity > 1e-12 else float("inf"),
                "inward_strength_ratio": inward_strength_kn_m / inward_strength_capacity if inward_strength_capacity > 1e-12 else float("inf"),
                "inward_service_ratio": inward_service_kn_m / inward_service_capacity if inward_service_capacity > 1e-12 else float("inf"),
            }
            candidates.append(candidate)
            if (
                outward_kn_m <= outward_capacity + 1e-9
                and inward_strength_kn_m <= inward_strength_capacity + 1e-9
                and inward_service_kn_m <= inward_service_capacity + 1e-9
            ):
                passing.append({
                    "status": "OK",
                    "section": section,
                    "area_mm2": self.metroll_section_area_mm2(section),
                    "depth_mm": depth_mm,
                    "thickness_tenths_mm": thickness_tenths_mm,
                    "span_row_mm": outward_row["span_mm"],
                    "interpolated": outward_row.get("interpolated", False) or inward_row.get("interpolated", False),
                    "tables": f"{table_ids[0]} / {table_ids[1]}",
                    "outward_capacity_kn_m": outward_capacity,
                    "inward_strength_capacity_kn_m": inward_strength_capacity,
                    "inward_service_capacity_kn_m": inward_service_capacity,
                    "capped_by_double_span": capped_by_double,
                })
        if passing:
            lightest_area = min(item["area_mm2"] for item in passing)
            near_lightest = [
                item for item in passing
                if item["area_mm2"] <= lightest_area * 1.05
            ]
            selected = min(near_lightest, key=lambda item: (-item["depth_mm"], item["area_mm2"], item["section"]))
            failed = [
                item for item in candidates
                if max(item["outward_ratio"], item["inward_strength_ratio"], item["inward_service_ratio"]) > 1.0
            ]
            failed.sort(key=lambda item: max(item["outward_ratio"], item["inward_strength_ratio"], item["inward_service_ratio"]))
            selected["near_misses"] = failed[:3]
            return selected
        return {"status": "NOT OK", "message": "No listed Metroll section satisfies the calculated line loads."}

    def purlin_check_from_inputs(self, purlin_layout, bay_size_m, roof_loads, wind_inputs, span_type):
        if purlin_layout is None or purlin_layout.max_spacing_mm <= 0.0 or bay_size_m <= 0.0:
            return None
        spacing_m = purlin_layout.max_spacing_mm / 1000.0
        span_mm = bay_size_m * 1000.0
        local_area_m2 = spacing_m * bay_size_m
        local_dimension_a_m = local_pressure_dimension_a(wind_inputs)
        edge_rows = max(1, math.ceil(local_dimension_a_m / spacing_m))
        permanent_kpa = abs(roof_loads.permanent_vertical_kpa)
        live_kpa = abs(roof_loads.q_vertical_kpa)
        inward_gravity_uls_kpa = 1.2 * permanent_kpa + 1.5 * live_kpa
        inward_service_kpa = permanent_kpa + live_kpa
        wind_upward_kpa, wind_downward_kpa = self.purlin_design_pressures(wind_inputs, edge_zone=False)
        outward_kn_m = wind_upward_kpa * spacing_m
        inward_wind_uls_kpa = 1.2 * permanent_kpa + wind_downward_kpa
        inward_strength_kn_m = max(inward_gravity_uls_kpa, inward_wind_uls_kpa) * spacing_m
        inward_service_kn_m = inward_service_kpa * spacing_m
        match = self.match_metroll_purlin(span_type, span_mm, outward_kn_m, inward_strength_kn_m, inward_service_kn_m)
        local_wind_upward_kpa, local_wind_downward_kpa, local_pressure_info = self.purlin_design_pressures(
            wind_inputs,
            tributary_area_m2=local_area_m2,
            edge_zone=True,
            return_info=True,
        )
        local_outward_kn_m = local_wind_upward_kpa * spacing_m
        local_inward_wind_uls_kpa = 1.2 * permanent_kpa + local_wind_downward_kpa
        local_inward_strength_kn_m = max(inward_gravity_uls_kpa, local_inward_wind_uls_kpa) * spacing_m
        local_match = self.match_metroll_purlin(span_type, span_mm, local_outward_kn_m, local_inward_strength_kn_m, inward_service_kn_m)
        return {
            "span_type": span_type,
            "spacing_mm": purlin_layout.max_spacing_mm,
            "span_mm": span_mm,
            "local_area_m2": local_area_m2,
            "edge_rows": edge_rows,
            "permanent_kpa": permanent_kpa,
            "live_kpa": live_kpa,
            "wind_upward_kpa": wind_upward_kpa,
            "wind_downward_kpa": wind_downward_kpa,
            "outward_kn_m": outward_kn_m,
            "inward_strength_kn_m": inward_strength_kn_m,
            "inward_gravity_uls_kpa": inward_gravity_uls_kpa,
            "inward_wind_uls_kpa": inward_wind_uls_kpa,
            "inward_service_kn_m": inward_service_kn_m,
            "match": match,
            "local_wind_upward_kpa": local_wind_upward_kpa,
            "local_wind_downward_kpa": local_wind_downward_kpa,
            "local_outward_kn_m": local_outward_kn_m,
            "local_inward_strength_kn_m": local_inward_strength_kn_m,
            "local_pressure_factor": local_pressure_info["local_pressure_factor"],
            "local_dimension_a_m": local_pressure_info["local_dimension_a_m"],
            "local_inward_wind_uls_kpa": local_inward_wind_uls_kpa,
            "local_match": local_match,
        }

    def wall_girt_check_from_inputs(self, wind_inputs, bay_size_m, eave_height_mm, span_type):
        if bay_size_m <= 0.0:
            return None
        spacing_mm = self.actual_wall_girt_spacing_mm(eave_height_mm, wind_inputs.wind_region)
        spacing_m = spacing_mm / 1000.0
        span_mm = bay_size_m * 1000.0
        wind_outward_kpa, wind_inward_kpa, wind_service_kpa = self.wall_girt_design_pressures(wind_inputs)
        outward_kn_m = wind_outward_kpa * spacing_m
        inward_strength_kn_m = wind_inward_kpa * spacing_m
        inward_service_kn_m = wind_service_kpa * spacing_m
        match = self.match_metroll_purlin(span_type, span_mm, outward_kn_m, inward_strength_kn_m, inward_service_kn_m)
        return {
            "span_type": span_type,
            "spacing_mm": spacing_mm,
            "max_spacing_mm": self.wall_girt_spacing_mm(wind_inputs.wind_region),
            "span_mm": span_mm,
            "wind_outward_kpa": wind_outward_kpa,
            "wind_inward_kpa": wind_inward_kpa,
            "wind_service_kpa": wind_service_kpa,
            "outward_kn_m": outward_kn_m,
            "inward_strength_kn_m": inward_strength_kn_m,
            "inward_service_kn_m": inward_service_kn_m,
            "match": match,
        }

    def wind_inputs_from_ui(self, span_mm, pitch_deg, eave_mm):
        building_length_m = float(self.building_length_input.text())
        wall_states = self.envelope_plan_selector.wall_states()
        frame_type = self.frame_type_from_wall_states(wall_states)
        return WindInputs(
            wind_region=self.wind_region_combo.currentText(),
            importance_level=int(self.importance_combo.currentText()),
            terrain_category=float(self.terrain_combo.currentText()),
            eave_height_m=eave_mm / 1000.0,
            roof_pitch_deg=pitch_deg,
            building_width_m=span_mm / 1000.0,
            building_length_m=building_length_m,
            bay_size_m=self.bay_size_m_from_ui(),
            orientation=self.wind_orientation_selector.currentText(),
            frame_type=frame_type,
            cpi_case="Cpi +",
            left_wall_clad=wall_states["left"],
            right_wall_clad=wall_states["right"],
            front_wall_clad=wall_states["front"],
            back_wall_clad=wall_states["back"],
            left_canopy_length_m=(max(float(self.left_canopy_length_input.text()), 0.0) / 1000.0 if self.side_addition_selected(self.left_canopy_type_combo) else 0.0),
            right_canopy_length_m=(max(float(self.right_canopy_length_input.text()), 0.0) / 1000.0 if self.side_addition_selected(self.right_canopy_type_combo) else 0.0),
            reduction_factor=float(self.wind_reduction_input.text()),
        )

    def wind_source_direction_for_case(self, north_up_direction, cpe_case):
        directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
        if north_up_direction not in directions:
            return north_up_direction
        index = directions.index(north_up_direction)
        if str(cpe_case).startswith("Left"):
            return directions[(index - 2) % len(directions)]
        if str(cpe_case).startswith("Right"):
            return directions[(index + 2) % len(directions)]
        if str(cpe_case).startswith("End"):
            return north_up_direction
        return north_up_direction

    def midspan_deflection_divisor(self):
        return 150.0 if (self.serviceability_importance_level or 1) == 1 else 300.0

    def midspan_component_deflection_divisors(self):
        if (self.serviceability_importance_level or 1) == 1:
            return {"G": 240.0, "Q": 180.0, "Ws": 100.0}
        return {"G": 360.0, "Q": 240.0, "Ws": 150.0}

    def column_half_width_for_span_offset(self, profile):
        return (profile.depth or profile.flange_width or 0.0) / 2.0

    def update_license_status_label(self, refresh_server=False):
        label = getattr(self, "license_status_label", None)
        if label is None:
            return
        from licensing import license_enforced

        if not license_enforced():
            label.setText("Not enforced for this run")
            label.setStyleSheet("color: #777;")
            return
        try:
            from licensing import LicenseManager

            manager = LicenseManager()
            status = manager.refresh() if refresh_server else manager.local_status()
        except Exception as exc:
            label.setText(f"Check failed: {exc}")
            label.setStyleSheet("color: #b00020;")
            return
        if status.valid:
            details = []
            if status.customer:
                details.append(status.customer)
            if status.plan:
                details.append(status.plan)
            if status.days_remaining is not None:
                details.append(f"{status.days_remaining} days remaining")
            if status.warning:
                details.append(status.warning)
            suffix = f" ({', '.join(details)})" if details else ""
            label.setText(f"Valid{suffix}")
            color = "#9a6700" if status.warning else "#1b7f3a"
            label.setStyleSheet(f"color: {color}; font-weight: bold;")
        else:
            label.setText(status.reason)
            label.setStyleSheet("color: #b00020; font-weight: bold;")

    def update_web_section_visibility(self):
        rows = []
        rows.extend(getattr(self, "web_section_rows", []))
        rows.extend(getattr(self, "web_check_rows", []))
        rows.extend(getattr(self, "left_canopy_web_section_rows", []))
        rows.extend(getattr(self, "left_canopy_web_check_rows", []))
        rows.extend(getattr(self, "right_canopy_web_section_rows", []))
        rows.extend(getattr(self, "right_canopy_web_check_rows", []))
        for widget in rows:
            if widget is not None:
                widget.setVisible(False)

    def update_internal_column_visibility(self):
        has_internal_columns = self.internal_column_count_combo.currentText() != "None"
        for widget in getattr(self, "internal_column_offset_rows", []):
            if widget is not None:
                widget.setVisible(has_internal_columns)
        for widget in getattr(self, "internal_column_section_rows", []):
            if widget is not None:
                widget.setVisible(has_internal_columns)
        for widget in getattr(self, "internal_column_check_rows", []):
            if widget is not None:
                widget.setVisible(has_internal_columns)

    def update_frame_system_visibility(self):
        is_truss = getattr(self, "frame_system_combo", None) is None or self.frame_system_combo.currentText() == "Truss"
        for widget in getattr(self, "bottom_chord_check_rows", []):
            if widget is not None:
                widget.setVisible(is_truss)
        for widget in getattr(self, "truss_depth_rows", []):
            if widget is not None:
                widget.setVisible(is_truss)
        for widget in getattr(self, "rafter_haunch_rows", []):
            if widget is not None:
                widget.setVisible(not is_truss)
        for widget in getattr(self, "web_section_rows", []):
            if widget is not None:
                widget.setVisible(False)
        for widget in getattr(self, "web_check_rows", []):
            if widget is not None:
                widget.setVisible(False)

    def update_eave_restraint_visibility(self):
        restraint = self.eave_x_restraint_combo.currentText()
        show_manual_spring = restraint in {"Left eave", "Right eave", "Both eaves"}
        show_end_brace = restraint == "End wall bracing approx"
        for widget in getattr(self, "eave_x_spring_rows", []):
            if widget is not None:
                widget.setVisible(show_manual_spring)
        for widget in getattr(self, "end_wall_brace_rows", []):
            if widget is not None:
                widget.setVisible(show_end_brace)

    def update_current_canopy_visibility(self):
        def positive_length(input_widget):
            try:
                return float(input_widget.text()) > 0.0
            except ValueError:
                return False

        left_selected = self.side_addition_selected(self.left_canopy_type_combo)
        right_selected = self.side_addition_selected(self.right_canopy_type_combo)
        left_has_canopy = left_selected and positive_length(self.left_canopy_length_input)
        right_has_canopy = right_selected and positive_length(self.right_canopy_length_input)
        left_has_lean_to = left_has_canopy and self.left_canopy_type_combo.currentText() == "Lean-to"
        right_has_lean_to = right_has_canopy and self.right_canopy_type_combo.currentText() == "Lean-to"
        for widget in getattr(self, "left_canopy_geometry_rows", []):
            if widget is not None:
                widget.setVisible(left_selected)
        for widget in getattr(self, "right_canopy_geometry_rows", []):
            if widget is not None:
                widget.setVisible(right_selected)
        if getattr(self, "envelope_plan_selector", None) is not None and hasattr(self.envelope_plan_selector, "set_lean_to_visibility"):
            self.envelope_plan_selector.set_lean_to_visibility(left_has_lean_to, right_has_lean_to)
        self.update_canopy_section_visibility(left_has_canopy, right_has_canopy, left_has_lean_to, right_has_lean_to)
        for rows in [
            getattr(self, "left_canopy_wind_model_rows", []),
            getattr(self, "left_canopy_underside_rows", []),
            getattr(self, "left_canopy_manual_cpn_rows", []),
            getattr(self, "right_canopy_wind_model_rows", []),
            getattr(self, "right_canopy_underside_rows", []),
            getattr(self, "right_canopy_manual_cpn_rows", []),
        ]:
            for widget in rows:
                if widget is not None:
                    widget.setVisible(False)

    def update_snow_visibility(self):
        snow_enabled = self.snow_region_combo.currentText() != "None"
        for widget in getattr(self, "snow_ahd_height_rows", []):
            if widget is not None:
                widget.setVisible(snow_enabled)

    def update_canopy_section_visibility(self, left_has_canopy, right_has_canopy, left_has_lean_to, right_has_lean_to):
        for widget in getattr(self, "left_canopy_section_rows", []):
            if widget is not None:
                widget.setVisible(left_has_canopy)
        for widget in getattr(self, "left_canopy_check_rows", []):
            if widget is not None:
                widget.setVisible(left_has_canopy)
        for widget in getattr(self, "left_canopy_load_rows", []):
            if widget is not None:
                widget.setVisible(left_has_canopy)
        for widget in getattr(self, "left_canopy_web_section_rows", []):
            if widget is not None:
                widget.setVisible(False)
        for widget in getattr(self, "left_canopy_web_check_rows", []):
            if widget is not None:
                widget.setVisible(False)
        for widget in getattr(self, "left_lean_section_rows", []):
            if widget is not None:
                widget.setVisible(left_has_lean_to)
        for widget in getattr(self, "left_lean_check_rows", []):
            if widget is not None:
                widget.setVisible(left_has_lean_to)
        for widget in getattr(self, "left_lean_enclosure_rows", []):
            if widget is not None:
                widget.setVisible(left_has_lean_to)
        for widget in getattr(self, "right_canopy_section_rows", []):
            if widget is not None:
                widget.setVisible(right_has_canopy)
        for widget in getattr(self, "right_canopy_check_rows", []):
            if widget is not None:
                widget.setVisible(right_has_canopy)
        for widget in getattr(self, "right_canopy_load_rows", []):
            if widget is not None:
                widget.setVisible(right_has_canopy)
        for widget in getattr(self, "right_canopy_web_section_rows", []):
            if widget is not None:
                widget.setVisible(False)
        for widget in getattr(self, "right_canopy_web_check_rows", []):
            if widget is not None:
                widget.setVisible(False)
        for widget in getattr(self, "right_lean_section_rows", []):
            if widget is not None:
                widget.setVisible(right_has_lean_to)
        for widget in getattr(self, "right_lean_check_rows", []):
            if widget is not None:
                widget.setVisible(right_has_lean_to)
        for widget in getattr(self, "right_lean_enclosure_rows", []):
            if widget is not None:
                widget.setVisible(right_has_lean_to)

    def generate_and_solve(self):
        try:
            outside_span = float(self.span_input.text())
            pitch = float(self.pitch_input.text())
            eave = float(self.eave_input.text())
            depth = float(self.depth_input.text())
            frame_system = self.frame_system_combo.currentText()
            is_rafter = frame_system == "Rafter"
            rafter_haunch_length = max(float(self.rafter_haunch_length_input.text()), 0.0)
            internal_column_count = self.internal_column_count_combo.currentText()
            internal_column_offset = max(float(self.internal_column_offset_input.text()), 0.0)
            left_canopy_length = max(float(self.left_canopy_length_input.text()), 0.0)
            right_canopy_length = max(float(self.right_canopy_length_input.text()), 0.0)
            left_canopy_type = self.left_canopy_type_combo.currentText()
            right_canopy_type = self.right_canopy_type_combo.currentText()
            if not self.side_addition_selected(self.left_canopy_type_combo):
                left_canopy_length = 0.0
            if not self.side_addition_selected(self.right_canopy_type_combo):
                right_canopy_length = 0.0
            left_canopy_wind_model = self.left_canopy_wind_model_combo.currentText()
            right_canopy_wind_model = self.right_canopy_wind_model_combo.currentText()
            left_canopy_underside = self.left_canopy_underside_combo.currentText()
            right_canopy_underside = self.right_canopy_underside_combo.currentText()
            left_canopy_cpn_uplift = float(self.left_canopy_cpn_uplift_input.text())
            left_canopy_cpn_downward = float(self.left_canopy_cpn_downward_input.text())
            right_canopy_cpn_uplift = float(self.right_canopy_cpn_uplift_input.text())
            right_canopy_cpn_downward = float(self.right_canopy_cpn_downward_input.text())
            left_lean_wall_states = self.left_lean_enclosure_selector.wall_states()
            right_lean_wall_states = self.right_lean_enclosure_selector.wall_states()
            left_lean_outer_wall_clad = left_lean_wall_states.get("left", False)
            right_lean_outer_wall_clad = right_lean_wall_states.get("right", False)
            left_canopy_eave_height = self.optional_height_from_input(self.left_canopy_eave_height_input)
            right_canopy_eave_height = self.optional_height_from_input(self.right_canopy_eave_height_input)
            left_canopy_pitch = float(self.left_canopy_pitch_input.text())
            right_canopy_pitch = float(self.right_canopy_pitch_input.text())
            left_canopy_shallow_depth = max(float(self.left_canopy_shallow_depth_input.text()), 0.0)
            right_canopy_shallow_depth = max(float(self.right_canopy_shallow_depth_input.text()), 0.0)
            eave_x_spring_kn_mm = max(float(self.eave_x_spring_input.text()), 0.0)
            end_wall_brace_dia_mm = max(float(self.end_wall_brace_dia_input.text()), 0.0)
            bay_size_m = self.bay_size_m_from_ui()
            building_length_m = float(self.building_length_input.text())
            entered_q_kpa = abs(float(self.q_load_input.text()))
            snow_loads = self.snow_loads_from_ui()
            effective_q_kpa = max(entered_q_kpa, snow_loads["service_kpa"])
            roof_loads = BasicRoofLoads(
                bay_size_m=bay_size_m,
                g=self.downward_roof_load_from_input(self.g_load_input),
                q=-effective_q_kpa,
                solar=self.downward_roof_load_from_input(self.solar_load_input),
                fire_service=self.downward_roof_load_from_input(self.fire_service_load_input),
                hvac=self.downward_roof_load_from_input(self.hvac_load_input),
                other=self.downward_roof_load_from_input(self.other_load_input),
                placements={
                    "G": self.roof_load_placement_from_ui("g"),
                    "Q": self.roof_load_placement_from_ui("q"),
                    "Solar": self.roof_load_placement_from_ui("solar"),
                    "Other": self.roof_load_placement_from_ui("other"),
                },
            )
            crane_rating_t = max(float(self.crane_rating_input.text()), 0.0)
            left_canopy_loads = BasicRoofLoads(
                bay_size_m=bay_size_m,
                g=self.downward_roof_load_from_input(self.left_canopy_g_load_input),
                q=self.downward_roof_load_from_input(self.left_canopy_q_load_input),
                solar=self.downward_roof_load_from_input(self.left_canopy_solar_load_input),
            )
            right_canopy_loads = BasicRoofLoads(
                bay_size_m=bay_size_m,
                g=self.downward_roof_load_from_input(self.right_canopy_g_load_input),
                q=self.downward_roof_load_from_input(self.right_canopy_q_load_input),
                solar=self.downward_roof_load_from_input(self.right_canopy_solar_load_input),
            )
            wall_loads = WallLoads(
                left_plus_x=0.0,
                right_minus_x=0.0,
            )
        except ValueError:
            self.results_box.setText("Input error: please enter valid numbers.")
            return

        if outside_span <= 0 or eave <= 0 or (not is_rafter and depth <= 0):
            self.results_box.setText("Input error: span, eave height and truss depth must be positive.")
            return

        left_column_profile = self.selected_profile(self.left_column_profile_combo)
        right_column_profile = self.selected_profile(self.right_column_profile_combo)
        internal_column_profile = self.selected_profile(self.internal_column_profile_combo)
        top_profile = self.selected_profile(self.top_profile_combo)
        bottom_profile = self.selected_profile(self.bottom_profile_combo)
        web_profile = self.selected_profile(self.web_profile_combo)
        left_canopy_top_profile = self.selected_profile(self.left_canopy_top_profile_combo)
        left_canopy_bottom_profile = self.selected_profile(self.left_canopy_bottom_profile_combo)
        left_canopy_web_profile = self.selected_profile(self.left_canopy_web_profile_combo)
        left_lean_column_profile = self.selected_profile(self.left_lean_column_profile_combo)
        right_canopy_top_profile = self.selected_profile(self.right_canopy_top_profile_combo)
        right_canopy_bottom_profile = self.selected_profile(self.right_canopy_bottom_profile_combo)
        right_canopy_web_profile = self.selected_profile(self.right_canopy_web_profile_combo)
        right_lean_column_profile = self.selected_profile(self.right_lean_column_profile_combo)
        left_has_canopy = left_canopy_length > 0.0
        right_has_canopy = right_canopy_length > 0.0
        left_has_lean_to = left_has_canopy and left_canopy_type == "Lean-to"
        right_has_lean_to = right_has_canopy and right_canopy_type == "Lean-to"
        left_lean_outer_wall_clad = left_lean_outer_wall_clad and left_has_lean_to
        right_lean_outer_wall_clad = right_lean_outer_wall_clad and right_has_lean_to
        left_lean_wall_states = {key: bool(value) and left_has_lean_to for key, value in left_lean_wall_states.items()}
        right_lean_wall_states = {key: bool(value) and right_has_lean_to for key, value in right_lean_wall_states.items()}
        main_structure_open = not any(self.envelope_plan_selector.wall_states().values())
        whole_structure_open = main_structure_open and not any(left_lean_wall_states.values()) and not any(right_lean_wall_states.values())
        main_roof_is_free_roof = whole_structure_open
        left_canopy_is_free_roof = left_has_canopy and whole_structure_open
        right_canopy_is_free_roof = right_has_canopy and whole_structure_open
        self.update_internal_column_visibility()
        self.update_canopy_section_visibility(left_has_canopy, right_has_canopy, left_has_lean_to, right_has_lean_to)
        if hasattr(self.envelope_plan_selector, "set_lean_to_visibility"):
            self.envelope_plan_selector.set_lean_to_visibility(left_has_lean_to, right_has_lean_to)
        self.update_frame_system_visibility()

        left_column_offset = self.column_half_width_for_span_offset(left_column_profile)
        right_column_offset = self.column_half_width_for_span_offset(right_column_profile)
        span = outside_span - left_column_offset - right_column_offset
        if span <= 0:
            self.results_box.setText("Input error: outside span must be greater than the selected column widths.")
            return
        wind_inputs = self.wind_inputs_from_ui(outside_span, pitch, eave)

        if is_rafter:
            num_panels = max(4, int(math.ceil(span / 2500.0)))
            if num_panels % 2 != 0:
                num_panels += 1
        else:
            target_web_angle_deg = 50.0
            target_panel = depth / math.tan(math.radians(target_web_angle_deg))
            ideal_panels = span / max(target_panel, 1.0)
            panels_per_side = round(ideal_panels / 2)
            if panels_per_side % 2 != 0:
                panels_per_side += 1
            panels_per_side = max(2, panels_per_side)
            num_panels = panels_per_side * 2
        panel = span / num_panels
        web_angle = 0.0 if is_rafter else math.degrees(math.atan(depth / panel))

        self.panel_label.setText(f"{panel:.0f} mm")
        self.web_angle_label.setText(f"{web_angle:.1f}°")
        self.panel_count_label.setText(str(num_panels))

        structure = Structure2D()
        E = 200000  # MPa = N/mm²

        half_span = span / 2
        rise = math.tan(math.radians(pitch)) * half_span
        ridge_y = eave + rise

        top_nodes = []
        bottom_nodes = []
        node_id = 1

        for i in range(num_panels + 1):
            local_x = i * panel
            x = left_column_offset + local_x
            if local_x <= half_span:
                y = eave + local_x * math.tan(math.radians(pitch))
            else:
                y = ridge_y - (local_x - half_span) * math.tan(math.radians(pitch))
            node = Node(node_id, x, y)
            structure.add_node(node)
            top_nodes.append(node)
            node_id += 1

        if is_rafter:
            bottom_nodes = top_nodes
            left_column_bottom_chord_node = top_nodes[0]
            right_column_bottom_chord_node = top_nodes[-1]
        else:
            for top_node in top_nodes:
                node = Node(node_id, top_node.x, top_node.y - depth)
                structure.add_node(node)
                bottom_nodes.append(node)
                node_id += 1

            left_column_bottom_chord_node = bottom_nodes[0]
            right_column_bottom_chord_node = bottom_nodes[-1]

        if not is_rafter:
            # The bottom chord seats on the column but should not impose full
            # rotational continuity into the through-column. Coincident nodes share
            # translations for force transfer while keeping separate rotations.
            left_truss_bottom_node = Node(
                node_id,
                left_column_bottom_chord_node.x,
                left_column_bottom_chord_node.y,
                ux=left_column_bottom_chord_node.ux,
                uy=left_column_bottom_chord_node.uy,
            )
            structure.add_node(left_truss_bottom_node)
            node_id += 1

            right_truss_bottom_node = Node(
                node_id,
                right_column_bottom_chord_node.x,
                right_column_bottom_chord_node.y,
                ux=right_column_bottom_chord_node.ux,
                uy=right_column_bottom_chord_node.uy,
            )
            structure.add_node(right_truss_bottom_node)
            node_id += 1

            bottom_nodes[0] = left_truss_bottom_node
            bottom_nodes[-1] = right_truss_bottom_node

        def bottom_y_at_x(x):
            ordered = sorted(bottom_nodes, key=lambda node: node.x)
            if x <= ordered[0].x:
                return ordered[0].y
            if x >= ordered[-1].x:
                return ordered[-1].y
            for n1, n2 in zip(ordered[:-1], ordered[1:]):
                if min(n1.x, n2.x) - 1e-9 <= x <= max(n1.x, n2.x) + 1e-9:
                    dx = n2.x - n1.x
                    if abs(dx) <= 1e-9:
                        return n1.y
                    t = (x - n1.x) / dx
                    return n1.y + t * (n2.y - n1.y)
            return ordered[-1].y

        internal_column_xs = []
        centre_x = left_column_offset + span / 2.0
        if internal_column_count == "1 central":
            internal_column_xs = [centre_x]
        elif internal_column_count == "2 columns":
            min_x = left_column_offset + 500.0
            max_x = outside_span - right_column_offset - 500.0
            internal_column_xs = [
                max(min_x, min(centre_x - internal_column_offset, max_x)),
                max(min_x, min(centre_x + internal_column_offset, max_x)),
            ]
            internal_column_xs = sorted(set(round(x, 6) for x in internal_column_xs))

        internal_column_tops = []
        for x in internal_column_xs:
            existing = next((node for node in bottom_nodes if abs(node.x - x) < 1e-6), None)
            if existing is not None:
                top_node = existing
            else:
                top_node = Node(node_id, x, bottom_y_at_x(x))
                structure.add_node(top_node)
                node_id += 1
            internal_column_tops.append(top_node)

        bottom_chord_nodes = sorted(bottom_nodes + [node for node in internal_column_tops if node not in bottom_nodes], key=lambda node: node.x)

        left_base = Node(node_id, left_column_offset, 0)
        structure.add_node(left_base)
        node_id += 1

        right_base = Node(node_id, outside_span - right_column_offset, 0)
        structure.add_node(right_base)
        node_id += 1

        internal_column_bases = []
        for top_node in internal_column_tops:
            base = Node(node_id, top_node.x, 0)
            structure.add_node(base)
            node_id += 1
            internal_column_bases.append(base)

        crane_load_y = left_column_bottom_chord_node.y - 2500.0
        left_crane_node = None
        right_crane_node = None
        if crane_rating_t > 0.0 and crane_load_y > 1e-6:
            left_crane_node = Node(node_id, left_column_bottom_chord_node.x, crane_load_y)
            structure.add_node(left_crane_node)
            node_id += 1
            right_crane_node = Node(node_id, right_column_bottom_chord_node.x, crane_load_y)
            structure.add_node(right_crane_node)
            node_id += 1

        element_id = 1

        def add_element(n1, n2, group, profile, analysis_i_factor=1.0):
            nonlocal element_id
            structure.add_element(FrameElement2D(element_id, n1, n2, E, profile, group, analysis_i_factor=analysis_i_factor))
            element_id += 1

        def pinned_connection_node(host_node):
            nonlocal node_id
            node = Node(node_id, host_node.x, host_node.y, ux=host_node.ux, uy=host_node.uy)
            structure.add_node(node)
            node_id += 1
            return node

        def existing_node_at(x, y, tolerance=1e-6):
            for node in structure.nodes:
                if abs(node.x - x) <= tolerance and abs(node.y - y) <= tolerance:
                    return node
            return None

        def node_at_or_new(x, y):
            nonlocal node_id
            existing = existing_node_at(x, y)
            if existing is not None:
                return existing
            node = Node(node_id, x, y)
            structure.add_node(node)
            node_id += 1
            return node

        def add_canopy(side, length_mm, canopy_type, side_eave_height, side_pitch, side_shallow_depth, side_top_profile, side_bottom_profile, side_web_profile, side_lean_column_profile):
            nonlocal node_id
            if length_mm <= 0.0:
                return [], [], [], None
            outward = -1.0 if side == "left" else 1.0
            prefix = "LEFT" if side == "left" else "RIGHT"
            default_anchor_top = top_nodes[0] if side == "left" else top_nodes[-1]
            if side_eave_height is None or abs(side_eave_height - default_anchor_top.y) <= 1e-6:
                anchor_top = default_anchor_top
            else:
                anchor_top = Node(node_id, default_anchor_top.x, side_eave_height)
                structure.add_node(anchor_top)
                node_id += 1
            target_canopy_web_angle_deg = 40.0
            target_canopy_panel = side_shallow_depth / math.tan(math.radians(target_canopy_web_angle_deg))
            canopy_panels = max(1, round(length_mm / max(target_canopy_panel, 1.0)))
            panel_len = length_mm / canopy_panels
            pitch_drop = math.tan(math.radians(side_pitch))
            outer_top_y = anchor_top.y - pitch_drop * length_mm
            is_lean_to = str(canopy_type).lower().startswith("lean")
            anchor_bottom_y = anchor_top.y - side_shallow_depth
            outer_bottom_y = outer_top_y - side_shallow_depth if is_lean_to else outer_top_y - side_shallow_depth
            bottom_y = anchor_bottom_y if is_lean_to else outer_bottom_y
            anchor_bottom = node_at_or_new(anchor_top.x, bottom_y)

            canopy_top = [anchor_top]
            canopy_bottom = [anchor_bottom]
            for i in range(canopy_panels + 1):
                x = anchor_top.x + outward * i * panel_len
                if i > 0:
                    top_node = Node(node_id, x, anchor_top.y - pitch_drop * i * panel_len)
                    structure.add_node(top_node)
                    node_id += 1
                    canopy_top.append(top_node)
                    if is_lean_to:
                        bottom_node_y = top_node.y - side_shallow_depth
                    else:
                        bottom_node_y = bottom_y
                    bottom_node = Node(node_id, x, bottom_node_y)
                    structure.add_node(bottom_node)
                    node_id += 1
                    canopy_bottom.append(bottom_node)

            for n1, n2 in zip(canopy_top[:-1], canopy_top[1:]):
                add_element(n1, n2, f"{prefix}_CANOPY_TOP", side_top_profile)
            for n1, n2 in zip(canopy_bottom[:-1], canopy_bottom[1:]):
                add_element(n1, n2, f"{prefix}_CANOPY_BOTTOM", side_bottom_profile)
            for i in range(canopy_panels):
                if i % 2 == 0:
                    add_element(pinned_connection_node(canopy_bottom[i]), pinned_connection_node(canopy_top[i + 1]), f"{prefix}_CANOPY_WEB", side_web_profile)
                else:
                    add_element(pinned_connection_node(canopy_top[i]), pinned_connection_node(canopy_bottom[i + 1]), f"{prefix}_CANOPY_WEB", side_web_profile)
            add_element(pinned_connection_node(canopy_top[0]), pinned_connection_node(canopy_bottom[0]), f"{prefix}_CANOPY_POST", side_web_profile)
            add_element(pinned_connection_node(canopy_top[-1]), pinned_connection_node(canopy_bottom[-1]), f"{prefix}_CANOPY_POST", side_web_profile)
            outer_column_base = None
            if is_lean_to:
                outer_column_base = Node(node_id, canopy_bottom[-1].x, 0)
                structure.add_node(outer_column_base)
                node_id += 1
                add_element(outer_column_base, canopy_bottom[-1], f"{prefix}_LEAN_COLUMN", side_lean_column_profile)
            column_nodes = [anchor_top, anchor_bottom]
            return canopy_top, canopy_bottom, column_nodes, outer_column_base

        top_profile_depth = self.profile_depth_mm(top_profile)

        def rafter_haunch_i_factor(n1, n2):
            if not is_rafter or rafter_haunch_length <= 0.0 or top_profile_depth <= 0.0:
                return 1.0
            mid_x = (n1.x + n2.x) / 2.0
            mid_y = (n1.y + n2.y) / 2.0
            left_distance = math.hypot(mid_x - top_nodes[0].x, mid_y - top_nodes[0].y)
            right_distance = math.hypot(mid_x - top_nodes[-1].x, mid_y - top_nodes[-1].y)
            distance_from_eave = min(left_distance, right_distance)
            if distance_from_eave >= rafter_haunch_length:
                return 1.0
            taper = 1.0 - distance_from_eave / rafter_haunch_length
            effective_depth = top_profile_depth * (1.0 + taper)
            return (effective_depth / top_profile_depth) ** 3

        for i in range(len(top_nodes) - 1):
            add_element(top_nodes[i], top_nodes[i + 1], "TOP", top_profile, analysis_i_factor=rafter_haunch_i_factor(top_nodes[i], top_nodes[i + 1]))

        if not is_rafter:
            for i in range(len(bottom_chord_nodes) - 1):
                add_element(bottom_chord_nodes[i], bottom_chord_nodes[i + 1], "BOTTOM", bottom_profile)

            def warren_web_top_node_for_bottom_x(x):
                for i, (b1, b2) in enumerate(zip(bottom_nodes[:-1], bottom_nodes[1:])):
                    if min(b1.x, b2.x) - 1e-6 <= x <= max(b1.x, b2.x) + 1e-6:
                        return top_nodes[i + 1] if i % 2 == 0 else top_nodes[i]
                return min(top_nodes, key=lambda node: abs(node.x - x))

            for i in range(num_panels):
                if i % 2 == 0:
                    add_element(pinned_connection_node(bottom_nodes[i]), pinned_connection_node(top_nodes[i + 1]), "WEB", web_profile)
                else:
                    add_element(pinned_connection_node(top_nodes[i]), pinned_connection_node(bottom_nodes[i + 1]), "WEB", web_profile)

            add_element(pinned_connection_node(top_nodes[0]), pinned_connection_node(bottom_nodes[0]), "POST", web_profile)
            add_element(pinned_connection_node(top_nodes[-1]), pinned_connection_node(bottom_nodes[-1]), "POST", web_profile)

            # Central apex vertical web/post. This keeps the ridge node tied directly
            # to the bottom chord and gives a more realistic Warren truss apex detail.
            apex_index = num_panels // 2
            add_element(pinned_connection_node(top_nodes[apex_index]), pinned_connection_node(bottom_nodes[apex_index]), "APEX", web_profile)

        for base_node, top_node in zip(internal_column_bases, internal_column_tops):
            add_element(base_node, pinned_connection_node(top_node), "INTERNAL_COLUMN", internal_column_profile)
            if not is_rafter:
                web_target_node = warren_web_top_node_for_bottom_x(top_node.x)
                add_element(pinned_connection_node(top_node), pinned_connection_node(web_target_node), "WEB", web_profile)

        left_canopy_top_nodes, left_canopy_bottom_nodes, left_canopy_column_node, left_lean_base = add_canopy(
            "left", left_canopy_length, left_canopy_type, left_canopy_eave_height, left_canopy_pitch, left_canopy_shallow_depth,
            left_canopy_top_profile, left_canopy_bottom_profile, left_canopy_web_profile, left_lean_column_profile,
        )
        right_canopy_top_nodes, right_canopy_bottom_nodes, right_canopy_column_node, right_lean_base = add_canopy(
            "right", right_canopy_length, right_canopy_type, right_canopy_eave_height, right_canopy_pitch, right_canopy_shallow_depth,
            right_canopy_top_profile, right_canopy_bottom_profile, right_canopy_web_profile, right_lean_column_profile,
        )
        roof_top_nodes = (
            list(reversed(left_canopy_top_nodes[1:]))
            + top_nodes
            + right_canopy_top_nodes[1:]
        )
        continuous_roof_top_nodes = (
            (list(reversed(left_canopy_top_nodes[1:])) if not left_canopy_is_free_roof else [])
            + top_nodes
            + (right_canopy_top_nodes[1:] if not right_canopy_is_free_roof else [])
        )

        def add_column_segments(nodes, group, profile):
            ordered = []
            for node in sorted([n for n in nodes if n is not None], key=lambda n: n.y):
                if ordered and abs(node.y - ordered[-1].y) < 1e-6:
                    continue
                ordered.append(node)
            for n1, n2 in zip(ordered[:-1], ordered[1:]):
                add_element(n1, n2, group, profile)

        add_column_segments(
            [left_base, left_crane_node, left_column_bottom_chord_node] + left_canopy_column_node + [top_nodes[0]],
            "LEFT_COLUMN",
            left_column_profile,
        )
        add_column_segments(
            [right_base, right_crane_node, right_column_bottom_chord_node] + right_canopy_column_node + [top_nodes[-1]],
            "RIGHT_COLUMN",
            right_column_profile,
        )

        left_support = self.left_support_combo.currentText()
        right_support = self.right_support_combo.currentText()
        eave_x_restraint = self.eave_x_restraint_combo.currentText()
        eave_bracing_info = None
        active_eave_restraint = eave_x_restraint
        active_eave_x_spring_kn_mm = eave_x_spring_kn_mm
        if eave_x_restraint == "End wall bracing approx":
            eave_bracing_info = self.end_wall_bracing_eave_spring(
                building_length_m,
                bay_size_m,
                end_wall_brace_dia_mm,
            )
            active_eave_restraint = "Both eaves"
            active_eave_x_spring_kn_mm = eave_bracing_info["effective_spring_kn_mm"]
        eave_x_spring_n_mm = active_eave_x_spring_kn_mm * 1000.0
        self.apply_base_support(structure, left_base, left_support)
        self.apply_base_support(structure, right_base, right_support)
        eave_restraint_nodes = []
        eave_restraint_types = []
        if active_eave_restraint in {"Left eave", "Both eaves"}:
            if eave_x_spring_n_mm > 0.0:
                structure.add_spring(top_nodes[0].ux, eave_x_spring_n_mm)
                eave_restraint_types.append("X spring")
            else:
                structure.add_support(top_nodes[0].ux)
                eave_restraint_types.append("X restraint")
            eave_restraint_nodes.append(top_nodes[0])
        if active_eave_restraint in {"Right eave", "Both eaves"}:
            if eave_x_spring_n_mm > 0.0:
                structure.add_spring(top_nodes[-1].ux, eave_x_spring_n_mm)
                eave_restraint_types.append("X spring")
            else:
                structure.add_support(top_nodes[-1].ux)
                eave_restraint_types.append("X restraint")
            eave_restraint_nodes.append(top_nodes[-1])
        for internal_base in internal_column_bases:
            self.apply_base_support(structure, internal_base, "Pinned")
        for lean_base in [left_lean_base, right_lean_base]:
            if lean_base is not None:
                self.apply_base_support(structure, lean_base, "Pinned")
        self.purlin_layout = self.build_purlin_layout(
            top_nodes,
            wind_inputs.wind_region,
            wind_inputs.importance_level,
        )
        self.purlin_check = self.purlin_check_from_inputs(
            self.purlin_layout,
            bay_size_m,
            roof_loads,
            wind_inputs,
            self.purlin_span_type_combo.currentText(),
        )
        self.wall_girt_check = self.wall_girt_check_from_inputs(
            wind_inputs,
            bay_size_m,
            eave,
            self.wall_girt_span_type_combo.currentText(),
        )
        self.member_check_restraints = self.member_check_restraint_lengths(
            wind_inputs,
            left_base,
            right_base,
            bottom_nodes,
            self.purlin_layout,
            column_top_deduction_mm=(2.0 * top_profile_depth if is_rafter and rafter_haunch_length > 0.0 else 0.0),
        )
        internal_lengths = [abs(top.y - base.y) for base, top in zip(internal_column_bases, internal_column_tops)]
        if internal_lengths:
            internal_length = max(internal_lengths)
            self.member_check_restraints["INTERNAL_COLUMN"] = {
                "major_mm": internal_length,
                "minor_mm": internal_length,
                "ke": 1.0,
                "am": 1.0,
                "kt": 1.0,
                "kl": 1.0,
                "kr": 1.0,
                "clad": False,
            }
        cladding = self.column_cladding_restraints(wind_inputs)
        wall_cladding_segments = []
        if cladding.get("LEFT_COLUMN"):
            wall_cladding_segments.append({"x": top_nodes[0].x, "y1": left_base.y, "y2": top_nodes[0].y, "label": "Left wall"})
        if cladding.get("RIGHT_COLUMN"):
            wall_cladding_segments.append({"x": top_nodes[-1].x, "y1": right_base.y, "y2": top_nodes[-1].y, "label": "Right wall"})
        if left_lean_outer_wall_clad and left_lean_base is not None and left_canopy_bottom_nodes:
            wall_cladding_segments.append({"x": left_lean_base.x, "y1": left_lean_base.y, "y2": left_canopy_bottom_nodes[-1].y, "label": "Left lean-to wall"})
        if right_lean_outer_wall_clad and right_lean_base is not None and right_canopy_bottom_nodes:
            wall_cladding_segments.append({"x": right_lean_base.x, "y1": right_lean_base.y, "y2": right_canopy_bottom_nodes[-1].y, "label": "Right lean-to wall"})

        selected_combination = self.load_combinations[self.load_combination_combo.currentIndex()]
        wind_result_cache = {}

        def get_wind_result(cpe_case):
            source_direction = self.wind_source_direction_for_case(wind_inputs.orientation, cpe_case)
            if source_direction not in wind_result_cache:
                wind_result_cache[source_direction] = calculate_wind(
                    replace(wind_inputs, orientation=source_direction)
                )
            return wind_result_cache[source_direction]

        continuous_roof_wind_result_cache = {}

        def get_continuous_roof_wind_result(cpe_case):
            source_direction = self.wind_source_direction_for_case(wind_inputs.orientation, cpe_case)
            if source_direction not in continuous_roof_wind_result_cache:
                continuous_roof_wind_result_cache[source_direction] = calculate_wind(
                    replace(
                        wind_inputs,
                        orientation=source_direction,
                        left_canopy_length_m=(left_canopy_length / 1000.0 if not left_canopy_is_free_roof else 0.0),
                        right_canopy_length_m=(right_canopy_length / 1000.0 if not right_canopy_is_free_roof else 0.0),
                    )
                )
            return continuous_roof_wind_result_cache[source_direction]

        def slope_path_length_m(nodes):
            total = 0.0
            for n1, n2 in zip(nodes[:-1], nodes[1:]):
                total += math.hypot((n2.x - n1.x) / 1000.0, (n2.y - n1.y) / 1000.0)
            return total

        def wind_interval_load_kn(strips, start_m, end_m, scale):
            total_kn = 0.0
            for a, b, q_kn_m in strips:
                overlap = max(0.0, min(end_m, b) - max(start_m, a))
                total_kn += q_kn_m * scale * overlap
            return total_kn

        def cumulative_roof_distances(nodes):
            distances = [0.0]
            for n1, n2 in zip(nodes[:-1], nodes[1:]):
                distances.append(distances[-1] + math.hypot((n2.x - n1.x) / 1000.0, (n2.y - n1.y) / 1000.0))
            return distances

        def continuous_roof_strips_for_nodes(wind_result, cpe_case, nodes):
            case = str(cpe_case).strip()
            if case.startswith("End") or wind_result.inputs.roof_pitch_deg < 10.0 or len(nodes) < 2:
                return wind_result.case_roof_strips(cpe_case)

            distances = cumulative_roof_distances(nodes)
            total = distances[-1]
            if total <= 1e-9:
                return []
            ridge_index = max(range(len(nodes)), key=lambda index: nodes[index].y)
            ridge_s = min(max(distances[ridge_index], 0.0), total)
            use_max = case.endswith("+")

            def zone_line_load(name):
                zone = next((item for item in wind_result.roof_zones if item.name == name), None)
                if zone is None:
                    return 0.0
                return zone.line_load_max_kn_m if use_max else zone.line_load_min_kn_m

            upwind_q = zone_line_load("Upwind rafter")
            downwind_q = zone_line_load("Downwind rafter")
            if case.startswith("Right"):
                strips = [(0.0, ridge_s, downwind_q), (ridge_s, total, upwind_q)]
            else:
                strips = [(0.0, ridge_s, upwind_q), (ridge_s, total, downwind_q)]
            return [(a, b, q) for a, b, q in strips if b - a > 1e-9]

        def apply_roof_wind_strips_to_nodes(target_structure, nodes, strips, wind_result, wind_pressure):
            scale = wind_pressure_scale(wind_result, wind_pressure)
            distances = cumulative_roof_distances(nodes)
            arrows = []
            for index, (n1, n2) in enumerate(zip(nodes[:-1], nodes[1:])):
                s1 = distances[index]
                s2 = distances[index + 1]
                panel_len = max(s2 - s1, 1e-9)
                total_kn = wind_interval_load_kn(strips, s1, s2, scale)
                total_n = total_kn * 1000.0
                target_structure.add_load(n1.uy, total_n / 2.0)
                target_structure.add_load(n2.uy, total_n / 2.0)
                display_n_per_m = (total_kn / panel_len) * 1000.0
                arrows.append(((n1.x + n2.x) / 2.0, (n1.y + n2.y) / 2.0 + 800.0, 0.0, display_n_per_m))
            return arrows

        def canopy_wind_model(side_key):
            return "AS/NZS B.5 free roof" if canopy_uses_free_roof(side_key) else "Continuous roof Cpe + chamber Cpi"

        def canopy_uses_free_roof(side_key):
            return left_canopy_is_free_roof if side_key == "left" else right_canopy_is_free_roof

        def canopy_underside(side_key):
            label = left_canopy_underside if side_key == "left" else right_canopy_underside
            return "blocked" if str(label).startswith("Blocked") else "empty"

        def canopy_free_roof_part(side_key, cpe_case):
            case = str(cpe_case).strip()
            if case.startswith("End"):
                return "windward"
            left_wind = case.startswith("Left")
            if side_key == "left":
                return "windward" if left_wind else "leeward"
            return "leeward" if left_wind else "windward"

        def manual_canopy_cpn(side_key, cpe_case):
            use_uplift = str(cpe_case).strip().endswith("+")
            if side_key == "left":
                return left_canopy_cpn_uplift if use_uplift else -abs(left_canopy_cpn_downward)
            return right_canopy_cpn_uplift if use_uplift else -abs(right_canopy_cpn_downward)

        def canopy_free_roof_values(wind_result, side_key, cpe_case):
            use_uplift = str(cpe_case).strip().endswith("+")
            part = canopy_free_roof_part(side_key, cpe_case)
            underside = canopy_underside(side_key)
            pitch_deg = left_canopy_pitch if side_key == "left" else right_canopy_pitch
            cpn_as = free_roof_cpn(
                pitch_deg,
                wind_result.h_over_d,
                part,
                underside,
                use_uplift,
            )
            return cpn_as, -cpn_as

        def main_roof_free_roof_part(segment_mid_x, cpe_case):
            case = str(cpe_case).strip()
            if case.startswith("End"):
                return None
            apex_x = max(top_nodes, key=lambda node: node.y).x
            left_wind = case.startswith("Left")
            if left_wind:
                return "windward" if segment_mid_x <= apex_x else "leeward"
            return "leeward" if segment_mid_x <= apex_x else "windward"

        def main_roof_free_roof_values(wind_result, segment_mid_x, cpe_case):
            part = main_roof_free_roof_part(segment_mid_x, cpe_case)
            if part is None:
                return None
            use_uplift = str(cpe_case).strip().endswith("+")
            cpn_as = free_roof_cpn(
                wind_result.inputs.roof_pitch_deg,
                wind_result.h_over_d,
                part,
                "empty",
                use_uplift,
            )
            return cpn_as, -cpn_as

        def add_vertical_line_load_to_nodes(target_structure, nodes, line_kn_m):
            arrows = []
            for n1, n2 in zip(nodes[:-1], nodes[1:]):
                dx = (n2.x - n1.x) / 1000.0
                dy = (n2.y - n1.y) / 1000.0
                panel_len_m = math.hypot(dx, dy)
                total_n = line_kn_m * panel_len_m * 1000.0
                target_structure.add_load(n1.uy, total_n / 2.0)
                target_structure.add_load(n2.uy, total_n / 2.0)
                arrows.append(((n1.x + n2.x) / 2.0, (n1.y + n2.y) / 2.0 + 800.0, 0.0, total_n))
            return arrows

        def add_normal_line_load_to_nodes(target_structure, nodes, line_kn_m):
            arrows = []
            for n1, n2 in zip(nodes[:-1], nodes[1:]):
                dx_m = (n2.x - n1.x) / 1000.0
                dy_m = (n2.y - n1.y) / 1000.0
                panel_len_m = math.hypot(dx_m, dy_m)
                if panel_len_m <= 1e-12:
                    continue
                # Positive free-roof Cpn frame load is uplift, so choose the
                # segment normal with positive global-Y component.
                nx = -dy_m / panel_len_m
                ny = dx_m / panel_len_m
                if ny < 0.0:
                    nx = -nx
                    ny = -ny
                total_fx_n = line_kn_m * panel_len_m * 1000.0 * nx
                total_fy_n = line_kn_m * panel_len_m * 1000.0 * ny
                target_structure.add_load(n1.ux, total_fx_n / 2.0)
                target_structure.add_load(n1.uy, total_fy_n / 2.0)
                target_structure.add_load(n2.ux, total_fx_n / 2.0)
                target_structure.add_load(n2.uy, total_fy_n / 2.0)
                arrows.append(((n1.x + n2.x) / 2.0, (n1.y + n2.y) / 2.0 + 800.0, total_fx_n, total_fy_n))
            return arrows

        def apply_main_roof_free_roof_cpn(target_structure, wind_result, cpe_case, wind_pressure):
            scale = wind_pressure_scale(wind_result, wind_pressure)
            arrows = []
            for n1, n2 in zip(top_nodes[:-1], top_nodes[1:]):
                mid_x = (n1.x + n2.x) / 2.0
                values = main_roof_free_roof_values(wind_result, mid_x, cpe_case)
                if values is None:
                    continue
                _, frame_cpn = values
                line_kn_m = frame_cpn * wind_result.wu_kn_m2 * scale * wind_result.inputs.bay_size_m
                arrows.extend(add_normal_line_load_to_nodes(target_structure, [n1, n2], line_kn_m))
            return arrows

        def apply_canopy_wind_cpe(target_structure, wind_result, cpe_case, wind_pressure):
            scale = wind_pressure_scale(wind_result, wind_pressure)
            strips = wind_result.case_roof_strips(cpe_case)
            roof_length_m = wind_result.roof_slope_length_m
            arrows = []
            for side_key, canopy_nodes, start_m in [
                ("left", left_canopy_top_nodes, 0.0),
                ("right", right_canopy_top_nodes, roof_length_m - slope_path_length_m(right_canopy_top_nodes)),
            ]:
                side_slope_m = slope_path_length_m(canopy_nodes)
                if side_slope_m <= 1e-9:
                    continue
                if canopy_uses_free_roof(side_key):
                    _, frame_cpn = canopy_free_roof_values(wind_result, side_key, cpe_case)
                    line_kn_m = frame_cpn * wind_result.wu_kn_m2 * scale * wind_result.inputs.bay_size_m
                    arrows.extend(add_normal_line_load_to_nodes(target_structure, canopy_nodes, line_kn_m))
            return arrows

        def enclosed_cpi_for_case(cpi_case):
            return 0.10 if str(cpi_case).strip() == "Cpi +" else -0.20

        def chamber_wall_opening_area_m2(chamber_width_m, wall_key):
            height_m = max(eave / 1000.0, 0.0)
            if wall_key in {"left", "right"}:
                return max(building_length_m, 0.0) * height_m
            return max(chamber_width_m, 0.0) * height_m

        def table_5_1b_cpi_for_opening(wind_result, cpi_case, cpe_case, wall_key, opening_area_m2, other_open_area_m2):
            surface = wall_surface_for_cpe_case(cpe_case, wall_key)
            cpe = wall_cpe_for_surface(wind_result, surface)
            ratio = float("inf") if other_open_area_m2 <= 1e-12 else opening_area_m2 / other_open_area_m2
            ka = opening_area_reduction_factor(surface, opening_area_m2)
            kl = 1.0
            signed_case = str(cpi_case).strip()

            if ratio <= 0.5:
                return 0.0 if signed_case == "Cpi +" else -0.3, ratio, surface, ka, kl
            if ratio < 2.0:
                if surface == "windward_wall":
                    return 0.2 if signed_case == "Cpi +" else -0.1, ratio, surface, ka, kl
                return 0.0 if signed_case == "Cpi +" else -0.3, ratio, surface, ka, kl
            if surface == "windward_wall" and ratio < 3.0:
                multiplier = 0.7
            elif surface == "windward_wall" and ratio < 6.0:
                multiplier = 0.85
            else:
                multiplier = 1.0
            return multiplier * ka * kl * cpe, ratio, surface, ka, kl

        def chamber_cpi_from_openings(wind_result, cpi_case, cpe_case, wall_states, external_keys, chamber_width_m):
            openings = []
            for key in external_keys:
                if not bool(wall_states.get(key, False)):
                    area = chamber_wall_opening_area_m2(chamber_width_m, key)
                    if area > 1e-12:
                        openings.append((key, area))
            if not openings:
                return enclosed_cpi_for_case(cpi_case), "enclosed chamber Cpi"

            largest_area = max(area for _, area in openings)
            candidates = []
            for wall_key, area in openings:
                if abs(area - largest_area) > 1e-9:
                    continue
                other_area = sum(other for other_key, other in openings if other_key != wall_key)
                candidates.append(table_5_1b_cpi_for_opening(wind_result, cpi_case, cpe_case, wall_key, area, other_area))

            pick = max if str(cpi_case).strip() == "Cpi +" else min
            cpi, ratio, surface, ka, kl = pick(candidates, key=lambda item: item[0])
            ratio_text = "6+" if ratio >= 6.0 else f"{ratio:.2g}"
            return cpi, f"Table 5.1(B) {surface} opening r={ratio_text}, Ka {ka:.2f}"

        def main_chamber_cpi(wind_result, cpi_case, cpe_case):
            return chamber_cpi_from_openings(
                wind_result,
                cpi_case,
                cpe_case,
                {
                    "left": wind_result.inputs.left_wall_clad,
                    "right": wind_result.inputs.right_wall_clad,
                    "front": wind_result.inputs.front_wall_clad,
                    "back": wind_result.inputs.back_wall_clad,
                },
                ["left", "right", "front", "back"],
                outside_span / 1000.0,
            )

        def lean_outer_wall_is_clad(side_key):
            return left_lean_outer_wall_clad if side_key == "left" else right_lean_outer_wall_clad

        def lean_to_wall_states(side_key):
            return left_lean_wall_states if side_key == "left" else right_lean_wall_states

        def lean_to_frame_type(side_key):
            states = lean_to_wall_states(side_key)
            inner_key = "right" if side_key == "left" else "left"
            exterior_count = sum(1 for key, value in states.items() if key != inner_key and value)
            if exterior_count == 0:
                return "Roof Only"
            clad_count = exterior_count + (1 if states.get(inner_key) else 0)
            if clad_count == 4:
                return "Enclosed"
            if clad_count == 3:
                return "3 Sided"
            if clad_count == 2:
                return "2 Sided"
            if clad_count == 1:
                return "1 Sided"
            return "Roof Only"

        def lean_to_cpi_for_case(wind_result, cpi_case, cpe_case, side_key):
            states = lean_to_wall_states(side_key)
            external_keys = ["left", "front", "back"] if side_key == "left" else ["right", "front", "back"]
            chamber_width_m = left_canopy_length / 1000.0 if side_key == "left" else right_canopy_length / 1000.0
            return chamber_cpi_from_openings(wind_result, cpi_case, cpe_case, states, external_keys, chamber_width_m)

        def lean_to_enclosure_summary(side_key):
            labels = (
                {"left": "outer", "front": "front", "back": "back"}
                if side_key == "left"
                else {"right": "outer", "front": "front", "back": "back"}
            )
            states = lean_to_wall_states(side_key)
            return ", ".join(f"{label} {'clad' if states.get(key) else 'open'}" for key, label in labels.items())

        def lean_outer_wall_nodes(side_key):
            if side_key == "left" and left_lean_base is not None and left_canopy_bottom_nodes:
                return left_lean_base, left_canopy_bottom_nodes[-1]
            if side_key == "right" and right_lean_base is not None and right_canopy_bottom_nodes:
                return right_lean_base, right_canopy_bottom_nodes[-1]
            return None, None

        def cpi_for_canopy_side(wind_result, cpi_case, cpe_case, side_key):
            if (side_key == "left" and left_has_lean_to) or (side_key == "right" and right_has_lean_to):
                cpi, _ = lean_to_cpi_for_case(wind_result, cpi_case, cpe_case, side_key)
                return cpi
            chamber_width_m = left_canopy_length / 1000.0 if side_key == "left" else right_canopy_length / 1000.0
            cpi, _ = chamber_cpi_from_openings(
                wind_result,
                cpi_case,
                cpe_case,
                {"left": False} if side_key == "left" else {"right": False},
                ["left"] if side_key == "left" else ["right"],
                chamber_width_m,
            )
            return cpi

        def add_horizontal_line_load(target_structure, base, top, line_kn_m, direction, arrow_count=5):
            if base is None or top is None:
                return []
            height_m = abs(top.y - base.y) / 1000.0
            if height_m <= 1e-9:
                return []
            total_n = line_kn_m * height_m * 1000.0 * direction
            target_structure.add_load(base.ux, total_n / 2.0)
            target_structure.add_load(top.ux, total_n / 2.0)
            arrows = []
            count = max(2, arrow_count)
            for i in range(count):
                t = (i + 0.5) / count
                x = base.x + (top.x - base.x) * t
                y = base.y + (top.y - base.y) * t
                arrows.append((x, y, total_n / count, 0.0))
            return arrows

        def add_signed_horizontal_line_load(target_structure, base, top, signed_line_kn_m, arrow_count=5):
            if abs(signed_line_kn_m) <= 1e-12:
                return []
            direction = +1.0 if signed_line_kn_m > 0.0 else -1.0
            return add_horizontal_line_load(target_structure, base, top, abs(signed_line_kn_m), direction, arrow_count)

        def signed_internal_cpi_line_kn_m(wind_result, cpi, wind_pressure):
            return (
                cpi
                * wind_result.wu_kn_m2
                * wind_pressure_scale(wind_result, wind_pressure)
                * internal_pressure_factor(wind_result, cpi)
                * wind_result.inputs.bay_size_m
            )

        def internal_wall_net_cpi_line(wind_result, side_key, cpi_case, wind_pressure, cpe_case):
            if side_key == "left":
                if not (left_has_lean_to and wind_result.inputs.left_wall_clad):
                    return 0.0, None
                lean_cpi_for = lambda case: lean_to_cpi_for_case(wind_result, case, cpe_case, "left")[0]
                main_sign = -1.0
                lean_sign = +1.0
            else:
                if not (right_has_lean_to and wind_result.inputs.right_wall_clad):
                    return 0.0, None
                lean_cpi_for = lambda case: lean_to_cpi_for_case(wind_result, case, cpe_case, "right")[0]
                main_sign = +1.0
                lean_sign = -1.0

            candidates = []
            for main_case in CPI_CASE_OPTIONS:
                main_cpi, _ = main_chamber_cpi(wind_result, main_case, cpe_case)
                main_line = signed_internal_cpi_line_kn_m(wind_result, main_cpi, wind_pressure)
                for lean_case in CPI_CASE_OPTIONS:
                    lean_cpi = lean_cpi_for(lean_case)
                    lean_line = signed_internal_cpi_line_kn_m(wind_result, lean_cpi, wind_pressure)
                    signed_line = main_sign * main_line + lean_sign * lean_line
                    candidates.append((signed_line, main_case, main_cpi, lean_case, lean_cpi))
            if not candidates:
                return 0.0, None
            pick = max if str(cpi_case).strip() == "Cpi +" else min
            signed_line, main_case, main_cpi, lean_case, lean_cpi = pick(candidates, key=lambda item: item[0])
            return signed_line, {
                "side": side_key,
                "main_case": main_case,
                "main_cpi": main_cpi,
                "lean_case": lean_case,
                "lean_cpi": lean_cpi,
            }

        def apply_wall_cpe_with_lean_to(target_structure, wind_result, cpe_case, wind_pressure):
            frame_type = wind_result.inputs.frame_type
            if frame_type == "No Wind":
                return []
            if frame_type == "Roof Only" and not (left_lean_outer_wall_clad or right_lean_outer_wall_clad):
                return []
            arrows = []
            if str(cpe_case).startswith("End"):
                side_cpe = wall_cpe_for_surface(wind_result, "side_wall")
                wall_area_m2 = wind_result.inputs.bay_size_m * wind_result.inputs.eave_height_m
                side_line = wall_cpe_line_load_kn_m(wind_result, side_cpe, wind_pressure, "side_wall", wall_area_m2)
                if frame_type not in {"Roof Only", "No Wind"}:
                    if wind_result.inputs.left_wall_clad and not left_has_lean_to:
                        arrows.extend(add_horizontal_line_load(target_structure, left_base, top_nodes[0], side_line, -1.0))
                    if wind_result.inputs.right_wall_clad and not right_has_lean_to:
                        arrows.extend(add_horizontal_line_load(target_structure, right_base, top_nodes[-1], side_line, +1.0))
                if left_lean_outer_wall_clad:
                    base, top = lean_outer_wall_nodes("left")
                    arrows.extend(add_horizontal_line_load(target_structure, base, top, side_line, -1.0))
                if right_lean_outer_wall_clad:
                    base, top = lean_outer_wall_nodes("right")
                    arrows.extend(add_horizontal_line_load(target_structure, base, top, side_line, +1.0))
                return arrows
            left_wind = str(cpe_case).startswith("Left")
            wall_area_m2 = wind_result.inputs.bay_size_m * wind_result.inputs.eave_height_m
            ww_line = wall_cpe_line_load_kn_m(wind_result, 0.7, wind_pressure, "windward_wall", wall_area_m2)
            lw_line = wall_cpe_line_load_kn_m(wind_result, wind_result.leeward_wall_cpe, wind_pressure, "leeward_wall", wall_area_m2)

            if frame_type not in {"Roof Only", "No Wind"}:
                if left_wind:
                    if wind_result.inputs.left_wall_clad and not left_lean_outer_wall_clad:
                        arrows.extend(add_horizontal_line_load(target_structure, left_base, top_nodes[0], ww_line, +1.0))
                    if wind_result.inputs.right_wall_clad and not right_lean_outer_wall_clad:
                        arrows.extend(add_horizontal_line_load(target_structure, right_base, top_nodes[-1], lw_line, +1.0))
                else:
                    if wind_result.inputs.right_wall_clad and not right_lean_outer_wall_clad:
                        arrows.extend(add_horizontal_line_load(target_structure, right_base, top_nodes[-1], ww_line, -1.0))
                    if wind_result.inputs.left_wall_clad and not left_lean_outer_wall_clad:
                        arrows.extend(add_horizontal_line_load(target_structure, left_base, top_nodes[0], lw_line, -1.0))

            if left_lean_outer_wall_clad:
                base, top = lean_outer_wall_nodes("left")
                arrows.extend(add_horizontal_line_load(target_structure, base, top, ww_line if left_wind else lw_line, +1.0 if left_wind else -1.0))
            if right_lean_outer_wall_clad:
                base, top = lean_outer_wall_nodes("right")
                arrows.extend(add_horizontal_line_load(target_structure, base, top, lw_line if left_wind else ww_line, +1.0 if left_wind else -1.0))
            return arrows

        def apply_wall_cpi_with_lean_to(target_structure, wind_result, cpi_case, wind_pressure, cpe_case):
            frame_type = wind_result.inputs.frame_type
            if frame_type == "No Wind":
                return []
            if frame_type == "Roof Only" and not (left_lean_outer_wall_clad or right_lean_outer_wall_clad):
                return []
            arrows = []
            main_cpi, _ = main_chamber_cpi(wind_result, cpi_case, cpe_case)
            main_line = wall_cpi_line_load_kn_m(wind_result, main_cpi, wind_pressure)
            main_left_direction = -1.0 if main_cpi > 0 else +1.0
            main_right_direction = +1.0 if main_cpi > 0 else -1.0

            if frame_type not in {"Roof Only", "No Wind"} and abs(main_line) > 1e-12:
                if wind_result.inputs.left_wall_clad and not left_has_lean_to:
                    arrows.extend(add_horizontal_line_load(target_structure, left_base, top_nodes[0], main_line, main_left_direction))
                if wind_result.inputs.right_wall_clad and not right_has_lean_to:
                    arrows.extend(add_horizontal_line_load(target_structure, right_base, top_nodes[-1], main_line, main_right_direction))

            left_net_line, _ = internal_wall_net_cpi_line(wind_result, "left", cpi_case, wind_pressure, cpe_case)
            arrows.extend(add_signed_horizontal_line_load(target_structure, left_base, top_nodes[0], left_net_line))
            right_net_line, _ = internal_wall_net_cpi_line(wind_result, "right", cpi_case, wind_pressure, cpe_case)
            arrows.extend(add_signed_horizontal_line_load(target_structure, right_base, top_nodes[-1], right_net_line))

            if left_lean_outer_wall_clad:
                base, top = lean_outer_wall_nodes("left")
                left_lean_cpi, _ = lean_to_cpi_for_case(wind_result, cpi_case, cpe_case, "left")
                left_lean_line = wall_cpi_line_load_kn_m(wind_result, left_lean_cpi, wind_pressure)
                direction = -1.0 if left_lean_cpi > 0 else +1.0
                arrows.extend(add_horizontal_line_load(target_structure, base, top, left_lean_line, direction))
            if right_lean_outer_wall_clad:
                base, top = lean_outer_wall_nodes("right")
                right_lean_cpi, _ = lean_to_cpi_for_case(wind_result, cpi_case, cpe_case, "right")
                right_lean_line = wall_cpi_line_load_kn_m(wind_result, right_lean_cpi, wind_pressure)
                direction = +1.0 if right_lean_cpi > 0 else -1.0
                arrows.extend(add_horizontal_line_load(target_structure, base, top, right_lean_line, direction))
            return arrows

        def internal_wall_cpi_summary_lines(wind_result, combination):
            lines = []
            for label, side_key in [("Left", "left"), ("Right", "right")]:
                signed_line, detail = internal_wall_net_cpi_line(
                    wind_result,
                    side_key,
                    combination.cpi_case,
                    combination.wind_pressure,
                    combination.cpe_case,
                )
                if detail is None or abs(signed_line) <= 1e-12:
                    continue
                direction = "+X" if signed_line > 0.0 else "-X"
                lines.append(
                    f"{label} internal wall Cpi: {signed_line:+.3f} kN/m ({direction}), "
                    f"main {detail['main_case']} {detail['main_cpi']:+.3f}, "
                    f"lean {detail['lean_case']} {detail['lean_cpi']:+.3f}"
                )
            return lines

        def canopy_wind_summary_lines(wind_result, combination, active_cpi):
            left_slope_m = slope_path_length_m(left_canopy_top_nodes)
            right_slope_m = slope_path_length_m(right_canopy_top_nodes)
            if left_slope_m <= 1e-9 and right_slope_m <= 1e-9:
                return []

            scale = wind_pressure_scale(wind_result, combination.wind_pressure)
            strips = continuous_roof_strips_for_nodes(wind_result, combination.cpe_case, continuous_roof_top_nodes)
            roof_length_m = slope_path_length_m(continuous_roof_top_nodes)
            lines = ["", "CANOPY / LEAN-TO WIND LOADS"]
            for label, side_key, side_slope_m, start_m, end_m in [
                ("Left", "left", left_slope_m, 0.0, left_slope_m),
                ("Right", "right", right_slope_m, roof_length_m - right_slope_m, roof_length_m),
            ]:
                if side_slope_m <= 1e-9:
                    continue
                if canopy_uses_free_roof(side_key):
                    cpn_as, frame_cpn = canopy_free_roof_values(wind_result, side_key, combination.cpe_case)
                    line_kn_m = frame_cpn * wind_result.wu_kn_m2 * scale * wind_result.inputs.bay_size_m
                    total_kn = line_kn_m * side_slope_m
                    part = canopy_free_roof_part(side_key, combination.cpe_case)
                    model_text = f"AS/NZS B.5 free roof, {part}, {canopy_underside(side_key)} under"
                    lines.append(f"{label:<5} wind model   : {model_text}")
                    lines.append(
                        f"{label:<5} Cpn net      : AS {cpn_as:+8.3f}, frame {frame_cpn:+.3f}, {line_kn_m:+.3f} kN/m "
                        f"({total_kn:+.3f} kN)"
                    )
                    continue
                cpi_source = "wall Cpi"
                if (side_key == "left" and left_has_lean_to) or (side_key == "right" and right_has_lean_to):
                    _, cpi_source = lean_to_cpi_for_case(wind_result, combination.cpi_case, combination.cpe_case, side_key)
                canopy_cpi = cpi_for_canopy_side(wind_result, combination.cpi_case, combination.cpe_case, side_key)
                canopy_cpi_line_kn_m = (
                    canopy_cpi
                    * wind_result.wu_kn_m2
                    * scale
                    * internal_pressure_factor(wind_result, canopy_cpi)
                    * wind_result.inputs.bay_size_m
                )
                cpe_total_kn = wind_interval_load_kn(strips, start_m, end_m, scale)
                cpe_avg_kn_m = cpe_total_kn / side_slope_m
                cpi_total_kn = canopy_cpi_line_kn_m * side_slope_m
                combined_total_kn = cpe_total_kn + cpi_total_kn
                combined_avg_kn_m = combined_total_kn / side_slope_m
                lines.append(
                    f"{label:<5} Cpe avg      : {cpe_avg_kn_m:+8.3f} kN/m over {side_slope_m:.3f} m "
                    f"({cpe_total_kn:+.3f} kN)"
                )
                lines.append(
                    f"{label:<5} {cpi_source:<18}: {canopy_cpi:+8.3f} coeff, {canopy_cpi_line_kn_m:+.3f} kN/m "
                    f"({cpi_total_kn:+.3f} kN)"
                )
                if (side_key == "left" and left_has_lean_to) or (side_key == "right" and right_has_lean_to):
                    lines.append(f"{label:<5} lean-to walls: {lean_to_enclosure_summary(side_key)}")
                lines.append(
                    f"{label:<5} combined avg: {combined_avg_kn_m:+8.3f} kN/m "
                    f"({combined_total_kn:+.3f} kN)"
                )
            return lines

        def apply_roof_and_canopy_cpi(target_structure, wind_result, cpi_case, wind_pressure, cpe_case):
            arrows = []
            main_cpi, _ = main_chamber_cpi(wind_result, cpi_case, cpe_case)
            arrows.extend(
                apply_roof_cpi_to_top_nodes(
                    target_structure,
                    top_nodes,
                    wind_result,
                    main_cpi,
                    wind_pressure,
                )
            )
            for side_key, canopy_nodes in [
                ("left", left_canopy_top_nodes),
                ("right", right_canopy_top_nodes),
            ]:
                if len(canopy_nodes) < 2:
                    continue
                if canopy_uses_free_roof(side_key):
                    continue
                canopy_cpi = cpi_for_canopy_side(wind_result, cpi_case, cpe_case, side_key)
                arrows.extend(
                    apply_roof_cpi_to_top_nodes(
                        target_structure,
                        canopy_nodes,
                        wind_result,
                        canopy_cpi,
                        wind_pressure,
                    )
                )
            return arrows

        def arrow_collector():
            class LoadOnlyStructure:
                def add_load(self, dof, value):
                    return None

            return LoadOnlyStructure()

        def apply_crane_load(target_structure, q_factor, crane_position="both"):
            if (
                crane_rating_t <= 0.0
                or q_factor <= 0.0
                or crane_position not in {"both", "left", "right"}
                or left_crane_node is None
                or right_crane_node is None
            ):
                return []
            load_kn = 17.0 * crane_rating_t * q_factor
            load_n = load_kn * 1000.0
            horizontal_kn = 1.25 * crane_rating_t * q_factor
            horizontal_n = horizontal_kn * 1000.0
            arrows = []
            crane_targets = []
            if crane_position in {"both", "left"}:
                crane_targets.append((left_crane_node, 300.0, -horizontal_n))
            if crane_position in {"both", "right"}:
                crane_targets.append((right_crane_node, -300.0, horizontal_n))
            for node, eccentricity_mm, horizontal_load_n in crane_targets:
                target_structure.add_load(node.uy, -load_n)
                target_structure.add_load(node.rz, eccentricity_mm * -load_n)
                target_structure.add_load(node.ux, horizontal_load_n)
                arrows.append((node.x + eccentricity_mm, node.y, 0.0, -load_n))
                arrows.append((node.x, node.y, horizontal_load_n, 0.0))
            return arrows

        def crane_summary_lines(q_factor, crane_position="both"):
            if crane_rating_t <= 0.0 or q_factor <= 0.0 or crane_position not in {"both", "left", "right"}:
                return []
            lines = ["", "CRANE LOAD"]
            if left_crane_node is None or right_crane_node is None:
                lines.append("Crane load not applied: bottom chord is less than 2500 mm above the base.")
                return lines
            load_kn = 17.0 * crane_rating_t * q_factor
            horizontal_kn = 1.25 * crane_rating_t * q_factor
            if crane_position == "left":
                target_text = "left column only"
                direction_text = "left column -X from inside face"
            elif crane_position == "right":
                target_text = "right column only"
                direction_text = "right column +X from inside face"
            else:
                target_text = "both main columns"
                direction_text = "left column -X, right column +X from inside faces"
            lines.append(f"{'Rating':<13}: {crane_rating_t:>8.3f} t")
            lines.append(f"{'Q factor':<13}: {q_factor:>8.3f}")
            lines.append(f"{'Position':<13}: {target_text}")
            lines.append(f"{'Column load':<13}: {load_kn:>8.3f} kN vertical per loaded column")
            lines.append(f"{'Horiz load':<13}: {horizontal_kn:>8.3f} kN per loaded column ({direction_text})")
            lines.append(f"{'Load height':<13}: {left_crane_node.y:>8.0f} mm above base, 2500 mm below bottom chord")
            lines.append(f"{'Eccentricity':<13}: {300.0:>8.0f} mm inward to portal")
            return lines

        def wind_case_for_single_label(label):
            if label.startswith("Wu Cpe Left +"):
                return "Left +", "ultimate"
            if label.startswith("Wu Cpe Left -"):
                return "Left -", "ultimate"
            if label.startswith("Wu Cpe Right +"):
                return "Right +", "ultimate"
            if label.startswith("Wu Cpe Right -"):
                return "Right -", "ultimate"
            if label.startswith("Ws Cpe Left +"):
                return "Left +", "serviceability"
            if label.startswith("Ws Cpe Left -"):
                return "Left -", "serviceability"
            if label.startswith("Ws Cpe Right +"):
                return "Right +", "serviceability"
            if label.startswith("Ws Cpe Right -"):
                return "Right -", "serviceability"
            if label.startswith("Wu Cpe End +"):
                return "End +", "ultimate"
            if label.startswith("Wu Cpe End -"):
                return "End -", "ultimate"
            if label.startswith("Ws Cpe End +"):
                return "End +", "serviceability"
            if label.startswith("Ws Cpe End -"):
                return "End -", "serviceability"
            return None, None

        def build_single_load_arrows():
            cases = {}

            for label, kpa, placement in roof_loads.factored_components(1.0, 1.0):
                if label in {"G", "Q", "Solar", "Fire service", "HVAC", "Other"}:
                    cases[label] = apply_roof_vertical_loads_in_zones(
                        arrow_collector(),
                        top_nodes,
                        kpa * roof_loads.bay_size_m,
                        placement.side,
                        placement.from_percent,
                        placement.to_percent,
                    )

            left_wall_arrow = apply_wall_horizontal_loads(arrow_collector(), left_base, top_nodes[0], wall_loads.left_plus_x, direction=+1)
            right_wall_arrow = apply_wall_horizontal_loads(arrow_collector(), right_base, top_nodes[-1], wall_loads.right_minus_x, direction=-1)
            if left_wall_arrow:
                cases["Left wall G"] = left_wall_arrow
            if right_wall_arrow:
                cases["Right wall G"] = right_wall_arrow

            for crane_label, crane_position in [
                ("Crane Q both", "both"),
                ("Crane Q left", "left"),
                ("Crane Q right", "right"),
            ]:
                crane_arrows = apply_crane_load(arrow_collector(), 1.0, crane_position)
                if crane_arrows:
                    cases[crane_label] = crane_arrows

            self_weight_arrows = apply_structure_self_weight(arrow_collector(), structure.elements, 1.0).arrows
            if self_weight_arrows:
                cases["Self weight"] = self_weight_arrows

            for side_label, canopy_nodes, side_canopy_loads in [
                ("Left canopy", left_canopy_top_nodes, left_canopy_loads),
                ("Right canopy", right_canopy_top_nodes, right_canopy_loads),
            ]:
                if len(canopy_nodes) < 2:
                    continue
                for label, kpa, _ in side_canopy_loads.factored_components(1.0, 1.0):
                    if label in {"G", "Q", "Solar"}:
                        arrows = apply_roof_vertical_loads(
                            arrow_collector(),
                            canopy_nodes,
                            kpa * side_canopy_loads.bay_size_m,
                        )
                        if arrows:
                            cases[f"{side_label} {label}"] = arrows

            for label in ["Wu Cpe Left +", "Wu Cpe Left -", "Wu Cpe Right +", "Wu Cpe Right -", "Wu Cpe End +", "Wu Cpe End -", "Ws Cpe Left +", "Ws Cpe Left -", "Ws Cpe Right +", "Ws Cpe Right -", "Ws Cpe End +", "Ws Cpe End -"]:
                cpe_case, pressure = wind_case_for_single_label(label)
                wind_result = get_wind_result(cpe_case)
                continuous_roof_wind_result = get_continuous_roof_wind_result(cpe_case)
                arrows = []
                if main_roof_is_free_roof and not str(cpe_case).startswith("End"):
                    arrows.extend(apply_main_roof_free_roof_cpn(arrow_collector(), continuous_roof_wind_result, cpe_case, pressure))
                else:
                    continuous_strips = continuous_roof_strips_for_nodes(continuous_roof_wind_result, cpe_case, continuous_roof_top_nodes)
                    arrows.extend(apply_roof_wind_strips_to_nodes(arrow_collector(), continuous_roof_top_nodes, continuous_strips, continuous_roof_wind_result, pressure))
                arrows.extend(apply_canopy_wind_cpe(arrow_collector(), wind_result, cpe_case, pressure))
                arrows.extend(apply_wall_cpe_with_lean_to(arrow_collector(), wind_result, cpe_case, pressure))
                if arrows:
                    cases[label] = arrows

            for cpe_case in CPE_CASE_OPTIONS:
                wind_result = get_wind_result(cpe_case)
                for cpi_case in CPI_CASE_OPTIONS:
                    arrows = []
                    arrows.extend(apply_roof_and_canopy_cpi(arrow_collector(), wind_result, cpi_case, "ultimate", cpe_case))
                    arrows.extend(apply_wall_cpi_with_lean_to(arrow_collector(), wind_result, cpi_case, "ultimate", cpe_case))
                    if arrows:
                        cases[f"Wu {cpi_case} with {cpe_case}"] = arrows
                    sls_arrows = []
                    sls_arrows.extend(apply_roof_and_canopy_cpi(arrow_collector(), wind_result, cpi_case, "serviceability", cpe_case))
                    sls_arrows.extend(apply_wall_cpi_with_lean_to(arrow_collector(), wind_result, cpi_case, "serviceability", cpe_case))
                    if sls_arrows:
                        cases[f"Ws {cpi_case} with {cpe_case}"] = sls_arrows

            return cases

        def apply_combination_loads(combination, include_summary=False, target_structure=None):
            load_structure = target_structure or structure
            load_structure.loads = {}
            basic_load_case = BasicLoadCase(roof_loads, wall_loads)
            load_application = basic_load_case.apply_factored(
                load_structure,
                top_nodes,
                left_base,
                right_base,
                combination.g_factor,
                combination.q_factor,
            )
            arrows = list(load_application.arrows)
            for side_label, canopy_nodes, side_canopy_loads in [
                ("Left canopy", left_canopy_top_nodes, left_canopy_loads),
                ("Right canopy", right_canopy_top_nodes, right_canopy_loads),
            ]:
                if len(canopy_nodes) < 2:
                    continue
                for _, factored_kpa, _ in side_canopy_loads.factored_components(combination.g_factor, combination.q_factor):
                    arrows.extend(
                        apply_roof_vertical_loads(
                            load_structure,
                            canopy_nodes,
                            factored_kpa * side_canopy_loads.bay_size_m,
                        )
                    )
            self_weight_application = apply_structure_self_weight(
                load_structure,
                load_structure.elements,
                combination.g_factor,
            )
            arrows.extend(self_weight_application.arrows)
            arrows.extend(apply_crane_load(load_structure, combination.q_factor, combination.crane_position))

            summary_lines = []
            if include_summary:
                summary_lines = [
                    "LOAD COMBINATION",
                    combination.name,
                    "",
                ]
                summary_lines.extend(load_application.summary_lines)
                if snow_loads["enabled"]:
                    q_source = "snow" if snow_loads["service_kpa"] >= entered_q_kpa else "entered Q"
                    summary_lines.append("")
                    summary_lines.append("SNOW LOAD")
                    summary_lines.append(f"{'Region':<13}: {snow_loads['region']}")
                    summary_lines.append(f"{'Importance':<13}: {snow_loads['importance']} (from wind inputs)")
                    summary_lines.append(f"{'h0':<13}: {snow_loads['h0_m']:>8.3f} m")
                    summary_lines.append(f"{'Roof pitch':<13}: {snow_loads['pitch_deg']:>8.3f} deg")
                    summary_lines.append(f"{'K1 / kp':<13}: {snow_loads['k1']:>8.3f} / {snow_loads['kp']:.3f}")
                    summary_lines.append(f"{'mu':<13}: {snow_loads['mu']:>8.3f}")
                    summary_lines.append(f"{'sg':<13}: {snow_loads['sg_kpa']:>8.3f} kPa")
                    summary_lines.append(f"{'LLsn ult':<13}: {snow_loads['ultimate_kpa']:>8.3f} kPa")
                    summary_lines.append(f"{'LLsn serv':<13}: {snow_loads['service_kpa']:>8.3f} kPa")
                    summary_lines.append(f"{'Entered Q':<13}: {entered_q_kpa:>8.3f} kPa")
                    summary_lines.append(f"{'Q used':<13}: {effective_q_kpa:>8.3f} kPa ({q_source})")
                summary_lines.extend(crane_summary_lines(combination.q_factor, combination.crane_position))
                for side_label, side_length, side_canopy_loads in [
                    ("LEFT CANOPY LOAD INPUTS", left_canopy_length, left_canopy_loads),
                    ("RIGHT CANOPY LOAD INPUTS", right_canopy_length, right_canopy_loads),
                ]:
                    if side_length > 0.0 and any(abs(value) > 1e-12 for value in [side_canopy_loads.g, side_canopy_loads.q, side_canopy_loads.solar]):
                        summary_lines.append("")
                        summary_lines.append(side_label)
                        summary_lines.append(f"{'G':<13}: {side_canopy_loads.g:>8.3f} kPa")
                        summary_lines.append(f"{'Q':<13}: {side_canopy_loads.q:>8.3f} kPa")
                        summary_lines.append(f"{'Solar':<13}: {side_canopy_loads.solar:>8.3f} kPa")
                        summary_lines.append(f"{'Factors':<13}: G x {combination.g_factor:.3f}, Q x {combination.q_factor:.3f}")
                summary_lines.append("")
                summary_lines.extend(self_weight_application.summary_lines)

            if combination.includes_wind:
                wind_result = get_wind_result(combination.cpe_case)
                continuous_roof_wind_result = get_continuous_roof_wind_result(combination.cpe_case)
                if main_roof_is_free_roof and not str(combination.cpe_case).startswith("End"):
                    arrows.extend(
                        apply_main_roof_free_roof_cpn(
                            load_structure,
                            continuous_roof_wind_result,
                            combination.cpe_case,
                            combination.wind_pressure,
                        )
                    )
                else:
                    continuous_strips = continuous_roof_strips_for_nodes(
                        continuous_roof_wind_result,
                        combination.cpe_case,
                        continuous_roof_top_nodes,
                    )
                    arrows.extend(
                        apply_roof_wind_strips_to_nodes(
                            load_structure,
                            continuous_roof_top_nodes,
                            continuous_strips,
                            continuous_roof_wind_result,
                            combination.wind_pressure,
                        )
                    )
                arrows.extend(apply_canopy_wind_cpe(load_structure, wind_result, combination.cpe_case, combination.wind_pressure))
                arrows.extend(apply_wall_cpe_with_lean_to(load_structure, wind_result, combination.cpe_case, combination.wind_pressure))
                arrows.extend(
                    apply_roof_and_canopy_cpi(
                        load_structure,
                        wind_result,
                        combination.cpi_case,
                        combination.wind_pressure,
                        combination.cpe_case,
                    )
                )
                arrows.extend(apply_wall_cpi_with_lean_to(load_structure, wind_result, combination.cpi_case, combination.wind_pressure, combination.cpe_case))
                if include_summary:
                    effective_cpi_case = effective_cpi_case_for_wind(
                        wind_result,
                        combination.cpi_case,
                        combination.cpe_case,
                    )
                    active_cpi, active_cpi_source = main_chamber_cpi(
                        wind_result,
                        combination.cpi_case,
                        combination.cpe_case,
                    )
                    applied_kce, applied_kci = action_combination_factors(wind_result, active_cpi)
                    summary_lines.append("")
                    summary_lines.append(
                        f"APPLIED WIND CASES: {combination.cpe_case} + {combination.cpi_case} "
                        f"({combination.wind_pressure})"
                    )
                    summary_lines.append(f"Section north-up    : {wind_inputs.orientation}")
                    summary_lines.append(f"Wind source dir     : {wind_result.inputs.orientation}")
                    if effective_cpi_case != combination.cpi_case:
                        summary_lines.append(f"Open-side Cpi rule  : {combination.cpi_case} overridden to {effective_cpi_case}")
                    summary_lines.append(f"Main chamber Cpi    : {active_cpi:+.3f} ({active_cpi_source})")
                    summary_lines.append(f"Applied Kce / Kci   : {applied_kce:.3f} / {applied_kci:.3f}")
                    summary_lines.extend(internal_wall_cpi_summary_lines(wind_result, combination))
                    summary_lines.extend(canopy_wind_summary_lines(continuous_roof_wind_result, combination, active_cpi))
                    summary_lines.extend(wind_result.summary_lines())
            elif include_summary:
                summary_lines.append("")
                summary_lines.append("Wind: not included in this combination")

            return arrows, summary_lines

        try:
            load_arrows, load_summary_lines = apply_combination_loads(selected_combination, include_summary=True)
        except Exception as exc:
            self.results_box.setText(f"Load/wind input calculation error: {exc}")
            return

        try:
            d, R, F, analysis_info = self.solve_structure(structure)
        except np.linalg.LinAlgError:
            self.results_box.setText(
                "Solver error: stiffness matrix is singular.\n\n"
                "Try using fixed supports, checking the section properties, or reviewing the generated geometry."
            )
            return

        lean_bases = [base for base in [left_lean_base, right_lean_base] if base is not None]
        foundation_nodes = [("Left column", left_base), ("Right column", right_base)]
        foundation_nodes.extend((f"Internal column {index + 1}", base) for index, base in enumerate(internal_column_bases))
        if left_lean_base is not None:
            foundation_nodes.append(("Left lean-to column", left_lean_base))
        if right_lean_base is not None:
            foundation_nodes.append(("Right lean-to column", right_lean_base))

        try:
            for cpe_case in CPE_CASE_OPTIONS:
                get_wind_result(cpe_case)
            design_check_lines, design_envelope_actions = self.member_check_envelope_lines(
                structure,
                self.load_combinations,
                apply_combination_loads,
                top_nodes,
                span,
                foundation_nodes,
            )
        except Exception as exc:
            self.results_box.setText(f"Member check envelope error: {exc}")
            return

        self.foundation_summary = self.foundation_design_from_envelopes(design_envelope_actions.get("foundation_reactions", []))

        try:
            load_arrows, load_summary_lines = apply_combination_loads(selected_combination, include_summary=True)
            d, R, F, analysis_info = self.solve_structure(structure)
        except Exception as exc:
            self.results_box.setText(f"Selected load combination error after envelope checks: {exc}")
            return

        self.structure = structure
        self.displacements = d
        self.reactions = R
        self.force_vector = F
        self.analysis_info = analysis_info
        self.costing_summary = self.costing_from_structure(structure, building_length_m, bay_size_m)
        self.base_nodes = [left_base, right_base] + internal_column_bases + lean_bases + eave_restraint_nodes
        self.support_types = [left_support, right_support] + ["Pinned"] * (len(internal_column_bases) + len(lean_bases)) + eave_restraint_types
        self.load_arrows = load_arrows
        self.wall_cladding_segments = wall_cladding_segments
        self.concept_roof_top_nodes = roof_top_nodes
        self.concept_main_top_nodes = top_nodes
        self.concept_building_length_m = building_length_m
        self.concept_bay_size_m = bay_size_m
        self.concept_wall_states = wind_inputs.wall_enclosure if hasattr(wind_inputs, "wall_enclosure") else self.envelope_plan_selector.wall_states()
        self.single_load_arrows = build_single_load_arrows()
        previous_single_load = getattr(self, "pending_single_load_text", None) or self.single_load_combo.currentText()
        self.pending_single_load_text = None
        self.single_load_combo.blockSignals(True)
        self.single_load_combo.clear()
        for label in self.single_load_arrows:
            self.single_load_combo.addItem(label)
        if previous_single_load:
            self.set_combo_text_if_present(self.single_load_combo, previous_single_load)
        self.single_load_combo.blockSignals(False)
        self.design_envelope_actions = design_envelope_actions
        self.serviceability_height_mm = eave
        self.serviceability_span_mm = span
        self.serviceability_importance_level = wind_inputs.importance_level
        self.serviceability_crane_active = crane_rating_t > 0.0 and left_crane_node is not None and right_crane_node is not None
        self.update_results_text(
            num_panels, panel, web_angle, left_support, right_support,
            left_base, right_base, left_column_profile, right_column_profile, top_profile, bottom_profile, web_profile,
            load_summary_lines, design_check_lines,
            outside_span, span, left_column_offset, right_column_offset,
            left_canopy_length, right_canopy_length, left_canopy_pitch, left_canopy_shallow_depth,
            right_canopy_pitch, right_canopy_shallow_depth, left_canopy_type, right_canopy_type,
            left_canopy_eave_height, right_canopy_eave_height,
            eave_x_restraint,
            active_eave_x_spring_kn_mm,
            eave_bracing_info,
            rafter_haunch_length=rafter_haunch_length if is_rafter else 0.0,
        )
        self.update_section_checks_text(
            left_column_profile,
            right_column_profile,
            internal_column_profile,
            top_profile,
            bottom_profile,
            web_profile,
            left_canopy_top_profile,
            left_canopy_bottom_profile,
            left_canopy_web_profile,
            left_lean_column_profile,
            right_canopy_top_profile,
            right_canopy_bottom_profile,
            right_canopy_web_profile,
            right_lean_column_profile,
        )
        self.update_costing_text()
        self.update_foundation_text()
        self.refresh_plot()

    def member_check_envelope_lines(self, structure, combinations, apply_combination_loads, top_nodes, span, foundation_nodes=None):
        is_rafter = getattr(self, "frame_system_combo", None) is not None and self.frame_system_combo.currentText() == "Rafter"
        foundation_nodes = foundation_nodes or []
        use_second_order = getattr(self, "second_order_checkbox", None) is not None and self.second_order_checkbox.isChecked()

        def empty_action():
            return {"value": -1.0, "combo": "-", "member": "-", "group": "-", "length_mm": 0.0}

        def update_action(envelope, key, value, combination, element, coincident=None):
            if value > envelope[key]["value"]:
                envelope[key] = {
                    "value": value,
                    "combo": combination.name,
                    "member": element.id,
                    "group": element.group,
                    "length_mm": element.length(),
                }
                if coincident:
                    envelope[key].update(coincident)

        def update_average_action(envelope, key, value, combination, group, length_mm):
            if value > envelope[key]["value"]:
                envelope[key] = {
                    "value": value,
                    "combo": combination.name,
                    "member": f"{group} average",
                    "group": group,
                    "length_mm": length_mm,
                }

        def action_line(label, item, unit):
            member_value = str(item["member"])
            member = f"Member {member_value}" if member_value.isdigit() else member_value
            return f"{label:<18}: {item['value']:>8.2f} {unit:<3} | {member:<10} | {item['combo']}"

        def new_member_envelope():
            return {
                "compression_kn": empty_action(),
                "tension_kn": empty_action(),
                "shear_kn": empty_action(),
                "moment_knm": empty_action(),
            }

        column_envelopes = {
            "LEFT_COLUMN": new_member_envelope(),
            "RIGHT_COLUMN": new_member_envelope(),
            "INTERNAL_COLUMN": new_member_envelope(),
            "LEFT_LEAN_COLUMN": new_member_envelope(),
            "RIGHT_LEAN_COLUMN": new_member_envelope(),
            "WEB": new_member_envelope(),
            "LEFT_CANOPY_WEB": new_member_envelope(),
            "RIGHT_CANOPY_WEB": new_member_envelope(),
        }
        column_envelope = {
            "compression_kn": empty_action(),
            "tension_kn": empty_action(),
            "shear_kn": empty_action(),
            "moment_knm": empty_action(),
        }
        chord_envelopes = {
            "TOP": new_member_envelope(),
            "BOTTOM": new_member_envelope(),
            "LEFT_CANOPY_TOP": new_member_envelope(),
            "LEFT_CANOPY_BOTTOM": new_member_envelope(),
            "RIGHT_CANOPY_TOP": new_member_envelope(),
            "RIGHT_CANOPY_BOTTOM": new_member_envelope(),
        }
        sls_column_top_deflection = {"value": -1.0, "combo": "-", "node": "-"}
        sls_column_midspan_ws_deflection = {"value": -1.0, "combo": "-", "member": "-", "group": "-", "height_mm": 0.0}
        def new_truss_midspan_deflections():
            return {
                "G": {"value": -1.0, "combo": "-", "node": "-"},
                "Q": {"value": -1.0, "combo": "-", "node": "-"},
                "Ws": {"value": -1.0, "combo": "-", "node": "-"},
            }

        sls_truss_midspan_deflections = new_truss_midspan_deflections()
        def new_foundation_reactions():
            reactions = []
            for label, node in foundation_nodes:
                reactions.append({
                    "label": label,
                    "node": node,
                    "node_id": node.id,
                    "compression_kn": 0.0,
                    "compression_combo": "-",
                    "compression_limit_state": "SLS",
                    "uplift_kn": 0.0,
                    "uplift_combo": "-",
                    "uplift_limit_state": "ULS",
                    "shear_kn": 0.0,
                    "shear_combo": "-",
                    "moment_knm": 0.0,
                    "moment_combo": "-",
                })
            return reactions

        foundation_reactions = new_foundation_reactions()

        top_left_x = min(node.x for node in top_nodes)
        top_right_x = max(node.x for node in top_nodes)
        midspan_x = (top_left_x + top_right_x) / 2.0
        midspan_node = min(top_nodes, key=lambda node: abs(node.x - midspan_x))
        chord_span = top_right_x - top_left_x
        inner_from = top_left_x + chord_span * 0.125
        inner_to = top_left_x + chord_span * 0.875
        bottom_inner_from = top_left_x + chord_span * 0.30
        bottom_inner_to = top_left_x + chord_span * 0.70
        uls_count = 0
        sls_count = 0

        def is_main_chord_centre_panel(element):
            return (
                not is_rafter
                and
                element.group in {"TOP", "BOTTOM"}
                and (
                    abs(element.start.x - midspan_x) <= 1e-6
                    or abs(element.end.x - midspan_x) <= 1e-6
                )
            )

        def is_canopy_chord_end_panel(element, group_key):
            if not (group_key.endswith("CANOPY_TOP") or group_key.endswith("CANOPY_BOTTOM")):
                return False
            group_elements = [item for item in structure.elements if item.group == group_key]
            if len(group_elements) <= 2:
                return False
            min_x = min(min(item.start.x, item.end.x) for item in group_elements)
            max_x = max(max(item.start.x, item.end.x) for item in group_elements)
            element_min_x = min(element.start.x, element.end.x)
            element_max_x = max(element.start.x, element.end.x)
            return abs(element_min_x - min_x) <= 1e-6 or abs(element_max_x - max_x) <= 1e-6

        def element_global_displacement_at(element, displacements, t):
            t = min(max(t, 0.0), 1.0)
            local = element.local_displacements(displacements)
            length = element.length()
            u = (1.0 - t) * local[0] + t * local[3]
            n1 = 1.0 - 3.0 * t**2 + 2.0 * t**3
            n2 = length * (t - 2.0 * t**2 + t**3)
            n3 = 3.0 * t**2 - 2.0 * t**3
            n4 = length * (-t**2 + t**3)
            v = n1 * local[1] + n2 * local[2] + n3 * local[4] + n4 * local[5]
            theta = element.angle()
            c = math.cos(theta)
            s = math.sin(theta)
            return c * u - s * v, s * u + c * v

        def column_midheight_deflection_items(candidate_structure, displacements):
            items = []
            for group in ["LEFT_COLUMN", "RIGHT_COLUMN", "INTERNAL_COLUMN", "LEFT_LEAN_COLUMN", "RIGHT_LEAN_COLUMN"]:
                group_elements = [element for element in candidate_structure.elements if element.group == group]
                if not group_elements:
                    continue
                nodes = [node for element in group_elements for node in [element.start, element.end]]
                min_y = min(node.y for node in nodes)
                max_y = max(node.y for node in nodes)
                height = max_y - min_y
                if height <= 1e-9:
                    continue
                target_y = min_y + 0.5 * height
                chosen = None
                chosen_t = 0.5
                chosen_distance = float("inf")
                for element in group_elements:
                    y1 = element.start.y
                    y2 = element.end.y
                    if abs(y2 - y1) > 1e-9:
                        t = (target_y - y1) / (y2 - y1)
                    else:
                        t = 0.5
                    clamped_t = min(max(t, 0.0), 1.0)
                    y_at_t = y1 + (y2 - y1) * clamped_t
                    distance = abs(y_at_t - target_y)
                    if distance < chosen_distance:
                        chosen = element
                        chosen_t = clamped_t
                        chosen_distance = distance
                if chosen is None:
                    continue
                ux, _ = element_global_displacement_at(chosen, displacements, chosen_t)
                items.append({
                    "value": abs(ux),
                    "combo": "-",
                    "member": chosen.id,
                    "group": group,
                    "height_mm": height,
                })
            return items

        def solve_candidate_structure(candidate_structure):
            if use_second_order:
                return candidate_structure.solve_second_order()
            d, reactions, force_vector = candidate_structure.solve()
            return d, reactions, force_vector, None

        def merge_action_envelope(target, source):
            for key, source_item in source.items():
                if source_item["value"] > target[key]["value"]:
                    target[key] = source_item

        def merge_foundation_reactions(target, source):
            for target_item, source_item in zip(target, source):
                for key, combo_key in [
                    ("compression_kn", "compression_combo"),
                    ("uplift_kn", "uplift_combo"),
                    ("shear_kn", "shear_combo"),
                    ("moment_knm", "moment_combo"),
                ]:
                    if source_item[key] > target_item[key]:
                        target_item[key] = source_item[key]
                        target_item[combo_key] = source_item[combo_key]

        def merge_combination_result(result):
            nonlocal sls_column_top_deflection, sls_column_midspan_ws_deflection, sls_truss_midspan_deflections, uls_count, sls_count
            uls_count += result["uls_count"]
            sls_count += result["sls_count"]
            for group, envelope in result["column_envelopes"].items():
                merge_action_envelope(column_envelopes[group], envelope)
            merge_action_envelope(column_envelope, result["column_envelope"])
            for group, envelope in result["chord_envelopes"].items():
                merge_action_envelope(chord_envelopes[group], envelope)
            merge_foundation_reactions(foundation_reactions, result["foundation_reactions"])
            if result["sls_column_top_deflection"]["value"] > sls_column_top_deflection["value"]:
                sls_column_top_deflection = result["sls_column_top_deflection"]
            if result["sls_column_midspan_ws_deflection"]["value"] > sls_column_midspan_ws_deflection["value"]:
                sls_column_midspan_ws_deflection = result["sls_column_midspan_ws_deflection"]
            for component, item in result["sls_truss_midspan_deflections"].items():
                if item["value"] > sls_truss_midspan_deflections[component]["value"]:
                    sls_truss_midspan_deflections[component] = item

        def evaluate_combination(combination):
            candidate_structure = copy.deepcopy(structure)
            local_column_envelopes = {
                "LEFT_COLUMN": new_member_envelope(),
                "RIGHT_COLUMN": new_member_envelope(),
                "INTERNAL_COLUMN": new_member_envelope(),
                "LEFT_LEAN_COLUMN": new_member_envelope(),
                "RIGHT_LEAN_COLUMN": new_member_envelope(),
                "WEB": new_member_envelope(),
                "LEFT_CANOPY_WEB": new_member_envelope(),
                "RIGHT_CANOPY_WEB": new_member_envelope(),
            }
            local_column_envelope = new_member_envelope()
            local_chord_envelopes = {
                "TOP": new_member_envelope(),
                "BOTTOM": new_member_envelope(),
                "LEFT_CANOPY_TOP": new_member_envelope(),
                "LEFT_CANOPY_BOTTOM": new_member_envelope(),
                "RIGHT_CANOPY_TOP": new_member_envelope(),
                "RIGHT_CANOPY_BOTTOM": new_member_envelope(),
            }
            local_foundation_reactions = new_foundation_reactions()
            local_sls_column_top_deflection = {"value": -1.0, "combo": "-", "node": "-"}
            local_sls_column_midspan_ws_deflection = {"value": -1.0, "combo": "-", "member": "-", "group": "-", "height_mm": 0.0}
            local_sls_truss_midspan_deflections = new_truss_midspan_deflections()
            local_uls_count = 0
            local_sls_count = 0

            apply_combination_loads(combination, target_structure=candidate_structure)
            d, reactions, _, _ = solve_candidate_structure(candidate_structure)

            for item in local_foundation_reactions:
                node = item["node"]
                ry_kn = reactions[node.uy] / 1000.0
                rx_kn = reactions[node.ux] / 1000.0
                mz_knm = reactions[node.rz] / 1_000_000.0
                compression_kn = max(ry_kn, 0.0)
                uplift_kn = max(-ry_kn, 0.0)
                shear_kn = abs(rx_kn)
                moment_knm = abs(mz_knm)
                if combination.is_serviceability and compression_kn > item["compression_kn"]:
                    item["compression_kn"] = compression_kn
                    item["compression_combo"] = combination.name
                if combination.is_ultimate and uplift_kn > item["uplift_kn"]:
                    item["uplift_kn"] = uplift_kn
                    item["uplift_combo"] = combination.name
                if combination.is_ultimate and shear_kn > item["shear_kn"]:
                    item["shear_kn"] = shear_kn
                    item["shear_combo"] = combination.name
                if combination.is_ultimate and moment_knm > item["moment_knm"]:
                    item["moment_knm"] = moment_knm
                    item["moment_combo"] = combination.name

            if combination.is_ultimate:
                local_uls_count += 1
                chord_case_items = {group: [] for group in local_chord_envelopes}
                for element in candidate_structure.elements:
                    summary = element.force_summary(d)
                    group_key = element.group
                    if element.group in {"LEFT_CANOPY_WEB", "LEFT_CANOPY_POST"}:
                        group_key = "LEFT_CANOPY_WEB"
                    elif element.group in {"RIGHT_CANOPY_WEB", "RIGHT_CANOPY_POST"}:
                        group_key = "RIGHT_CANOPY_WEB"
                    elif element.group in {"WEB", "POST", "APEX"}:
                        group_key = "WEB"
                    moment_knm = summary["moment_knm"]
                    if group_key == "INTERNAL_COLUMN":
                        strong_depth_mm = element.profile.depth or element.profile.flange_width or 0.0
                        eccentric_moment_knm = summary["compression_kn"] * (strong_depth_mm / 2.0) / 1000.0
                        moment_knm = max(moment_knm, eccentric_moment_knm)
                    if group_key in local_column_envelopes:
                        envelope = local_column_envelopes[group_key]
                        update_action(envelope, "compression_kn", summary["compression_kn"], combination, element)
                        update_action(envelope, "tension_kn", summary["tension_kn"], combination, element)
                        update_action(envelope, "shear_kn", summary["shear_kn"], combination, element)
                        update_action(envelope, "moment_knm", moment_knm, combination, element)
                        if group_key in {"LEFT_COLUMN", "RIGHT_COLUMN", "INTERNAL_COLUMN"}:
                            update_action(local_column_envelope, "compression_kn", summary["compression_kn"], combination, element)
                            update_action(local_column_envelope, "tension_kn", summary["tension_kn"], combination, element)
                            update_action(local_column_envelope, "shear_kn", summary["shear_kn"], combination, element)
                            update_action(local_column_envelope, "moment_knm", moment_knm, combination, element)
                    elif group_key in local_chord_envelopes:
                        if is_main_chord_centre_panel(element):
                            continue
                        if is_canopy_chord_end_panel(element, group_key):
                            continue
                        midpoint_x = (element.start.x + element.end.x) / 2.0
                        if is_rafter and group_key == "TOP":
                            include_chord = True
                        elif group_key == "BOTTOM":
                            include_chord = bottom_inner_from <= midpoint_x <= bottom_inner_to
                        else:
                            include_chord = group_key.endswith("CANOPY_TOP") or group_key.endswith("CANOPY_BOTTOM") or inner_from <= midpoint_x <= inner_to
                        if include_chord:
                            chord_case_items[group_key].append((element, summary))

                for group, items in chord_case_items.items():
                    if not items:
                        continue
                    total_length = sum(element.length() for element, _ in items)
                    average_moment = 0.0
                    if total_length > 0.0:
                        average_moment = sum(summary["moment_knm"] * element.length() for element, summary in items) / total_length
                    update_average_action(local_chord_envelopes[group], "moment_knm", average_moment, combination, group, total_length / len(items))
                    for element, summary in items:
                        update_action(
                            local_chord_envelopes[group],
                            "compression_kn",
                            summary["compression_kn"],
                            combination,
                            element,
                            coincident={
                                "coincident_tension_kn": summary["tension_kn"],
                                "coincident_shear_kn": summary["shear_kn"],
                                "coincident_moment_knm": average_moment,
                                "coincident_moment_member": f"{group} average",
                            },
                        )
                        update_action(local_chord_envelopes[group], "tension_kn", summary["tension_kn"], combination, element)
                        update_action(local_chord_envelopes[group], "shear_kn", summary["shear_kn"], combination, element)

            if combination.is_serviceability:
                local_sls_count += 1
                for node in [top_nodes[0], top_nodes[-1]]:
                    ux = abs(d[node.ux])
                    if ux > local_sls_column_top_deflection["value"]:
                        local_sls_column_top_deflection = {
                            "value": ux,
                            "combo": combination.name,
                            "node": node.id,
                        }
                if combination.includes_wind and combination.wind_pressure == "serviceability":
                    for item in column_midheight_deflection_items(candidate_structure, d):
                        if item["value"] > local_sls_column_midspan_ws_deflection["value"]:
                            local_sls_column_midspan_ws_deflection = {
                                **item,
                                "combo": combination.name,
                            }
                truss_component = getattr(combination, "serviceability_component", None)
                if truss_component in local_sls_truss_midspan_deflections:
                    midspan_uy = abs(d[midspan_node.uy])
                    if midspan_uy > local_sls_truss_midspan_deflections[truss_component]["value"]:
                        local_sls_truss_midspan_deflections[truss_component] = {
                            "value": midspan_uy,
                            "combo": combination.name,
                            "node": midspan_node.id,
                        }

            return {
                "column_envelopes": local_column_envelopes,
                "column_envelope": local_column_envelope,
                "chord_envelopes": local_chord_envelopes,
                "sls_column_top_deflection": local_sls_column_top_deflection,
                "sls_column_midspan_ws_deflection": local_sls_column_midspan_ws_deflection,
                "sls_truss_midspan_deflections": local_sls_truss_midspan_deflections,
                "foundation_reactions": local_foundation_reactions,
                "uls_count": local_uls_count,
                "sls_count": local_sls_count,
            }

        max_workers = min(len(combinations), max((os.cpu_count() or 2) - 1, 1), 8)
        if max_workers > 1:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                for result in executor.map(evaluate_combination, combinations):
                    merge_combination_result(result)
        else:
            for combination in combinations:
                merge_combination_result(evaluate_combination(combination))

        lines = ["DESIGN CHECK ENVELOPES"]
        lines.append(f"ULS cases checked : {uls_count}")
        lines.append(f"SLS cases checked : {sls_count}")
        lines.append("")
        lines.append("SERVICEABILITY ENVELOPES")
        lines.append(
            f"{'Column max UX':<18}: {sls_column_top_deflection['value']:>8.3f} mm  | "
            f"Node {sls_column_top_deflection['node']:<7} | {sls_column_top_deflection['combo']}"
        )
        lines.append(
            f"{'Column mid UX Ws':<18}: {sls_column_midspan_ws_deflection['value']:>8.3f} mm  | "
            f"{sls_column_midspan_ws_deflection['group']:<14} | {sls_column_midspan_ws_deflection['combo']}"
        )
        for component in ["G", "Q", "Ws"]:
            item = sls_truss_midspan_deflections[component]
            lines.append(
                f"{'Truss midspan ' + component:<18}: {item['value']:>8.3f} mm  | "
                f"Node {item['node']:<7} | {item['combo']}"
            )
        lines.append("")
        lines.append("STRENGTH ENVELOPES")
        lines.append("")
        lines.append("COLUMNS - ULS ENVELOPE")
        lines.append(action_line("Compression", column_envelope["compression_kn"], "kN"))
        lines.append(action_line("Tension", column_envelope["tension_kn"], "kN"))
        lines.append(action_line("Shear", column_envelope["shear_kn"], "kN"))
        lines.append(action_line("Bending", column_envelope["moment_knm"], "kNm"))
        lines.append("")
        lines.append("RAFTER - ULS ENVELOPE" if is_rafter else "TRUSS CHORDS - ULS ENVELOPE")
        if is_rafter:
            lines.append("Rafter envelope uses the full rafter length")
        else:
            lines.append(f"Ignored chord zone: x < {inner_from:.0f} mm and x > {inner_to:.0f} mm")
            lines.append("Ignored main chord centre panels: members connected to the midspan node")
            lines.append("Bottom chord envelope excludes the outer 30% from each eave")
        lines.append("Canopy/lean-to chord envelopes ignore end panels where an internal panel exists")
        lines.append("Chord bending demand: length-weighted average of chord member moments in the design zone")
        for group, title in [
            ("TOP", "Top chord"),
            ("BOTTOM", "Bottom chord"),
            ("LEFT_CANOPY_TOP", "Left canopy/lean-to top chord"),
            ("LEFT_CANOPY_BOTTOM", "Left canopy/lean-to bottom chord"),
            ("RIGHT_CANOPY_TOP", "Right canopy/lean-to top chord"),
            ("RIGHT_CANOPY_BOTTOM", "Right canopy/lean-to bottom chord"),
        ]:
            if chord_envelopes[group]["compression_kn"]["value"] < 0.0 and chord_envelopes[group]["moment_knm"]["value"] < 0.0:
                continue
            lines.append("")
            lines.append(title)
            lines.append(action_line("Compression", chord_envelopes[group]["compression_kn"], "kN"))
            lines.append(action_line("Tension", chord_envelopes[group]["tension_kn"], "kN"))
            lines.append(action_line("Shear", chord_envelopes[group]["shear_kn"], "kN"))
            lines.append(action_line("Bending", chord_envelopes[group]["moment_knm"], "kNm"))

        if column_envelopes["WEB"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.append("WEBS / POSTS - ULS ENVELOPE")
            lines.append(action_line("Compression", column_envelopes["WEB"]["compression_kn"], "kN"))
            lines.append(action_line("Tension", column_envelopes["WEB"]["tension_kn"], "kN"))
            lines.append(action_line("Shear", column_envelopes["WEB"]["shear_kn"], "kN"))
            lines.append(action_line("Bending", column_envelopes["WEB"]["moment_knm"], "kNm"))
        envelope_actions = {
            "COLUMN": column_envelope,
            "LEFT_COLUMN": column_envelopes["LEFT_COLUMN"],
            "RIGHT_COLUMN": column_envelopes["RIGHT_COLUMN"],
            "INTERNAL_COLUMN": column_envelopes["INTERNAL_COLUMN"],
            "LEFT_LEAN_COLUMN": column_envelopes["LEFT_LEAN_COLUMN"],
            "RIGHT_LEAN_COLUMN": column_envelopes["RIGHT_LEAN_COLUMN"],
            "WEB": column_envelopes["WEB"],
            "LEFT_CANOPY_WEB": column_envelopes["LEFT_CANOPY_WEB"],
            "RIGHT_CANOPY_WEB": column_envelopes["RIGHT_CANOPY_WEB"],
            "TOP": chord_envelopes["TOP"],
            "BOTTOM": chord_envelopes["BOTTOM"],
            "LEFT_CANOPY_TOP": chord_envelopes["LEFT_CANOPY_TOP"],
            "LEFT_CANOPY_BOTTOM": chord_envelopes["LEFT_CANOPY_BOTTOM"],
            "RIGHT_CANOPY_TOP": chord_envelopes["RIGHT_CANOPY_TOP"],
            "RIGHT_CANOPY_BOTTOM": chord_envelopes["RIGHT_CANOPY_BOTTOM"],
            "sls_column_top_deflection": sls_column_top_deflection,
            "sls_column_midspan_ws_deflection": sls_column_midspan_ws_deflection,
            "sls_truss_midspan_deflections": sls_truss_midspan_deflections,
            "uls_count": uls_count,
            "sls_count": sls_count,
            "foundation_reactions": [
                {key: value for key, value in item.items() if key != "node"}
                for item in foundation_reactions
            ],
        }
        return lines, envelope_actions

    def member_envelope_has_demands(self, envelope):
        if not envelope:
            return False
        return any(
            isinstance(item, dict) and item.get("value", -1.0) >= 0.0
            for item in envelope.values()
        )

    def member_envelope_value(self, envelope, key):
        return max(envelope[key]["value"], 0.0)

    def member_envelope_effective_length(self, envelope, key, fallback=0.0):
        return envelope[key].get("length_mm") or fallback

    def member_actions_from_envelope(self, envelope, group, restraint_override=None):
        restraint = dict((self.member_check_restraints or {}).get(group, {}))
        if restraint_override:
            restraint.update(restraint_override)
        major_length = restraint.get("major_mm") or self.member_envelope_effective_length(envelope, "compression_kn")
        minor_length = restraint.get("minor_mm") or major_length
        if major_length <= 0.0:
            lengths = [item.get("length_mm", 0.0) for item in envelope.values() if isinstance(item, dict)]
            major_length = max(lengths or [0.0])
        if minor_length <= 0.0:
            minor_length = major_length
        ke = restraint.get("ke", 1.0)
        moment_knm = self.member_envelope_value(envelope, "moment_knm")
        if group in {"TOP", "BOTTOM", "LEFT_CANOPY_TOP", "LEFT_CANOPY_BOTTOM", "RIGHT_CANOPY_TOP", "RIGHT_CANOPY_BOTTOM"}:
            moment_knm = max(envelope["compression_kn"].get("coincident_moment_knm", 0.0), 0.0)
        return MemberActions(
            compression_kn=self.member_envelope_value(envelope, "compression_kn"),
            tension_kn=self.member_envelope_value(envelope, "tension_kn"),
            shear_kn=self.member_envelope_value(envelope, "shear_kn"),
            moment_knm=moment_knm,
            effective_length_y_mm=(ke * minor_length) or None,
            effective_length_z_mm=(ke * major_length) or None,
            moment_modifier=restraint.get("am", 1.0),
            kt=restraint.get("kt", 1.0),
            kl=restraint.get("kl", 1.0),
            kr=restraint.get("kr", 1.0),
        )

    def member_check_result(self, profile, envelope, group, restraint_override=None):
        return check_member(profile.check_properties(), self.member_actions_from_envelope(envelope, group, restraint_override))

    def auto_size_profile_metric(self, profile, group):
        inertia_values = [value for value in [profile.Iyp, profile.Izp] if value > 0.0]
        if not inertia_values:
            return profile.A
        if group in {
            "TOP", "BOTTOM", "WEB",
            "LEFT_CANOPY_TOP", "LEFT_CANOPY_BOTTOM", "LEFT_CANOPY_WEB",
            "RIGHT_CANOPY_TOP", "RIGHT_CANOPY_BOTTOM", "RIGHT_CANOPY_WEB",
        }:
            return min(inertia_values)
        return max(inertia_values)

    def auto_size_profile_cost_key(self, profile):
        bucket = self.section_cost_bucket(profile)
        default_key = "shs_rate_per_kg" if bucket == "shs" else "ub_rate_per_kg"
        rate = self.costing_rate_from_input(
            self.shs_cost_input if bucket == "shs" else self.ub_cost_input,
            default_key,
        )
        kg_per_m = profile.A * 1e-6 * 7850.0
        return kg_per_m * rate

    def auto_size_candidate_indices(self, combo, group, min_metric_step=0.18):
        profiles = []
        for index in range(combo.count()):
            profile = self.catalog.get(combo.itemText(index))
            if profile is None:
                continue
            profiles.append((index, profile, self.auto_size_profile_metric(profile, group)))
        profiles = sorted(profiles, key=lambda item: (item[2], self.auto_size_profile_cost_key(item[1]), item[1].A, item[1].name))
        if not profiles:
            return []

        current_index = combo.currentIndex()
        ladder = []
        last_metric = None
        last_depth = None
        for index, profile, metric in profiles:
            depth = profile.depth or self.profile_name_depth(profile.name)
            keep = index == current_index or not ladder
            if last_metric and metric >= last_metric * (1.0 + min_metric_step):
                keep = True
            if last_depth and depth > last_depth:
                keep = True
            if keep:
                ladder.append(index)
                last_metric = metric
                last_depth = depth

        if current_index not in ladder:
            ladder.append(current_index)
        return sorted(set(ladder), key=lambda index: next(
            (self.auto_size_profile_metric(profile, group), self.auto_size_profile_cost_key(profile), profile.A, profile.name)
            for item_index, profile, _ in profiles if item_index == index
        ))

    def profile_name_depth(self, name):
        numbers = [float(value) for value in re.findall(r"\d+(?:\.\d+)?", str(name))]
        return numbers[0] if numbers else 0.0

    def auto_size_candidate_walk(self, combo, group, current_passes):
        ladder = self.auto_size_candidate_indices(combo, group)
        if not ladder:
            return []
        current_index = combo.currentIndex()
        if current_index not in ladder:
            ladder.append(current_index)
            ladder = sorted(
                set(ladder),
                key=lambda index: (
                    self.auto_size_profile_metric(self.catalog.get(combo.itemText(index)), group),
                    self.auto_size_profile_cost_key(self.catalog.get(combo.itemText(index))),
                    self.catalog.get(combo.itemText(index)).A,
                    self.catalog.get(combo.itemText(index)).name,
                ),
            )
        position = ladder.index(current_index)
        if current_passes:
            return list(reversed(ladder[:position]))
        return ladder[position + 1:]

    def auto_size_overall_ratio(self, active_specs):
        ratios = []
        envelopes = self.design_envelope_actions or {}
        for _, combo, group in active_specs:
            envelope = envelopes.get(group)
            if not self.member_envelope_has_demands(envelope):
                continue
            result = self.member_check_result(self.selected_profile(combo), envelope, group)
            ratios.append((result.governing_ratio, group, result.governing_check))

        height = self.serviceability_height_mm or 0.0
        if height > 0.0 and "sls_column_top_deflection" in envelopes:
            divisor = 250.0 if getattr(self, "serviceability_crane_active", False) else (80.0 if (self.serviceability_importance_level or 1) == 1 else 150.0)
            limit = height / divisor
            demand = max(envelopes["sls_column_top_deflection"]["value"], 0.0)
            ratios.append((demand / limit if limit > 1e-12 else float("inf"), "Column deflection", f"H/{divisor:.0f}"))
        if "sls_column_midspan_ws_deflection" in envelopes:
            item = envelopes["sls_column_midspan_ws_deflection"]
            limit = (item.get("height_mm") or height) / 150.0 if (item.get("height_mm") or height) > 0.0 else 0.0
            demand = max(item["value"], 0.0)
            ratios.append((demand / limit if limit > 1e-12 else float("inf"), "Column midspan Ws deflection", "H/150"))

        span = self.serviceability_span_mm or 0.0
        if span > 0.0 and "sls_truss_midspan_deflections" in envelopes:
            for component, divisor in self.midspan_component_deflection_divisors().items():
                item = envelopes["sls_truss_midspan_deflections"].get(component, {})
                limit = span / divisor
                demand = max(item.get("value", -1.0), 0.0)
                ratios.append((demand / limit if limit > 1e-12 else float("inf"), f"Midspan {component} deflection", f"Span/{divisor:.0f}"))

        return max(ratios, default=(0.0, "-", "-"), key=lambda item: item[0])

    def auto_size_adjust_truss_depth(self, active_specs):
        if getattr(self, "frame_system_combo", None) is not None and self.frame_system_combo.currentText() == "Rafter":
            return None
        try:
            current_depth = float(self.depth_input.text())
            min_depth = self.truss_depth_from_span(0.03)
            max_depth = self.truss_depth_from_span(0.045)
        except ValueError:
            return None

        current_depth = min(max(self.round_up_to_increment(current_depth, 50.0), min_depth), max_depth)
        self.depth_input.setText(f"{current_depth:.0f}")
        self.generate_and_solve()
        QApplication.processEvents()

        start_depth = current_depth
        best_depth = current_depth
        ratio, _, _ = self.auto_size_overall_ratio(active_specs)
        if ratio <= 1.0:
            trial_depth = current_depth - 50.0
            while trial_depth >= min_depth:
                self.depth_input.setText(f"{trial_depth:.0f}")
                self.generate_and_solve()
                QApplication.processEvents()
                trial_ratio, _, _ = self.auto_size_overall_ratio(active_specs)
                if trial_ratio > 1.0:
                    break
                best_depth = trial_depth
                trial_depth -= 50.0
        else:
            trial_depth = current_depth + 50.0
            while trial_depth <= max_depth:
                self.depth_input.setText(f"{trial_depth:.0f}")
                self.generate_and_solve()
                QApplication.processEvents()
                trial_ratio, _, _ = self.auto_size_overall_ratio(active_specs)
                best_depth = trial_depth
                if trial_ratio <= 1.0:
                    break
                trial_depth += 50.0

        self.depth_input.setText(f"{best_depth:.0f}")
        self.generate_and_solve()
        QApplication.processEvents()
        final_ratio, final_group, final_check = self.auto_size_overall_ratio(active_specs)
        return {
            "start": start_depth,
            "end": best_depth,
            "min": min_depth,
            "max": max_depth,
            "ratio": final_ratio,
            "governing_group": final_group,
            "governing_check": final_check,
        }

    def auto_size_group_specs(self):
        return [
            ("Left column", self.left_column_profile_combo, "LEFT_COLUMN"),
            ("Right column", self.right_column_profile_combo, "RIGHT_COLUMN"),
            ("Internal column", self.internal_column_profile_combo, "INTERNAL_COLUMN"),
            ("Top chord", self.top_profile_combo, "TOP"),
            ("Bottom chord", self.bottom_profile_combo, "BOTTOM"),
            ("Web/post", self.web_profile_combo, "WEB"),
            ("Left canopy/lean-to top", self.left_canopy_top_profile_combo, "LEFT_CANOPY_TOP"),
            ("Left canopy/lean-to bottom", self.left_canopy_bottom_profile_combo, "LEFT_CANOPY_BOTTOM"),
            ("Left canopy/lean-to web/post", self.left_canopy_web_profile_combo, "LEFT_CANOPY_WEB"),
            ("Left lean-to outer column", self.left_lean_column_profile_combo, "LEFT_LEAN_COLUMN"),
            ("Right canopy/lean-to top", self.right_canopy_top_profile_combo, "RIGHT_CANOPY_TOP"),
            ("Right canopy/lean-to bottom", self.right_canopy_bottom_profile_combo, "RIGHT_CANOPY_BOTTOM"),
            ("Right canopy/lean-to web/post", self.right_canopy_web_profile_combo, "RIGHT_CANOPY_WEB"),
            ("Right lean-to outer column", self.right_lean_column_profile_combo, "RIGHT_LEAN_COLUMN"),
        ]

    def active_auto_size_group_specs(self):
        envelopes = self.design_envelope_actions or {}
        return [
            spec for spec in self.auto_size_group_specs()
            if self.member_envelope_has_demands(envelopes.get(spec[2]))
        ]

    def auto_size_current_preset(self):
        self.auto_size_report_text = ""
        self.generate_and_solve()
        if not self.design_envelope_actions:
            QMessageBox.warning(self, "Auto-size unavailable", "Generate and solve the frame before auto-sizing.")
            return

        active_specs = self.active_auto_size_group_specs()
        if not active_specs:
            QMessageBox.information(self, "Auto-size current preset", "No active member groups were found to auto-size.")
            return

        starting_depth = self.depth_input.text()
        starting_sections = {group: combo.currentText() for _, combo, group in active_specs}
        selected_sections = {}
        unresolved = []
        depth_results = []

        depth_result = self.auto_size_adjust_truss_depth(active_specs)
        if depth_result:
            depth_results.append(depth_result)

        for label, combo, group in active_specs:
            starting_index = combo.currentIndex()
            chosen_index = starting_index
            last_result = None
            self.generate_and_solve()
            QApplication.processEvents()
            envelope = (self.design_envelope_actions or {}).get(group)
            if not self.member_envelope_has_demands(envelope):
                continue
            profile = self.selected_profile(combo)
            last_result = self.member_check_result(profile, envelope, group)
            current_passes = last_result.governing_ratio <= 1.0

            for index in self.auto_size_candidate_walk(combo, group, current_passes):
                combo.setCurrentIndex(index)
                self.generate_and_solve()
                QApplication.processEvents()
                if not current_passes:
                    chosen_index = index
                envelope = (self.design_envelope_actions or {}).get(group)
                if not self.member_envelope_has_demands(envelope):
                    continue
                profile = self.selected_profile(combo)
                last_result = self.member_check_result(profile, envelope, group)
                if last_result.governing_ratio <= 1.0:
                    chosen_index = index
                    if not current_passes:
                        break
                    continue
                if current_passes:
                    break

            combo.setCurrentIndex(chosen_index)
            self.generate_and_solve()
            QApplication.processEvents()
            envelope = (self.design_envelope_actions or {}).get(group)
            if self.member_envelope_has_demands(envelope):
                profile = self.selected_profile(combo)
                last_result = self.member_check_result(profile, envelope, group)
            if last_result is None or last_result.governing_ratio > 1.0:
                unresolved.append((label, group, last_result))
            selected_sections[group] = combo.currentText()

        depth_result = self.auto_size_adjust_truss_depth(active_specs)
        if depth_result:
            depth_results.append(depth_result)

        self.generate_and_solve()
        report_lines = []
        final_depth = self.depth_input.text()
        depth_change_text = f"{starting_depth} -> {final_depth}" if starting_depth != final_depth else f"{final_depth} (unchanged)"
        if depth_results:
            last_depth_result = depth_results[-1]
            report_lines.append(
                f"{'Truss depth':<31}: {depth_change_text} | bounds "
                f"{last_depth_result['min']:.0f}-{last_depth_result['max']:.0f} mm | "
                f"governs {last_depth_result['governing_group']} {last_depth_result['governing_check']} "
                f"{last_depth_result['ratio']:.3f}"
            )
        for label, combo, group in active_specs:
            envelope = (self.design_envelope_actions or {}).get(group)
            if not self.member_envelope_has_demands(envelope):
                continue
            profile = self.selected_profile(combo)
            result = self.member_check_result(profile, envelope, group)
            status = "OK" if result.governing_ratio <= 1.0 else "NOT OK"
            before = starting_sections.get(group, "-")
            after = selected_sections.get(group, combo.currentText())
            change_text = f"{before} -> {after}" if before != after else f"{after} (unchanged)"
            report_lines.append(
                f"{label:<31}: {change_text} | governs {result.governing_check} "
                f"{result.governing_ratio:.3f} | {status}"
            )

        if unresolved:
            unresolved_labels = ", ".join(label for label, _, _ in unresolved)
            report_lines.append("")
            report_lines.append(f"No passing catalog option found for: {unresolved_labels}")

        self.auto_size_report_text = "\n".join(report_lines)
        self.generate_and_solve()
        QMessageBox.information(
            self,
            "Auto-size current preset",
            self.auto_size_report_text or "Auto-size completed.",
        )

    def update_section_checks_text(
        self,
        left_column_profile,
        right_column_profile,
        internal_column_profile,
        top_profile,
        bottom_profile,
        web_profile,
        left_canopy_top_profile=None,
        left_canopy_bottom_profile=None,
        left_canopy_web_profile=None,
        left_lean_column_profile=None,
        right_canopy_top_profile=None,
        right_canopy_bottom_profile=None,
        right_canopy_web_profile=None,
        right_lean_column_profile=None,
    ):
        if not self.design_envelope_actions:
            self.section_checks_box.setText("Generate and solve the frame to view section checks.")
            return

        def resolved_section_check(profile, envelope, group):
            base_restraint = dict((self.member_check_restraints or {}).get(group, {}))
            base_actions = self.member_actions_from_envelope(envelope, group)
            base_result = self.member_check_result(profile, envelope, group)
            if group == "BOTTOM" and base_result.governing_ratio > 1.0:
                top_minor = ((self.member_check_restraints or {}).get("TOP", {}) or {}).get("minor_mm")
                bottom_minor = base_restraint.get("minor_mm")
                if top_minor and bottom_minor and top_minor < bottom_minor:
                    retry_restraint = {**base_restraint, "minor_mm": top_minor}
                    retry_result = self.member_check_result(profile, envelope, group, retry_restraint)
                    if retry_result.governing_ratio <= 1.0:
                        retry_actions = self.member_actions_from_envelope(envelope, group, retry_restraint)
                        return retry_result, retry_actions, retry_restraint, [
                            "Additional flybracing required: default every-second-purlin restraint failed; provide flybracing at every purlin over the checked bottom-chord region."
                        ]
            return base_result, base_actions, base_restraint, []

        def format_check_set(title, profile, envelope, group):
            result, actions, restraint, extra_notes = resolved_section_check(profile, envelope, group)
            status = "OK" if result.governing_ratio <= 1.0 else "NOT OK"
            clad_text = ""
            if restraint.get("clad") is not None:
                clad_text = f" | clad: {'yes' if restraint.get('clad') else 'no'}"
            lines = [
                title,
                f"Section: {profile.name}",
                f"Effective lengths: major {actions.effective_length_z_mm or 0.0:.0f} mm, minor {actions.effective_length_y_mm or 0.0:.0f} mm | ke {restraint.get('ke', 1.0):.2f} | am {actions.moment_modifier:.2f} | kt/kl/kr {actions.kt:.2f}/{actions.kl:.2f}/{actions.kr:.2f}{clad_text}",
                f"Governing: {result.governing_check} | ratio {result.governing_ratio:.3f} | {status}",
                "Check        | Demand      | Capacity    | Ratio  | Status | Source",
            ]
            if group == "INTERNAL_COLUMN":
                lines.append("Internal column bending includes AS4100 eccentricity moment Mx = Nc x D/2.")
            source_keys = {
                "compression": "compression_kn",
                "tension": "tension_kn",
                "shear": "shear_kn",
                "bending": "moment_knm",
            }
            for key in ["compression", "tension", "shear", "bending", "combined"]:
                check = result.checks[key]
                unit = f" {check.unit}" if check.unit else ""
                check_status = "OK" if check.ratio <= 1.0 else "NOT OK"
                if key == "bending" and group in {"TOP", "BOTTOM", "LEFT_CANOPY_TOP", "LEFT_CANOPY_BOTTOM", "RIGHT_CANOPY_TOP", "RIGHT_CANOPY_BOTTOM"}:
                    source = envelope["compression_kn"]
                    source_member = source.get("coincident_moment_member", f"{group} average")
                    source_text = f"{source_member} | {source['combo']} | compression case"
                elif key in source_keys:
                    source = envelope[source_keys[key]]
                    source_text = f"Member {source['member']} | {source['combo']}"
                else:
                    source_text = "Envelope interaction"
                lines.append(
                    f"{check.label:<12} | {check.demand:>8.2f}{unit:<4} | "
                    f"{check.capacity:>8.2f}{unit:<4} | {check.ratio:>6.3f} | {check_status:<6} | {source_text}"
                )
            notes = extra_notes + [check.notes for check in result.checks.values() if check.notes]
            if notes:
                lines.append("Notes:")
                for note in dict.fromkeys(notes):
                    lines.append(f"- {note}")
            return lines

        def section_result(profile, envelope, group):
            result, _, _, notes = resolved_section_check(profile, envelope, group)
            return result, notes

        def update_status_label(label, result_info):
            result, notes = result_info
            status = "OK" if result.governing_ratio <= 1.0 else "NOT OK"
            if notes and result.governing_ratio <= 1.0:
                color = "#a87900"
            elif result.governing_ratio >= 1.0:
                color = "#b42318"
            elif result.governing_ratio >= 0.95:
                color = "#c45a00"
            elif result.governing_ratio >= 0.90:
                color = "#a87900"
            else:
                color = "#1b7f3a"
            suffix = " - additional flybracing required" if notes else ""
            label.setText(f"{status}{suffix} - max utilisation {result.governing_ratio:.3f} ({result.governing_check})")
            label.setStyleSheet(f"font-weight: bold; color: {color};")

        def column_deflection_check():
            height = self.serviceability_height_mm or 0.0
            importance = self.serviceability_importance_level or 1
            if getattr(self, "serviceability_crane_active", False):
                divisor = 250.0
            else:
                divisor = 80.0 if importance == 1 else 150.0
            limit = height / divisor if height > 0.0 else 0.0
            demand = max(self.design_envelope_actions["sls_column_top_deflection"]["value"], 0.0)
            ratio = demand / limit if limit > 1e-12 else float("inf")
            status = "OK" if ratio <= 1.0 else "NOT OK"
            return demand, limit, ratio, status, divisor

        def column_midspan_ws_deflection_check():
            item = self.design_envelope_actions.get("sls_column_midspan_ws_deflection", {})
            height = item.get("height_mm") or self.serviceability_height_mm or 0.0
            divisor = 150.0
            limit = height / divisor if height > 0.0 else 0.0
            demand = max(item.get("value", -1.0), 0.0)
            ratio = demand / limit if limit > 1e-12 else float("inf")
            status = "OK" if ratio <= 1.0 else "NOT OK"
            return demand, limit, ratio, status, divisor, item

        def truss_deflection_checks():
            span = self.serviceability_span_mm or 0.0
            checks = {}
            items = self.design_envelope_actions.get("sls_truss_midspan_deflections", {})
            for component, divisor in self.midspan_component_deflection_divisors().items():
                item = items.get(component, {})
                limit = span / divisor if span > 0.0 else 0.0
                demand = max(item.get("value", -1.0), 0.0)
                ratio = demand / limit if limit > 1e-12 else float("inf")
                status = "OK" if ratio <= 1.0 else "NOT OK"
                checks[component] = {
                    "demand": demand,
                    "limit": limit,
                    "ratio": ratio,
                    "status": status,
                    "divisor": divisor,
                    "item": item,
                }
            return checks

        envelopes = self.design_envelope_actions
        is_rafter = getattr(self, "frame_system_combo", None) is not None and self.frame_system_combo.currentText() == "Rafter"
        update_status_label(self.left_column_check_label, section_result(left_column_profile, envelopes["LEFT_COLUMN"], "LEFT_COLUMN"))
        update_status_label(self.right_column_check_label, section_result(right_column_profile, envelopes["RIGHT_COLUMN"], "RIGHT_COLUMN"))
        if envelopes["INTERNAL_COLUMN"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.internal_column_check_label, section_result(internal_column_profile, envelopes["INTERNAL_COLUMN"], "INTERNAL_COLUMN"))
        else:
            self.internal_column_check_label.setText("Not checked")
            self.internal_column_check_label.setStyleSheet("font-weight: bold;")
        update_status_label(self.top_chord_check_label, section_result(top_profile, envelopes["TOP"], "TOP"))
        if is_rafter:
            self.bottom_chord_check_label.setText("Not checked")
            self.bottom_chord_check_label.setStyleSheet("font-weight: bold;")
            self.web_check_label.setText("Not checked")
            self.web_check_label.setStyleSheet("font-weight: bold;")
        else:
            update_status_label(self.bottom_chord_check_label, section_result(bottom_profile, envelopes["BOTTOM"], "BOTTOM"))
        purlin_check = self.purlin_check or {}
        if purlin_check:
            match = purlin_check.get("match", {})
            local_match = purlin_check.get("local_match", {})
            purlin_status = match.get("status", "NOT CHECKED")
            local_status = local_match.get("status", "NOT CHECKED")
            combined_status = "OK" if purlin_status == "OK" and local_status == "OK" else "NOT OK"
            purlin_color = "#1b7f3a" if combined_status == "OK" else "#b42318"
            if combined_status == "OK":
                self.purlin_check_label.setText(
                    f"OK - general {match.get('section', '-')}, edge {local_match.get('section', '-')} first {purlin_check.get('edge_rows', 0)} rows"
                )
            else:
                self.purlin_check_label.setText(
                    f"NOT OK - general {match.get('section', match.get('message', '-'))}, edge {local_match.get('section', local_match.get('message', '-'))}"
                )
            self.purlin_check_label.setStyleSheet(f"font-weight: bold; color: {purlin_color};")
        else:
            self.purlin_check_label.setText("Not checked")
            self.purlin_check_label.setStyleSheet("font-weight: bold;")
        wall_girt_check = self.wall_girt_check or {}
        if wall_girt_check:
            match = wall_girt_check.get("match", {})
            girt_status = match.get("status", "NOT CHECKED")
            girt_color = "#1b7f3a" if girt_status == "OK" else "#b42318"
            if girt_status == "OK":
                self.wall_girt_check_label.setText(f"OK - {match.get('section', '-')} ({match.get('tables', '-')})")
            else:
                self.wall_girt_check_label.setText(f"{girt_status} - {match.get('message', 'No match')}")
            self.wall_girt_check_label.setStyleSheet(f"font-weight: bold; color: {girt_color};")
        else:
            self.wall_girt_check_label.setText("Not checked")
            self.wall_girt_check_label.setStyleSheet("font-weight: bold;")
        if not is_rafter and envelopes["WEB"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.web_check_label, section_result(web_profile, envelopes["WEB"], "WEB"))
        elif not is_rafter:
            self.web_check_label.setText("Not checked")
            self.web_check_label.setStyleSheet("font-weight: bold;")
        if left_canopy_top_profile is not None and envelopes["LEFT_CANOPY_TOP"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.left_canopy_top_check_label, section_result(left_canopy_top_profile, envelopes["LEFT_CANOPY_TOP"], "LEFT_CANOPY_TOP"))
        if left_canopy_bottom_profile is not None and envelopes["LEFT_CANOPY_BOTTOM"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.left_canopy_bottom_check_label, section_result(left_canopy_bottom_profile, envelopes["LEFT_CANOPY_BOTTOM"], "LEFT_CANOPY_BOTTOM"))
        if left_canopy_web_profile is not None and envelopes["LEFT_CANOPY_WEB"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.left_canopy_web_check_label, section_result(left_canopy_web_profile, envelopes["LEFT_CANOPY_WEB"], "LEFT_CANOPY_WEB"))
        if left_lean_column_profile is not None and envelopes["LEFT_LEAN_COLUMN"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.left_lean_column_check_label, section_result(left_lean_column_profile, envelopes["LEFT_LEAN_COLUMN"], "LEFT_LEAN_COLUMN"))
        if right_canopy_top_profile is not None and envelopes["RIGHT_CANOPY_TOP"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.right_canopy_top_check_label, section_result(right_canopy_top_profile, envelopes["RIGHT_CANOPY_TOP"], "RIGHT_CANOPY_TOP"))
        if right_canopy_bottom_profile is not None and envelopes["RIGHT_CANOPY_BOTTOM"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.right_canopy_bottom_check_label, section_result(right_canopy_bottom_profile, envelopes["RIGHT_CANOPY_BOTTOM"], "RIGHT_CANOPY_BOTTOM"))
        if right_canopy_web_profile is not None and envelopes["RIGHT_CANOPY_WEB"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.right_canopy_web_check_label, section_result(right_canopy_web_profile, envelopes["RIGHT_CANOPY_WEB"], "RIGHT_CANOPY_WEB"))
        if right_lean_column_profile is not None and envelopes["RIGHT_LEAN_COLUMN"]["compression_kn"]["value"] >= 0.0:
            update_status_label(self.right_lean_column_check_label, section_result(right_lean_column_profile, envelopes["RIGHT_LEAN_COLUMN"], "RIGHT_LEAN_COLUMN"))
        deflection_demand, deflection_limit, deflection_ratio, deflection_status, deflection_divisor = column_deflection_check()
        mid_column_demand, mid_column_limit, mid_column_ratio, mid_column_status, mid_column_divisor, mid_column_item = column_midspan_ws_deflection_check()
        self.column_deflection_check_label.setText(
            f"Top {deflection_status} {deflection_ratio:.3f} (H/{deflection_divisor:.0f}); mid Ws {mid_column_status} {mid_column_ratio:.3f} (H/{mid_column_divisor:.0f})"
        )
        self.column_deflection_check_label.setStyleSheet(f"font-weight: bold; color: {'#b42318' if 'NOT OK' in [deflection_status, mid_column_status] else '#1b7f3a'};")
        truss_deflection_checks_by_component = truss_deflection_checks()
        governing_truss_deflection = max(
            truss_deflection_checks_by_component.items(),
            default=("-", {"ratio": 0.0, "status": "OK", "demand": 0.0, "limit": 0.0, "divisor": 1.0}),
            key=lambda entry: entry[1]["ratio"],
        )
        truss_deflection_status = "NOT OK" if any(item["status"] == "NOT OK" for item in truss_deflection_checks_by_component.values()) else "OK"
        truss_deflection_color = "#1b7f3a" if truss_deflection_status == "OK" else "#b42318"
        governing_component, governing_item = governing_truss_deflection
        self.truss_deflection_check_label.setText(
            f"{truss_deflection_status} - gov {governing_component} {governing_item['demand']:.2f} / {governing_item['limit']:.2f} mm = {governing_item['ratio']:.3f} (Span/{governing_item['divisor']:.0f})"
        )
        self.truss_deflection_check_label.setStyleSheet(f"font-weight: bold; color: {truss_deflection_color};")

        def truss_serviceability_line(component):
            item = truss_deflection_checks_by_component[component]
            combo_item = item.get("item", {})
            label = "Midspan wind deflection Ws" if component == "Ws" else f"Midspan sag {component}"
            return (
                f"{label}: {item['demand']:.3f} mm <= {item['limit']:.3f} mm "
                f"(Span/{item['divisor']:.0f}) | {combo_item.get('combo', '-')} | "
                f"utilisation {item['ratio']:.3f} | {item['status']}"
            )

        lines = [
            "SECTION CHECKS",
            "Uses ULS envelope demands from the Results tab.",
            "Column restraints: clad walls use 3000 mm major / 1500 mm minor; unclad walls use base to bottom of truss/rafter, with rafter haunches deducting rafter depth plus haunch depth.",
            f"{'Rafter' if is_rafter else 'Chord'} restraints: major length uses node spacing; top/rafter minor length uses purlin spacing {((self.member_check_restraints or {}).get('TOP', {}).get('minor_mm') or 0.0):.0f} mm; bottom minor length defaults to every second girt/flybrace {((self.member_check_restraints or {}).get('BOTTOM', {}).get('minor_mm') or 0.0):.0f} mm, with automatic every-purlin retry if required.",
            "Checks include axial, Clause 5.11-style shear, member bending, and axial plus bending interaction.",
            "Compactness/effective section modulus refinements are still to be verified against worked examples.",
            "",
            "SERVICEABILITY CHECKS",
            f"Column top deflection SLS: {deflection_demand:.3f} mm <= {deflection_limit:.3f} mm (H/{deflection_divisor:.0f}) | utilisation {deflection_ratio:.3f} | {deflection_status}",
            f"Column midspan deflection Ws: {mid_column_demand:.3f} mm <= {mid_column_limit:.3f} mm (H/{mid_column_divisor:.0f}) | {mid_column_item.get('group', '-')} | {mid_column_item.get('combo', '-')} | utilisation {mid_column_ratio:.3f} | {mid_column_status}",
            truss_serviceability_line("G"),
            truss_serviceability_line("Q"),
            truss_serviceability_line("Ws"),
            "",
            "STRENGTH CHECKS",
        ]
        auto_size_report = getattr(self, "auto_size_report_text", "")
        if auto_size_report:
            lines.extend(["AUTO-SIZE REPORT", auto_size_report, ""])
        lines.extend(format_check_set("LEFT COLUMN", left_column_profile, envelopes["LEFT_COLUMN"], "LEFT_COLUMN"))
        lines.append("")
        lines.extend(format_check_set("RIGHT COLUMN", right_column_profile, envelopes["RIGHT_COLUMN"], "RIGHT_COLUMN"))
        if envelopes["INTERNAL_COLUMN"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("INTERNAL COLUMN", internal_column_profile, envelopes["INTERNAL_COLUMN"], "INTERNAL_COLUMN"))
        lines.append("")
        lines.extend(format_check_set("RAFTER" if is_rafter else "TOP CHORD - CENTRAL 75% ENVELOPE", top_profile, envelopes["TOP"], "TOP"))
        if not is_rafter:
            lines.append("")
            lines.extend(format_check_set("BOTTOM CHORD - CENTRAL 40% ENVELOPE", bottom_profile, envelopes["BOTTOM"], "BOTTOM"))
        if purlin_check:
            match = purlin_check.get("match", {})
            local_match = purlin_check.get("local_match", {})
            lines.append("")
            lines.extend([
                "PURLIN CHECK",
                f"Span type: {purlin_check['span_type']} | table span {match.get('span_row_mm', 0):.0f} mm{' interpolated' if match.get('interpolated') else ''} | spacing {purlin_check['spacing_mm']:.0f} mm",
                f"Line loads: outward ULS {purlin_check['outward_kn_m']:.3f} kN/m, inward ULS {purlin_check['inward_strength_kn_m']:.3f} kN/m, inward SLS {purlin_check['inward_service_kn_m']:.3f} kN/m",
                f"Area loads: G {purlin_check['permanent_kpa']:.3f} kPa, Q {purlin_check['live_kpa']:.3f} kPa, wind up/down {purlin_check['wind_upward_kpa']:.3f}/{purlin_check['wind_downward_kpa']:.3f} kPa",
                f"Inward ULS basis: max(1.2G+1.5Q = {purlin_check['inward_gravity_uls_kpa']:.3f} kPa, 1.2G+Wd = {purlin_check['inward_wind_uls_kpa']:.3f} kPa)",
            ])
            if match.get("status") == "OK":
                lines.extend([
                    f"General matched section: {match['section']} from Metroll tables {match['tables']}",
                    f"Selection basis: deepest section within 5% of the lightest passing area; selected area {match.get('area_mm2', 0.0):.0f} mm2.",
                    f"Capacities: outward {match['outward_capacity_kn_m']:.3f} kN/m, inward strength {match['inward_strength_capacity_kn_m']:.3f} kN/m, inward L/150 {match['inward_service_capacity_kn_m']:.3f} kN/m",
                ])
                if match.get("capped_by_double_span"):
                    lines.append("Single-span capacities capped at corresponding double-span values where extracted table values were higher.")
                if match.get("near_misses"):
                    lines.append("Nearest failing sections:")
                    for miss in match["near_misses"]:
                        lines.append(
                            f"- {miss['section']}: outward {miss['outward_ratio']:.3f}, "
                            f"inward strength {miss['inward_strength_ratio']:.3f}, "
                            f"inward L/150 {miss['inward_service_ratio']:.3f}"
                        )
            else:
                lines.append(f"General matched section: {match.get('status', 'NOT CHECKED')} - {match.get('message', 'No match')}")
            lines.extend([
                f"Local edge rows: first {purlin_check['edge_rows']} rows within a={purlin_check.get('local_dimension_a_m', 0.0):.2f} m | tributary area {purlin_check['local_area_m2']:.2f} m2 | Kl {purlin_check.get('local_pressure_factor', 1.0):.1f}",
                f"Local line loads: outward ULS {purlin_check['local_outward_kn_m']:.3f} kN/m, inward ULS {purlin_check['local_inward_strength_kn_m']:.3f} kN/m",
                f"Local wind up/down {purlin_check['local_wind_upward_kpa']:.3f}/{purlin_check['local_wind_downward_kpa']:.3f} kPa",
                f"Local inward ULS basis: max(1.2G+1.5Q = {purlin_check['inward_gravity_uls_kpa']:.3f} kPa, 1.2G+Wd = {purlin_check['local_inward_wind_uls_kpa']:.3f} kPa)",
            ])
            if local_match.get("status") == "OK":
                lines.extend([
                    f"Local matched section: {local_match['section']} from Metroll tables {local_match['tables']}",
                    f"Local capacities: outward {local_match['outward_capacity_kn_m']:.3f} kN/m, inward strength {local_match['inward_strength_capacity_kn_m']:.3f} kN/m, inward L/150 {local_match['inward_service_capacity_kn_m']:.3f} kN/m",
                ])
                if local_match.get("capped_by_double_span"):
                    lines.append("Local single-span capacities capped at corresponding double-span values where extracted table values were higher.")
            else:
                lines.append(f"Local matched section: {local_match.get('status', 'NOT CHECKED')} - {local_match.get('message', 'No match')}")
            lines.append("Note: Metroll table capacities are limit-state kN/m values and do not include member self-weight.")
        wall_girt_check = self.wall_girt_check or {}
        if wall_girt_check:
            match = wall_girt_check.get("match", {})
            lines.append("")
            lines.extend([
                "WALL GIRT CHECK",
                f"Span type: {wall_girt_check['span_type']} | table span {match.get('span_row_mm', 0):.0f} mm{' interpolated' if match.get('interpolated') else ''} | spacing {wall_girt_check['spacing_mm']:.0f} mm (max {wall_girt_check['max_spacing_mm']:.0f} mm)",
                f"Line loads: outward ULS {wall_girt_check['outward_kn_m']:.3f} kN/m, inward ULS {wall_girt_check['inward_strength_kn_m']:.3f} kN/m, SLS {wall_girt_check['inward_service_kn_m']:.3f} kN/m",
                f"Wall pressures: outward {wall_girt_check['wind_outward_kpa']:.3f} kPa, inward {wall_girt_check['wind_inward_kpa']:.3f} kPa, service {wall_girt_check['wind_service_kpa']:.3f} kPa",
            ])
            if match.get("status") == "OK":
                lines.extend([
                    f"Matched section: {match['section']} from Metroll tables {match['tables']}",
                    f"Selection basis: deepest section within 5% of the lightest passing area; selected area {match.get('area_mm2', 0.0):.0f} mm2.",
                    f"Capacities: outward {match['outward_capacity_kn_m']:.3f} kN/m, inward strength {match['inward_strength_capacity_kn_m']:.3f} kN/m, inward L/150 {match['inward_service_capacity_kn_m']:.3f} kN/m",
                ])
                if match.get("capped_by_double_span"):
                    lines.append("Single-span capacities capped at corresponding double-span values where extracted table values were higher.")
                if match.get("near_misses"):
                    lines.append("Nearest failing sections:")
                    for miss in match["near_misses"]:
                        lines.append(
                            f"- {miss['section']}: outward {miss['outward_ratio']:.3f}, "
                            f"inward strength {miss['inward_strength_ratio']:.3f}, "
                            f"inward L/150 {miss['inward_service_ratio']:.3f}"
                        )
            else:
                lines.append(f"Matched section: {match.get('status', 'NOT CHECKED')} - {match.get('message', 'No match')}")
            lines.append("Note: wall girt spacing is capped at 1700 mm, or 1100 mm for wind Region C.")
        if not is_rafter and envelopes["WEB"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("WEB / POST", web_profile, envelopes["WEB"], "WEB"))
        if left_canopy_top_profile is not None and envelopes["LEFT_CANOPY_TOP"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("LEFT CANOPY / LEAN-TO TOP CHORD", left_canopy_top_profile, envelopes["LEFT_CANOPY_TOP"], "LEFT_CANOPY_TOP"))
        if left_canopy_bottom_profile is not None and envelopes["LEFT_CANOPY_BOTTOM"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("LEFT CANOPY / LEAN-TO BOTTOM CHORD", left_canopy_bottom_profile, envelopes["LEFT_CANOPY_BOTTOM"], "LEFT_CANOPY_BOTTOM"))
        if left_canopy_web_profile is not None and envelopes["LEFT_CANOPY_WEB"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("LEFT CANOPY / LEAN-TO WEB/POST", left_canopy_web_profile, envelopes["LEFT_CANOPY_WEB"], "LEFT_CANOPY_WEB"))
        if left_lean_column_profile is not None and envelopes["LEFT_LEAN_COLUMN"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("LEFT LEAN-TO OUTER COLUMN", left_lean_column_profile, envelopes["LEFT_LEAN_COLUMN"], "LEFT_LEAN_COLUMN"))
        if right_canopy_top_profile is not None and envelopes["RIGHT_CANOPY_TOP"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("RIGHT CANOPY / LEAN-TO TOP CHORD", right_canopy_top_profile, envelopes["RIGHT_CANOPY_TOP"], "RIGHT_CANOPY_TOP"))
        if right_canopy_bottom_profile is not None and envelopes["RIGHT_CANOPY_BOTTOM"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("RIGHT CANOPY / LEAN-TO BOTTOM CHORD", right_canopy_bottom_profile, envelopes["RIGHT_CANOPY_BOTTOM"], "RIGHT_CANOPY_BOTTOM"))
        if right_canopy_web_profile is not None and envelopes["RIGHT_CANOPY_WEB"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("RIGHT CANOPY / LEAN-TO WEB/POST", right_canopy_web_profile, envelopes["RIGHT_CANOPY_WEB"], "RIGHT_CANOPY_WEB"))
        if right_lean_column_profile is not None and envelopes["RIGHT_LEAN_COLUMN"]["compression_kn"]["value"] >= 0.0:
            lines.append("")
            lines.extend(format_check_set("RIGHT LEAN-TO OUTER COLUMN", right_lean_column_profile, envelopes["RIGHT_LEAN_COLUMN"], "RIGHT_LEAN_COLUMN"))

        self.section_checks_box.setText("\n".join(lines))

    def update_costing_text(self):
        if not getattr(self, "costing_box", None):
            return
        summary = self.costing_summary
        if not summary:
            self.costing_box.setText("Generate and solve the frame to view costing.")
            if getattr(self, "total_cost_label", None):
                self.total_cost_label.setText("Not calculated")
                self.total_cost_label.setStyleSheet("font-weight: bold;")
            return
        lines = [
            "COSTING SUMMARY",
            f"Building length: {summary['building_length_m']:.2f} m",
            f"Bay size: {summary['bay_size_m']:.2f} m",
            f"Bays / portals: {summary['bay_count']} / {summary['portal_count']}",
            "",
            "FRAME STEEL",
            f"Per portal SHS/RHS/CHS exact: {summary['per_portal_base_weights']['shs']:.1f} kg + 7% waste = {summary['per_portal_weights']['shs']:.1f} kg",
            f"Per portal UB/UC/PFC/open exact: {summary['per_portal_base_weights']['ub']:.1f} kg; cut length {summary['ub_cut_length_m_per_portal']:.2f} m rounded to stock length {summary['ub_stock_length_m_per_portal']:.2f} m = {summary['per_portal_weights']['ub']:.1f} kg",
            f"Per portal WB exact: {summary['per_portal_base_weights']['wb']:.1f} kg; cut length {summary['wb_cut_length_m_per_portal']:.2f} m rounded to stock length {summary['wb_stock_length_m_per_portal']:.2f} m = {summary['per_portal_weights']['wb']:.1f} kg",
            f"Total SHS/RHS/CHS incl. waste: {summary['frame_total_kg']['shs']:.1f} kg x ${summary['rates']['shs']:.2f}/kg = ${summary['costs']['shs']:,.0f}",
            f"Total UB/UC/PFC/open stock: {summary['frame_total_kg']['ub']:.1f} kg x ${summary['rates']['ub']:.2f}/kg = ${summary['costs']['ub']:,.0f}",
            f"Total WB stock: {summary['frame_total_kg']['wb']:.1f} kg x ${summary['rates']['wb']:.2f}/kg = ${summary['costs']['wb']:,.0f}",
            "",
            "PURLINS / GIRTS",
        ]
        if summary["purlin_items"]:
            for item in summary["purlin_items"]:
                line = (
                    f"{item['label']}: {item['count']} x {item['length_m']:.2f} m {item['section']} "
                    f"= {item['total_m']:.1f} m x {item['kg_m']:.2f} kg/m"
                )
                if item.get("lap_factor", 1.0) > 1.0:
                    line += f" x 1.15 lap = {item['kg']:.1f} kg"
                else:
                    line += f" = {item['kg']:.1f} kg"
                lines.append(line)
            lines.append(
                f"Total purlins/girts: {summary['purlin_total_kg']:.1f} kg x "
                f"${summary['rates']['purlin']:.2f}/kg = ${summary['costs']['purlin']:,.0f}"
            )
        else:
            lines.append("No purlin/girt match available.")
        lines.extend([
            "",
            f"TOTAL ESTIMATED STEEL COST: ${summary['total_cost']:,.0f}",
            "",
            "Rates are supply placeholders only; no allowance is made for cleats, bolts, plates, fabrication, coating, transport, waste, labour, GST, or mark-up.",
        ])
        self.costing_box.setText("\n".join(lines))
        if getattr(self, "total_cost_label", None):
            self.total_cost_label.setText(f"${summary['total_cost']:,.0f}")
            self.total_cost_label.setStyleSheet("font-weight: bold; color: #1b7f3a;")

    def update_foundation_text(self):
        if not getattr(self, "foundation_box", None):
            return
        summary = self.foundation_summary
        if not summary:
            self.foundation_box.setText("Generate and solve the frame to view foundation design.")
            return
        lines = [
            "FOUNDATION DESIGN",
            f"Allowable bearing: {summary['bearing_kpa']:.1f} kPa",
            f"Skin friction: {summary['skin_friction_kpa']:.1f} kPa from 1.0 m depth",
            f"Skin friction used for bearing: {'Yes' if summary.get('use_skin_friction_for_bearing') else 'No'}",
            "Standard pier diameters: " + ", ".join(f"{diameter:.0f} mm" for diameter in summary["diameters_mm"]),
            "Pier selection increases diameter before exceeding 4.0 m depth.",
            "",
            "SELECTED FOOTINGS",
        ]
        for item in summary["items"]:
            lines.append(
                f"{item['label']}: {item['diameter_mm']:.0f} mm dia x {item['depth_m']:.2f} m deep ({item.get('status', '-')})"
            )
        lines.extend([
            "",
            "DETAILS",
            "Column              | Demand envelope                         | Pier design",
            "--------------------|-----------------------------------------|-------------------------------",
        ])
        for item in summary["items"]:
            status = item.get("status", "-")
            extra = ""
            if item.get("required_bearing_diameter_mm", 0.0) > item["diameter_mm"]:
                extra = f" | bearing needs {item['required_bearing_diameter_mm']:.0f} mm dia"
            lines.append(
                f"{item['label']:<20}| "
                f"SLS C {item['compression_kn']:>7.1f} kN ({item['compression_combo']}), "
                f"ULS U {item['uplift_kn']:>7.1f} kN ({item['uplift_combo']}) | "
                f"{item['diameter_mm']:.0f} dia x {item['depth_m']:.2f} m deep | "
                f"bearing {item['bearing_capacity_kn']:.1f} kN, skin {item['skin_capacity_kn']:.1f} kN | {status}{extra}"
            )
            lines.append(
                f"{'':<20}| "
                f"H {item['shear_kn']:>7.1f} kN ({item['shear_combo']}), "
                f"M {item['moment_knm']:>7.1f} kNm ({item['moment_combo']}) | "
                "horizontal/moment shown for reference"
            )
        lines.extend([
            "",
            "Note: this is a preliminary circular pier sizing using vertical bearing and uplift skin friction only. It does not check lateral soil pressure, pier bending/shear, reinforcement, settlement, edge distance, group effects, or geotechnical reduction factors.",
        ])
        self.foundation_box.setText("\n".join(lines))

    def update_results_text(self, num_panels, panel, web_angle, left_support, right_support,
                            left_base, right_base, left_column_profile, right_column_profile, top_profile, bottom_profile, web_profile,
                            load_summary_lines, design_check_lines,
                            outside_span=None, centreline_span=None, left_column_offset=0.0, right_column_offset=0.0,
                            left_canopy_length=0.0, right_canopy_length=0.0,
                            left_canopy_pitch=0.0, left_canopy_shallow_depth=0.0,
                            right_canopy_pitch=0.0, right_canopy_shallow_depth=0.0,
                            left_canopy_type="Canopy", right_canopy_type="Canopy",
                            left_canopy_eave_height=None, right_canopy_eave_height=None,
                            eave_x_restraint="None", eave_x_spring_kn_mm=0.0,
                            eave_bracing_info=None,
                            rafter_haunch_length=0.0):
        d = self.displacements
        R = self.reactions
        structure = self.structure

        results = []
        results.append("=== ANALYSIS RESULTS ===\n")
        if outside_span is not None and centreline_span is not None:
            results.append(f"Outside width: {outside_span:.1f} mm")
            results.append(f"Column centreline span: {centreline_span:.1f} mm")
            results.append(f"Column offsets: left {left_column_offset:.1f} mm, right {right_column_offset:.1f} mm")
        if left_canopy_length > 0.0 or right_canopy_length > 0.0:
            left_height_text = "top column node" if left_canopy_eave_height is None else f"{left_canopy_eave_height:.1f} mm"
            right_height_text = "top column node" if right_canopy_eave_height is None else f"{right_canopy_eave_height:.1f} mm"
            results.append(
                f"Canopies: left {left_canopy_length:.1f} mm ({left_canopy_type}, eave {left_height_text}, pitch {left_canopy_pitch:.2f} deg, depth {left_canopy_shallow_depth:.1f} mm), "
                f"right {right_canopy_length:.1f} mm ({right_canopy_type}, eave {right_height_text}, pitch {right_canopy_pitch:.2f} deg, depth {right_canopy_shallow_depth:.1f} mm)"
            )
        results.append(f"Panels: {num_panels}")
        results.append(f"Panel length: {panel:.1f} mm")
        results.append(f"Approx web angle: {web_angle:.1f}°")
        if self.purlin_layout is not None:
            results.append(
                f"Purlins: max spacing {self.purlin_layout.max_spacing_mm:.0f} mm "
                f"(end {self.purlin_layout.end_spacing_limit_mm:.0f}, mid limit {self.purlin_layout.mid_spacing_limit_mm:.0f})"
            )
        results.append("Bottom chord connection: directly connected to split column nodes")
        if rafter_haunch_length > 0.0:
            results.append(
                f"Rafter haunch: tapered added depth from full rafter depth at columns to zero over {rafter_haunch_length:.0f} mm each side (stiffness approximation)"
            )
        results.append(f"Left support: {left_support}")
        results.append(f"Right support: {right_support}")
        eave_restraint_text = eave_x_restraint
        if eave_x_restraint != "None" and eave_x_spring_kn_mm > 0.0:
            eave_restraint_text += f" spring {eave_x_spring_kn_mm:.3f} kN/mm"
        elif eave_x_restraint != "None":
            eave_restraint_text += " rigid"
        results.append(f"Eave X restraint: {eave_restraint_text}\n")
        if eave_bracing_info is not None:
            results.append(
                "End wall bracing approximation: "
                f"{eave_bracing_info['squares_to_end']} x 45-degree bay squares to nearest end wall, "
                f"{eave_bracing_info['brace_diameter_mm']:.1f} mm dia brace "
                f"(A={eave_bracing_info['brace_area_mm2']:.0f} mm2, L={eave_bracing_info['diagonal_length_mm']:.0f} mm), "
                f"one-square spring {eave_bracing_info['base_spring_kn_mm']:.3f} kN/mm, "
                f"both end walls in parallel -> {eave_bracing_info['effective_spring_kn_mm']:.3f} kN/mm per eave\n"
            )

        analysis_info = self.analysis_info or {"type": "first_order", "converged": True, "iterations": 1, "displacement_amplification": 1.0}
        if analysis_info.get("type") == "second_order":
            status = "converged" if analysis_info.get("converged") else "not converged"
            results.append("ANALYSIS METHOD")
            results.append(
                f"Second-order geometric stiffness iteration: {status} in "
                f"{analysis_info.get('iterations', 0)} / {analysis_info.get('max_iterations', 0)} iterations"
            )
            results.append(
                f"Max displacement amplification: {analysis_info.get('displacement_amplification', 1.0):.3f} "
                f"({analysis_info.get('first_order_max_abs_mm', 0.0):.3f} mm -> "
                f"{analysis_info.get('second_order_max_abs_mm', 0.0):.3f} mm)\n"
            )
        else:
            results.append("ANALYSIS METHOD")
            results.append("First-order linear elastic analysis\n")

        results.extend(load_summary_lines)
        results.append("")

        results.append("SELECTED PROFILES")
        for label, p in [
            ("Left column", left_column_profile),
            ("Right column", right_column_profile),
            ("Top chord", top_profile),
            ("Bottom chord", bottom_profile),
            ("Webs/posts", web_profile),
        ]:
            results.append(f"{label:<13}: {p.name} | A={p.A:.0f} mm² | I={p.I_analysis:.3e} mm⁴")
        results.append("")

        results.append("SUPPORT REACTIONS")
        results.append(f"Left RX  = {R[left_base.ux] / 1000:.2f} kN")
        results.append(f"Left RY  = {R[left_base.uy] / 1000:.2f} kN")
        results.append(f"Left MZ  = {R[left_base.rz] / 1_000_000:.2f} kNm")
        results.append(f"Right RX = {R[right_base.ux] / 1000:.2f} kN")
        results.append(f"Right RY = {R[right_base.uy] / 1000:.2f} kN")
        results.append(f"Right MZ = {R[right_base.rz] / 1_000_000:.2f} kNm\n")

        min_uy_node = min(structure.nodes, key=lambda n: d[n.uy])
        max_uy_node = max(structure.nodes, key=lambda n: d[n.uy])
        min_ux_node = min(structure.nodes, key=lambda n: d[n.ux])
        max_ux_node = max(structure.nodes, key=lambda n: d[n.ux])
        max_abs_ux_node = max(structure.nodes, key=lambda n: abs(d[n.ux]))
        results.append("MAX DISPLACEMENTS")
        results.append(f"Max upward UY    = {d[max_uy_node.uy]:+.3f} mm at Node {max_uy_node.id}")
        results.append(f"Max downward UY  = {d[min_uy_node.uy]:+.3f} mm at Node {min_uy_node.id}")
        results.append(f"Max +UX          = {d[max_ux_node.ux]:+.3f} mm at Node {max_ux_node.id}")
        results.append(f"Max -UX          = {d[min_ux_node.ux]:+.3f} mm at Node {min_ux_node.id}")
        results.append(f"Max abs UX       = {d[max_abs_ux_node.ux]:+.3f} mm at Node {max_abs_ux_node.id}\n")

        results.extend(design_check_lines)
        results.append("")

        results.append("MEMBER MAX VALUES")
        results.append("Member | Group  | Profile            | Compression kN | Shear kN | Moment kNm")
        for element in structure.elements:
            s = element.force_summary(d)
            results.append(
                f"{element.id:>6} | {element.group:<6} | {element.profile.name[:18]:<18} | "
                f"{s['compression_kn']:>14.2f} | {s['shear_kn']:>8.2f} | {s['moment_knm']:>10.2f}"
            )

        self.results_box.setText("\n".join(results))

    def refresh_plot(self):
        if self.structure is None:
            return
        try:
            deflection_scale = float(self.deflection_scale_input.text())
        except (ValueError, AttributeError):
            deflection_scale = 100.0
        try:
            diagram_scale_factor = float(self.diagram_scale_input.text())
        except (ValueError, AttributeError):
            diagram_scale_factor = 1.0

        deflection_scale = max(deflection_scale, 0.0)
        diagram_scale_factor = max(diagram_scale_factor, 0.0)
        result_type = self.result_combo.currentText()
        load_arrows = self.load_arrows
        if (
            result_type == "Load Diagram"
            and getattr(self, "load_view_mode_combo", None) is not None
            and self.load_view_mode_combo.currentText() == "Single load"
        ):
            load_arrows = (self.single_load_arrows or {}).get(self.single_load_combo.currentText(), [])

        self.canvas.plot_structure(
            self.structure.nodes,
            self.structure.elements,
            self.base_nodes,
            self.support_types,
            load_arrows,
            result_type,
            self.displacements,
            scale=deflection_scale,
            diagram_scale_factor=diagram_scale_factor,
            purlin_layout=self.purlin_layout,
            wall_cladding_segments=self.wall_cladding_segments,
        )


# ==========================================================
# APPLICATION
# ==========================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    try:
        window = MainWindow()
        window.show()

        def enforce_license_after_startup():
            from licensing import license_enforced

            if not license_enforced():
                return
            try:
                from licensing import require_valid_license
                licensed = require_valid_license(window)
            except Exception as exc:
                QMessageBox.critical(
                    window,
                    "License check failed",
                    str(exc),
                )
                licensed = False
            if not licensed:
                QMessageBox.warning(
                    window,
                    "License required",
                    "PortalCalc needs a valid license to continue.",
                )
                window.close()
                app.quit()
            else:
                window.update_license_status_label()

        QTimer.singleShot(100, enforce_license_after_startup)
        sys.exit(app.exec())
    except Exception as exc:
        QMessageBox.critical(None, "Startup Error", str(exc))
        raise
